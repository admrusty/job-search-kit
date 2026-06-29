#!/usr/bin/env python3
"""
Job Scout Browser — a local, custom interface for "Job Scout Jobs.xlsx".

Runs entirely on your machine. Reads jobs from data.json (primary) or the
workbook (fallback). All write operations go to job-scout-state.json.
After state changes that affect the workbook, triggers a background rebuild
via jobscout_build.py.

Dependencies: openpyxl (already installed for the scheduled task). No internet,
no other packages. Python 3.8+.

Run:  python3 job_scout_browser.py
Then open the URL it prints (it also tries to open your browser automatically).
"""

import hashlib
import json
import os
import re
import subprocess
import sys
import shutil
import threading
import webbrowser
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import quote_plus

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install with:  pip3 install --user openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
SCOUT_DIR = HERE
# Reuse the scraper's deterministic classifier (salary/workplace/keywords/geoFit).
sys.path.insert(0, SCOUT_DIR)
try:
    from jobscout_core import load_state, save_state, merge_state
    import jobscout_core as core
except Exception:
    load_state = save_state = merge_state = None
    core = None
XLSX = os.path.join(HERE, "Job Scout Jobs.xlsx")
BACKUP = os.path.join(HERE, "Job Scout Jobs.bak.xlsx")
DATA_JSON = os.path.join(SCOUT_DIR, "data.json")
STATE_PATH = os.path.join(SCOUT_DIR, "job-scout-state.json")
# Legacy files — kept for fallback reads; new writes go to STATE_PATH.
OVERRIDES = os.path.join(SCOUT_DIR, "job-scout-overrides.json")
STARS = os.path.join(SCOUT_DIR, "job-scout-stars.json")
STATUS_FILE = os.path.join(SCOUT_DIR, "job-scout-status.json")
# Manual rank order (list of JobIDs, highest first) for the "Manual" sort.
ORDER_FILE = os.path.join(SCOUT_DIR, "job-scout-order.json")
CONFIG_FILE = os.path.join(SCOUT_DIR, "config", "jobscout_config.json")
SHEET = "Jobs"
PORT = 8733

def _load_third_party_set():
    try:
        with open(CONFIG_FILE) as f:
            cfg = json.load(f)
        return {s.lower() for s in cfg.get("thirdPartyPosters", [])}
    except Exception:
        return set()

THIRD_PARTY = _load_third_party_set()


def _job_export_dir() -> str:
    """Where the dashboard's '⤓ save' button writes a job's details as Markdown.

    Saves land in this project's own `inputs/` folder, so running `/draft`
    picks them up directly — the job finder and the resume drafter share one
    workspace.
    """
    return os.path.join(SCOUT_DIR, "inputs")


# ---------------------------------------------------------------------------
# Resume / cover-letter PDF lookup (for Application Tracker)
# ---------------------------------------------------------------------------

def _salary_norm_annual(raw: str) -> int:
    """Parse a raw salary string and return a normalized annual integer.

    Rules:
    - Extracts all dollar amounts (handles K-suffix and comma-formatted values).
    - Uses the maximum value found (top of range).
    - Detects hourly rates (value ≤ 200 or explicit /hr or /hour marker) and
      multiplies by 2080 (40 hrs × 52 wks).
    - Returns 0 if no dollar amount can be parsed.
    """
    if not raw:
        return 0
    import re as _re
    nums = []
    for m in _re.finditer(r'\$\s*([\d,]+(?:\.\d+)?)\s*[Kk]?', raw):
        token = m.group(0)
        val = float(_re.sub(r'[$,\s]', '', token.rstrip('Kk')))
        if _re.search(r'[Kk]', token):
            val *= 1000
        nums.append(val)
    if not nums:
        return 0
    val = max(nums)
    hourly = bool(_re.search(r'\bhr\b|\bhour\b|per hour|/hr', raw, _re.IGNORECASE)) or val <= 200
    if hourly:
        val = val * 2080
    return round(val)


_SOURCE_DOMAIN_MAP = [
    ("linkedin.com",           "linkedin"),
    ("indeed.com",             "indeed"),
    ("glassdoor.com",          "glassdoor"),
    ("greenhouse.io",          "greenhouse"),
    ("boards.greenhouse.io",   "greenhouse"),
    ("jobs.ashbyhq.com",       "ashby"),
    ("ashbyhq.com",            "ashby"),
    ("lever.co",               "lever"),
    ("jobs.lever.co",          "lever"),
    ("myworkdayjobs.com",      "workday"),
    ("paylocity.com",          "employer"),
]

def _infer_source(link: str) -> str:
    """Infer canonical source name from a job URL. Returns '' for manual/unknown."""
    if not link:
        return ""
    ll = link.lower()
    for fragment, source in _SOURCE_DOMAIN_MAP:
        if fragment in ll:
            return source
    if any(kw in ll for kw in ("/careers/", "/jobs/", "/job/", "/openings/")):
        return "employer"
    return ""


def _norm_company(s: str) -> str:
    """Normalize a company name for fuzzy folder matching."""
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    for w in ("inc", "llc", "ltd", "corp", "technologies", "technology",
              "group", "services", "solutions", "systems", "the", "and"):
        s = re.sub(r"\b" + w + r"\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# ── Folder configuration ──────────────────────────────────────────────────────
# Customize these if you rename or reorganize your Google Drive folders.
# RESUME_FOLDER_NAME:   exact name of the folder that contains per-company subfolders.
# JOB_SEARCH_PARENT:   substring matched (case-insensitive) against a My Drive subfolder
#                       that contains RESUME_FOLDER_NAME.  Leave blank to search all
#                       My Drive subfolders at depth 1 and 2.
RESUME_FOLDER_NAME  = "Job Materials"
JOB_SEARCH_PARENT   = "Job Search"    # matches "2026 Job Search", "2027 Job Search", etc.
# ──────────────────────────────────────────────────────────────────────────────


def _find_resume_folder() -> str:
    """Locate RESUME_FOLDER_NAME under Google Drive (any account).

    Search strategy (stops at first match):
      1. Any My Drive subfolder whose name contains JOB_SEARCH_PARENT (case-insensitive),
         then look for RESUME_FOLDER_NAME inside it.
      2. RESUME_FOLDER_NAME directly inside My Drive (no intermediate folder).
    """
    gdrive_root = os.path.expanduser("~/Library/CloudStorage")
    if not os.path.isdir(gdrive_root):
        return ""
    try:
        entries = os.listdir(gdrive_root)
    except OSError:
        return ""
    for entry in entries:
        if not entry.startswith("GoogleDrive-"):
            continue
        my_drive = os.path.join(gdrive_root, entry, "My Drive")
        if not os.path.isdir(my_drive):
            continue
        try:
            subdirs = os.listdir(my_drive)
        except OSError:
            continue
        # Pass 1: look inside any subfolder matching JOB_SEARCH_PARENT
        if JOB_SEARCH_PARENT:
            for sub in subdirs:
                if JOB_SEARCH_PARENT.lower() not in sub.lower():
                    continue
                candidate = os.path.join(my_drive, sub, RESUME_FOLDER_NAME)
                if os.path.isdir(candidate):
                    return candidate
        # Pass 2: look for RESUME_FOLDER_NAME directly in My Drive
        candidate = os.path.join(my_drive, RESUME_FOLDER_NAME)
        if os.path.isdir(candidate):
            return candidate
    return ""


RESUME_FOLDER: str = _find_resume_folder()
_pdf_cache: dict = {}   # norm_company → {"resume": path|None, "coverLetter": path|None}


def _get_company_pdfs(company: str, role: str = "") -> dict:
    """Return {"resume": path_or_None, "coverLetter": path_or_None} for a company+role."""
    norm = _norm_company(company)
    role_norm = _norm_company(role)
    cache_key = f"{norm}|{role_norm}"
    if cache_key in _pdf_cache:
        return _pdf_cache[cache_key]
    result = {"resume": None, "coverLetter": None, "folder": None}
    if not RESUME_FOLDER:
        _pdf_cache[cache_key] = result
        return result
    # Find the best-matching company subfolder
    best_folder, best_score = "", 0.0
    try:
        for d in os.listdir(RESUME_FOLDER):
            if d.startswith((".", "ARCHIVE")):
                continue
            full = os.path.join(RESUME_FOLDER, d)
            if not os.path.isdir(full):
                continue
            d_norm = _norm_company(d)
            if not d_norm or not norm:
                continue
            # Folders follow "Company Name - Role Title" or "Company (context) - Role".
            # Extract company portion (before the first " - " or "(") and role portion (after " - ").
            parts = re.split(r'\s[-–]\s', d, maxsplit=1)
            company_part = re.split(r'\s*\(', parts[0])[0].strip()
            folder_role_part = parts[1].strip() if len(parts) > 1 else ""
            cp_norm = _norm_company(company_part)
            folder_role_norm = _norm_company(folder_role_part)

            def _prefix_match(a, b):
                """True if a starts with b and next char is a word boundary."""
                return bool(b) and a.startswith(b) and (len(a) == len(b) or a[len(b)] in (' ', '-'))

            # Score company match
            if norm == d_norm:
                co_score = 1.0
            elif cp_norm == norm or _prefix_match(norm, cp_norm) or _prefix_match(cp_norm, norm):
                co_score = 0.9
            elif _prefix_match(d_norm, norm) or _prefix_match(norm, d_norm):
                co_score = 0.85
            else:
                shorter, longer = sorted([norm, d_norm], key=len)
                co_score = (len(shorter) / len(longer)) if shorter and shorter in longer else 0.0

            if co_score == 0.0:
                continue

            # Score role match — only applied when a role is provided and folder has a role portion
            if role_norm and folder_role_norm:
                if role_norm == folder_role_norm:
                    role_score = 1.0
                elif role_norm in folder_role_norm or folder_role_norm in role_norm:
                    shorter, longer = sorted([role_norm, folder_role_norm], key=len)
                    role_score = len(shorter) / len(longer)
                else:
                    # Word overlap ratio
                    r_words = set(role_norm.split())
                    f_words = set(folder_role_norm.split())
                    overlap = len(r_words & f_words)
                    role_score = overlap / max(len(r_words), len(f_words)) if r_words or f_words else 0.0
                score = co_score * 0.5 + role_score * 0.5
            else:
                score = co_score

            if score > best_score:
                best_score, best_folder = score, full
    except OSError:
        pass
    if best_folder and best_score > 0.45:
        try:
            pdfs = [f for f in os.listdir(best_folder) if f.lower().endswith(".pdf")]
        except OSError:
            pdfs = []
        res_cands, cl_cands = [], []
        for f in pdfs:
            fl = f.lower()
            if "cover letter" in fl or "cover_letter" in fl:
                cl_cands.append(f)
            elif "resume" in fl:
                res_cands.append(f)
        def _pick(cands):
            if not cands:
                return None
            # Pick the highest vN version found in any candidate filename
            import re as _re
            def _ver(name):
                m = _re.search(r'v(\d+)', name.lower())
                return int(m.group(1)) if m else -1
            versioned = [(c, _ver(c)) for c in cands if _ver(c) >= 0]
            if versioned:
                return os.path.join(best_folder, max(versioned, key=lambda x: x[1])[0])
            return os.path.join(best_folder, sorted(cands)[-1])
        result = {"resume": _pick(res_cands), "coverLetter": _pick(cl_cands), "folder": best_folder}
    _pdf_cache[norm] = result
    return result


# Load state once at startup; mutated in-process via update_job_state().
scout_state: dict = load_state(STATE_PATH) if load_state else {"jobs": {}}

# Column positions (1-based) in the Jobs sheet — must match jobscout_build.py COLS.
COL = {
    "JobID": 1, "Score": 2, "Title": 3, "Company": 4, "Workplace": 5,
    "Location": 6, "Salary": 7, "Category": 8, "Posted": 9, "Status": 10,
    "AppliedDate": 11, "Dismissed": 12, "Notes": 13, "Keywords": 14,
    "Why": 15, "Link": 16, "Starred": 17, "GeoFit": 18, "Description": 19,
}
STATUSES = ("new", "applied", "drafting", "interviewing", "rejected", "not-applying")
TRUTHY = {"x", "yes", "true", "1", "✓", "y"}

_write_lock = threading.Lock()


# ---------------------------------------------------------------------------
# State helpers (Wave 5)
# ---------------------------------------------------------------------------

def update_job_state(job_id: str, updates: dict, rebuild: bool = True) -> None:
    """Update a job's state entry and persist to job-scout-state.json."""
    entry = scout_state["jobs"].setdefault(str(job_id), {
        "starred": False,
        "userDismissed": False,
        "reviewStatus": "new",
        "manualNotes": "",
        "overrides": {},
        "updatedAt": None,
    })
    # Merge overrides sub-dict rather than replacing the whole thing
    if "overrides" in updates:
        existing_ov = entry.get("overrides") or {}
        existing_ov.update(updates.pop("overrides"))
        entry["overrides"] = existing_ov
    entry.update(updates)
    entry["updatedAt"] = datetime.now().astimezone().isoformat()
    if save_state:
        save_state(STATE_PATH, scout_state)
    if rebuild:
        trigger_rebuild()


def trigger_rebuild() -> None:
    """Run jobscout_build.py in the background to regenerate the workbook."""
    import subprocess
    build_script = os.path.join(SCOUT_DIR, "jobscout_build.py")
    if not os.path.exists(build_script):
        return
    subprocess.Popen(
        ["python3", build_script],
        cwd=SCOUT_DIR,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )


def read_jobs():
    """Read every job row fresh from the workbook."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    jobs = []
    for row in ws.iter_rows(min_row=2):
        jid = row[COL["JobID"] - 1].value
        if jid is None or str(jid).strip() == "":
            continue
        link_cell = row[COL["Link"] - 1]
        link = link_cell.hyperlink.target if link_cell.hyperlink else (link_cell.value or "")
        dval = str(row[COL["Dismissed"] - 1].value or "").strip().lower()
        status = str(row[COL["Status"] - 1].value or "").strip().lower() or "new"
        if status not in STATUSES:
            status = "new"
        kw = row[COL["Keywords"] - 1].value or ""
        jobs.append({
            "jobid": str(jid).strip(),
            "score": row[COL["Score"] - 1].value or 0,
            "title": row[COL["Title"] - 1].value or "",
            "company": row[COL["Company"] - 1].value or "",
            "workplace": row[COL["Workplace"] - 1].value or "",
            "location": row[COL["Location"] - 1].value or "",
            "salary": row[COL["Salary"] - 1].value or "",
            "category": row[COL["Category"] - 1].value or "",
            "posted": str(row[COL["Posted"] - 1].value or ""),
            "status": status,
            "applied": status != "new",
            "dismissed": dval in TRUTHY,
            "notes": row[COL["Notes"] - 1].value or "",
            "keywords": [k.strip() for k in str(kw).split(",") if k.strip()],
            "why": row[COL["Why"] - 1].value or "",
            "link": link,
        })
    wb.close()
    return jobs


def _manual_jobid(company: str, role: str) -> str:
    """Stable synthetic jobid for a manually-added applied.json entry."""
    key = (_norm_company(company) + "|" + _norm_company(role)).encode()
    return "manual_" + hashlib.md5(key).hexdigest()[:10]


def _load_manual_applied_jobs(known_pairs: set) -> list:
    """Return applied.json entries not already in data.json as synthetic job dicts."""
    applied_path = os.path.join(HERE, "job-scout-applied.json")
    try:
        with open(applied_path) as f:
            raw = json.load(f)
        entries = raw.get("appliedRoles", [])
    except Exception:
        return []

    result = []
    for m in entries:
        company = (m.get("company") or "").strip()
        role = (m.get("role") or "").strip()
        if not company or not role:
            continue
        if (_norm_company(company), _norm_company(role)) in known_pairs:
            continue  # already present from data.json

        jobid = _manual_jobid(company, role)
        applied_at = (m.get("appliedAt") or "")[:10]

        # Overlay persisted state (status, starred, dismissed, notes)
        state_entry = scout_state.get("jobs", {}).get(jobid, {})
        status = state_entry.get("reviewStatus") or "applied"
        if status not in STATUSES:
            status = "applied"

        result.append({
            "jobid": jobid,
            "score": 0,
            "title": role,
            "company": company,
            "workplace": "",
            "location": "",
            "salary": "",
            "category": "",
            "posted": "",
            "status": status,
            "applied": True,
            "dismissed": bool(state_entry.get("userDismissed", False)),
            "notes": state_entry.get("manualNotes") or "",
            "keywords": [],
            "why": "",
            "link": "",
            "starred": bool(state_entry.get("starred", False)),
            "thirdParty": False,
            "geofit": None,
            "description": "",
            "caExcluded": False,
            "caEligible": None,
            "compRisk": "",
            "addedAt": applied_at,
            "createdAt": applied_at,
            "updatedAt": state_entry.get("updatedAt") or "",
            "appliedAt": applied_at,
        })
    return result


def read_jobs_from_data() -> list:
    """Read jobs from data.json (primary) and overlay state from job-scout-state.json.

    Falls back to read_jobs() (xlsx) if data.json is missing or unreadable.
    The returned dicts use the legacy lowercase keys expected by the UI.
    """
    if not os.path.exists(DATA_JSON):
        return read_jobs()
    try:
        with open(DATA_JSON) as f:
            data = json.load(f)
    except Exception:
        return read_jobs()

    raw_jobs = data.get("jobs", [])
    if not raw_jobs:
        return read_jobs()

    jobs = []
    for rec in raw_jobs:
        # Apply state overlay (starred, userDismissed, reviewStatus, manualNotes, overrides)
        if merge_state:
            rec = merge_state(rec, scout_state)

        overrides = rec.get("_overrides") or rec.get("overrides") or {}
        workplace = overrides.get("workplace") or rec.get("workplaceNormalized") or rec.get("workplace") or ""
        salary = overrides.get("salary") or rec.get("salaryRaw") or rec.get("salary") or ""
        score_raw = overrides.get("score")
        if score_raw is None:
            score_raw = rec.get("score") or 0
        try:
            score = int(score_raw)
        except (TypeError, ValueError):
            score = 0

        review_status = rec.get("reviewStatus") or "new"
        if review_status not in STATUSES:
            review_status = "new"

        # Map data.json fields → legacy UI keys
        kw = rec.get("matchedKeywords") or rec.get("keywords") or []
        if isinstance(kw, str):
            kw = [k.strip() for k in kw.split(",") if k.strip()]

        link = rec.get("linkedinUrl") or rec.get("applyUrl") or rec.get("link") or ""
        posted = str(rec.get("postedAt") or rec.get("posted") or "")[:10]
        location = rec.get("locationRaw") or rec.get("location") or ""
        company = rec.get("companyName") or rec.get("company") or ""

        jobs.append({
            "jobid": str(rec.get("id") or "").strip(),
            "score": score,
            "title": rec.get("title") or "",
            "company": company,
            "workplace": workplace,
            "location": location,
            "salary": salary,
            "salaryNormAnnual": rec.get("salaryNormAnnual") or _salary_norm_annual(salary),
            "category": rec.get("searchLane") or rec.get("category") or "",
            "posted": posted,
            "status": review_status,
            "applied": review_status not in ("new", "dismissed"),
            "dismissed": bool(rec.get("userDismissed") or rec.get("dismissed")),
            "notes": rec.get("manualNotes") or rec.get("notes") or "",
            "keywords": kw if isinstance(kw, list) else [],
            "why": rec.get("scoreReason") or rec.get("why") or "",
            "source": rec.get("source") or ("manual" if (rec.get("src") or "").lower() == "manual" else _infer_source(link)),
            "link": link,
            "starred": bool(rec.get("starred")),
            "thirdParty": company.lower() in THIRD_PARTY,
            "geofit": rec.get("geoFit"),
            "description": rec.get("description") or "",
            "caExcluded": rec.get("caExcluded") or False,
            "caEligible": rec.get("caEligible"),
            "compRisk": rec.get("compRisk") or "",
            "addedAt": (rec.get("addedAt") or "")[:10],
            "createdAt": rec.get("createdAt") or "",
            "updatedAt": rec.get("_stateUpdatedAt") or rec.get("updatedAt") or "",
            "appliedAt": rec.get("appliedAt") or "",
            "targetEmployer": bool(rec.get("targetEmployer")),
            "targetEmployerName": rec.get("targetEmployerName") or "",
            "incentives": rec.get("incentives") or [],
        })
    result = [j for j in jobs if j["jobid"]]
    # Deduplicate by normalized company+title — keeps first (newest) when duplicates exist
    seen_pairs: set = set()
    deduped = []
    for j in result:
        pair = (_norm_company(j["company"]), _norm_company(j["title"]))
        if pair not in seen_pairs:
            seen_pairs.add(pair)
            deduped.append(j)
    result = deduped
    # Append manually-added applied.json entries not already scraped by Job Scout
    known_pairs = seen_pairs
    result += _load_manual_applied_jobs(known_pairs)
    return result


def export_csv(jobids):
    """Build a CSV (with full descriptions) for the given JobIDs, in the given order."""
    import csv, io
    order = [str(x).strip() for x in (jobids or []) if str(x).strip()]
    want = set(order)
    stars = load_stars()
    by_id = {}
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    for row in ws.iter_rows(min_row=2):
        jv = row[COL["JobID"] - 1].value
        if jv is None or str(jv).strip() == "":
            continue
        jid = str(jv).strip()
        if want and jid not in want:
            continue
        lc = row[COL["Link"] - 1]
        link = lc.hyperlink.target if lc.hyperlink else (lc.value or "")
        starred = (jid in stars) or bool(str(row[COL["Starred"] - 1].value or "").strip())
        by_id[jid] = [
            row[COL["Score"] - 1].value if row[COL["Score"] - 1].value is not None else "",
            row[COL["Title"] - 1].value or "",
            row[COL["Company"] - 1].value or "",
            row[COL["Workplace"] - 1].value or "",
            row[COL["Location"] - 1].value or "",
            row[COL["Salary"] - 1].value or "",
            row[COL["Category"] - 1].value or "",
            str(row[COL["Posted"] - 1].value or ""),
            row[COL["Status"] - 1].value or "new",
            row[COL["GeoFit"] - 1].value or "",
            "x" if starred else "",
            row[COL["Why"] - 1].value or "",
            row[COL["Notes"] - 1].value or "",
            row[COL["Keywords"] - 1].value or "",
            link,
            row[COL["Description"] - 1].value or "",
        ]
    wb.close()
    header = ["Score", "Title", "Company", "Workplace", "Location", "Salary", "Category",
              "Posted", "Status", "Geo fit", "Starred", "Why", "Notes", "Keywords", "Link", "Description"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    for jid in (order or list(by_id.keys())):
        if jid in by_id:
            w.writerow(by_id[jid])
    return "﻿" + buf.getvalue()


def export_single_job_csv(job: dict) -> tuple:
    """Export all fields for a single job to a CSV in ~/Downloads/.

    Returns (ok: bool, path_or_error: str).
    """
    import csv, io, re as _re
    jid = job.get("jobid") or job.get("id") or "unknown"
    company_slug = _re.sub(r"[^a-z0-9]+", "-", (job.get("company") or "").lower()).strip("-")[:30]
    title_slug = _re.sub(r"[^a-z0-9]+", "-", (job.get("title") or "").lower()).strip("-")[:30]
    fname = f"job-scout-export-{company_slug}-{title_slug}-{jid}.csv"
    downloads = os.path.expanduser("~/Downloads/")
    os.makedirs(downloads, exist_ok=True)
    dest = os.path.join(downloads, fname)

    state_entry = scout_state.get("jobs", {}).get(str(jid), {})
    overrides = state_entry.get("overrides") or {}

    fields = [
        ("id", jid),
        ("title", job.get("title") or ""),
        ("companyName", job.get("company") or ""),
        ("location", job.get("location") or ""),
        ("workplaceNormalized", job.get("workplace") or ""),
        ("score", job.get("score") or ""),
        ("reviewStatus", job.get("status") or ""),
        ("searchLane", job.get("category") or ""),
        ("scoreReason", job.get("why") or ""),
        ("salary", job.get("salary") or ""),
        ("compRisk", job.get("compRisk") or ""),
        ("geoFit", job.get("geofit") if job.get("geofit") is not None else ""),
        ("caEligible", job.get("caEligible") if job.get("caEligible") is not None else ""),
        ("caExcluded", job.get("caExcluded") or False),
        ("commuteEstimateMinutes", ""),
        ("description", job.get("description") or ""),
        ("url", job.get("link") or ""),
        ("appliedAt", job.get("appliedAt") or ""),
        ("starred", job.get("starred") or False),
        ("userDismissed", job.get("dismissed") or False),
        ("manualNotes", job.get("notes") or ""),
        ("overrides_score", overrides.get("score") or ""),
        ("overrides_salary", overrides.get("salary") or ""),
        ("overrides_workplace", overrides.get("workplace") or ""),
        ("createdAt", job.get("createdAt") or ""),
        ("updatedAt", job.get("updatedAt") or ""),
    ]
    try:
        with open(dest, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow([f[0] for f in fields])
            w.writerow([f[1] for f in fields])
        return True, dest
    except Exception as exc:
        return False, str(exc)


def read_last_built():
    try:
        wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
        if "About" in wb.sheetnames:
            for r in wb["About"].iter_rows(values_only=True):
                if r and r[0] and str(r[0]).startswith("Last built"):
                    wb.close()
                    return str(r[0])
        wb.close()
    except Exception:
        pass
    return ""


def _set_override(jobid, key, value):
    """Persist a manual salary/workplace override so the rebuild keeps it."""
    ov = {}
    if os.path.exists(OVERRIDES):
        try:
            with open(OVERRIDES) as f:
                ov = json.load(f)
        except Exception:
            ov = {}
    rec = ov.get(str(jobid), {})
    if value:
        rec[key] = value
    else:
        rec.pop(key, None)
    if rec:
        ov[str(jobid)] = rec
    else:
        ov.pop(str(jobid), None)
    os.makedirs(os.path.dirname(OVERRIDES), exist_ok=True)
    with open(OVERRIDES, "w") as f:
        json.dump(ov, f, indent=1)


def load_stars():
    if os.path.exists(STARS):
        try:
            with open(STARS) as f:
                return set(str(x) for x in json.load(f))
        except Exception:
            return set()
    return set()


def load_order():
    if os.path.exists(ORDER_FILE):
        try:
            with open(ORDER_FILE) as f:
                return [str(x) for x in json.load(f)]
        except Exception:
            return []
    return []


def save_order(order):
    seen, clean = set(), []
    for x in (order or []):
        if x is None:
            continue
        x = str(x).strip()
        if x and x not in seen:
            seen.add(x)
            clean.append(x)
    os.makedirs(os.path.dirname(ORDER_FILE), exist_ok=True)
    with open(ORDER_FILE, "w") as f:
        json.dump(clean, f, indent=1)
    return True


def set_star(jobid, starred):
    """Write star state to job-scout-state.json (and legacy stars file for compat)."""
    update_job_state(str(jobid), {"starred": bool(starred)}, rebuild=True)
    # Keep legacy file in sync so jobscout_build.py can still read it during
    # the transition period before the build is also fully state-first.
    try:
        s = load_stars()
        if starred:
            s.add(str(jobid))
        else:
            s.discard(str(jobid))
        os.makedirs(os.path.dirname(STARS), exist_ok=True)
        with open(STARS, "w") as f:
            json.dump(sorted(s), f, indent=1)
    except Exception:
        pass
    return True


def set_status(jobid, status, applied_at=None):
    """Persist a manual application status to job-scout-state.json."""
    status = (status or "").strip().lower()
    if status not in STATUSES:
        return False
    updates = {"reviewStatus": status}
    # Auto-set appliedAt to today when first moving into a post-application status,
    # unless an explicit override is provided or one already exists.
    if status in ("applied", "interviewing", "rejected", "not-applying"):
        existing = scout_state.get("jobs", {}).get(str(jobid), {}).get("appliedAt", "")
        if applied_at:
            updates["appliedAt"] = applied_at[:10]
        elif not existing:
            updates["appliedAt"] = datetime.now().strftime("%Y-%m-%d")
    update_job_state(str(jobid), updates, rebuild=True)
    # Keep legacy status file in sync for build compat.
    try:
        s = {}
        if os.path.exists(STATUS_FILE):
            with open(STATUS_FILE) as f:
                s = json.load(f)
        s[str(jobid)] = status
        os.makedirs(os.path.dirname(STATUS_FILE), exist_ok=True)
        with open(STATUS_FILE, "w") as f:
            json.dump(s, f, indent=1)
    except Exception:
        pass
    return True


def update_job(jobid, dismissed, notes, salary=None):
    """Write dismissed + notes (+ optional salary) for one JobID to state.

    No longer writes to the xlsx directly; the workbook is regenerated by
    jobscout_build.py after each state change.
    """
    updates: dict = {
        "userDismissed": bool(dismissed),
        "manualNotes": notes or "",
    }
    if dismissed:
        # Only override reviewStatus to dismissed if it isn't something more
        # meaningful (applied/interviewing) — let the caller set those via
        # set_status().
        entry = scout_state.get("jobs", {}).get(str(jobid), {})
        if entry.get("reviewStatus") not in ("applied", "interviewing"):
            updates["reviewStatus"] = "dismissed"
    else:
        # Undismiss: revert to new only if previously dismissed
        entry = scout_state.get("jobs", {}).get(str(jobid), {})
        if entry.get("reviewStatus") == "dismissed":
            updates["reviewStatus"] = "new"

    if salary is not None:
        updates["overrides"] = {"salary": (salary or "").strip()}

    # notes-only saves (called on a timer) skip rebuild to avoid hammering disk
    rebuild = (dismissed != (scout_state.get("jobs", {}).get(str(jobid), {}).get("userDismissed", False)))
    update_job_state(str(jobid), updates, rebuild=rebuild or (salary is not None))
    return True


def update_jobs_bulk(jobids, dismissed):
    """Set userDismissed for many JobIDs. Returns count updated."""
    wanted = [str(j).strip() for j in jobids if str(j).strip()]
    if not wanted:
        return 0
    for jid in wanted:
        updates: dict = {"userDismissed": bool(dismissed)}
        entry = scout_state.get("jobs", {}).get(jid, {})
        if dismissed and entry.get("reviewStatus") not in ("applied", "interviewing"):
            updates["reviewStatus"] = "dismissed"
        elif not dismissed and entry.get("reviewStatus") == "dismissed":
            updates["reviewStatus"] = "new"
        update_job_state(jid, updates, rebuild=False)
    trigger_rebuild()
    return len(wanted)


def _linkedin_id(link):
    m = re.search(r"/jobs/view/(\d+)", link or "")
    if m:
        return m.group(1)
    m = re.search(r"\b(\d{9,})\b", link or "")
    return m.group(1) if m else None


def _existing_ids():
    ids = set()
    try:
        wb = openpyxl.load_workbook(XLSX, read_only=True)
        ws = wb[SHEET]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                ids.add(str(row[0]).strip())
        wb.close()
    except Exception:
        pass
    return ids


def add_manual_job(f, initial_status="new"):
    """Append a manually-entered job to the spreadsheet + data.json immediately.
    Returns (ok, message). Pass initial_status='applied' when promoting from applied.json."""
    title = (f.get("title") or "").strip()
    company = (f.get("company") or "").strip()
    if not title or not company:
        return False, "Title and Company are required."
    link = (f.get("link") or "").strip()
    jid = _linkedin_id(link) or ("m" + datetime.now().strftime("%Y%m%d%H%M%S"))
    if jid in _existing_ids():
        return False, "That job is already in your list."
    location = (f.get("location") or "").strip() or "United States"
    desc = (f.get("description") or "").strip()
    derived = core.classify({"title": title, "companyName": company, "location": location,
                             "descriptionText": desc}, "manual") if core else {}
    workplace = (f.get("workplace") or "").strip() or derived.get("workplace", "Remote — unverified")
    salary = (f.get("salary") or "").strip() or derived.get("salary", "")
    keywords = derived.get("keywords", [])
    category = derived.get("category", "")
    geofit = core.geo_fit(workplace, location) if core else True
    try:
        score = max(0, min(10, int(f.get("score") or 0)))
    except Exception:
        score = 0
    posted = (f.get("posted") or "").strip() or datetime.now().strftime("%Y-%m-%d")
    reason = (f.get("reason") or "").strip() or "Manually added."
    notes = (f.get("notes") or "").strip()
    if not link:
        link = f"https://www.linkedin.com/jobs/view/{jid}" if jid.isdigit() else ""

    with _write_lock:
        try:
            shutil.copy2(XLSX, BACKUP)
        except Exception:
            pass
        wb = openpyxl.load_workbook(XLSX)
        ws = wb[SHEET]
        ws.append([jid, score, title, company, workplace, location, salary, category, posted,
                   initial_status, "", "", notes, ", ".join(keywords), reason, "", "x", "x" if geofit else "", desc])
        if link:
            lc = ws.cell(row=ws.max_row, column=COL["Link"])
            lc.value = "Open ↗"
            lc.hyperlink = link
        wb.save(XLSX)
        wb.close()

    # Persist into data.json so it survives the twice-daily rebuild.
    dpath = os.path.join(HERE, "data.json")
    try:
        with open(dpath) as fh:
            data = json.load(fh)
        rec = {"id": jid, "title": title, "company": company, "location": location, "posted": posted,
               "score": score, "reason": reason, "workplace": workplace, "salary": salary,
               "caExcluded": False, "keywords": keywords, "category": category,
               "ocTarget": (core.is_in_oc(location) if core else False), "geoFit": geofit,
               "description": desc, "dismissed": False, "notes": notes, "applied": False,
               "status": initial_status, "appliedDate": "", "addedAt": datetime.now(timezone.utc).isoformat(),
               "src": "manual", "manual": True, "needsScore": (score == 0 and bool(desc)),
               "link": link}
        data.setdefault("jobs", []).insert(0, rec)
        with open(dpath, "w") as fh:
            json.dump(data, fh, indent=1)
    except Exception:
        pass

    # Pin salary/workplace so the rebuild keeps the manual values.
    if salary:
        _set_override(jid, "salary", salary)
    _set_override(jid, "workplace", workplace)
    # Auto-star manually-added jobs so they're easy to find for follow-up.
    set_star(jid, True)
    return True, jid


def _update_applied_json_entry(manual_jobid: str, new_company: str, new_role: str):
    """Find the applied.json entry whose hash matches manual_jobid and update its
    company/role to the new values. Called on promotion so applied.json stays in sync
    and Source 2 dedup doesn't produce a stale duplicate under the old name."""
    applied_path = os.path.join(HERE, "job-scout-applied.json")
    if not os.path.exists(applied_path):
        return
    try:
        with open(applied_path) as fh:
            data = json.load(fh)
        roles = data.get("appliedRoles", [])
        import hashlib as _hl
        for role_entry in roles:
            nc = _norm_company(role_entry.get("company", ""))
            nr = _norm_company(role_entry.get("role", ""))
            synthetic = "manual_" + _hl.md5((nc + "|" + nr).encode()).hexdigest()[:10]
            if synthetic == manual_jobid:
                role_entry["company"] = new_company.strip()
                role_entry["role"] = new_role.strip()
                break
        with open(applied_path, "w") as fh:
            json.dump(data, fh, indent=2)
    except Exception:
        pass


def edit_job(jobid, f):
    """Update editable fields of an existing job in data.json + state.

    Overridable fields (salary, workplace, score) go to state["jobs"][id]["overrides"].
    Non-override fields (title, company, location, posted, notes, link, reason) go
    directly into data.json so they survive rebuilds.

    Synthetic manual_ jobs (from applied.json) are promoted to real xlsx/data.json
    rows on first edit, with state migrated to the new real jobid.
    """
    jobid = str(jobid or "").strip()
    if not jobid:
        return False, "missing job id"

    # Promote synthetic applied.json entries to real jobs on first edit.
    # If already promoted (a real record exists with matching company+title),
    # redirect to that real jobid so the rest of edit_job can find and update it.
    if jobid.startswith("manual_"):
        real_jid = None

        # If the user explicitly linked this entry to a specific job, use it directly.
        linked_jid = str(f.get("linkedJobId") or "").strip()
        if linked_jid and not linked_jid.startswith("manual_"):
            real_jid = linked_jid
            # Update applied.json to match the linked job's company/title so dedup works.
            try:
                with open(DATA_JSON) as fh:
                    d = json.load(fh)
                for rec in d.get("jobs", []):
                    if str(rec.get("id", "")) == real_jid:
                        _update_applied_json_entry(
                            jobid,
                            str(rec.get("companyName") or rec.get("company") or ""),
                            str(rec.get("title") or ""),
                        )
                        break
            except Exception:
                pass

        if not real_jid and os.path.exists(DATA_JSON):
            try:
                with open(DATA_JSON) as fh:
                    d = json.load(fh)
                nc = _norm_company(f.get("company") or "")
                nt = _norm_company(f.get("title") or "")
                for rec in d.get("jobs", []):
                    rjid = str(rec.get("id") or "")
                    if not rjid or rjid.startswith("manual_"):
                        continue
                    if (_norm_company(str(rec.get("companyName") or rec.get("company") or "")) == nc and
                            _norm_company(str(rec.get("title") or "")) == nt):
                        real_jid = rjid
                        break
            except Exception:
                pass
        if not real_jid:
            # First edit — promote to a real xlsx/data.json record (already applied)
            ok, new_jid = add_manual_job(f, initial_status="applied")
            if not ok:
                return False, new_jid
            real_jid = new_jid
        # Migrate any persisted state (status, starred, notes) from synthetic to real jobid.
        # Always run — even if old_state is empty we must set reviewStatus so the promoted
        # record appears in Source 1 (Applied) and not in New Jobs.
        old_state = scout_state.get("jobs", {}).pop(jobid, {})
        entry = scout_state["jobs"].setdefault(real_jid, {})
        for k, v in old_state.items():
            if k not in entry or not entry[k]:
                entry[k] = v
        # Promoted entries came from applied.json — always mark as applied unless the
        # migrated state already has a more specific status (interviewing, rejected, etc.).
        if not entry.get("reviewStatus"):
            entry["reviewStatus"] = "applied"
        if save_state:
            save_state(STATE_PATH, scout_state)

        # Update applied.json in place so the entry reflects the new company/role.
        # This prevents a stale applied.json entry (old name) from creating a duplicate
        # in Source 2 after promotion.
        _update_applied_json_entry(jobid, f.get("company",""), f.get("title",""))

        jobid = real_jid  # redirect all subsequent logic to the real record

    # --- Merge real job into master (dedup: dismiss current, migrate state) ---
    linked_jid = str(f.get("linkedJobId") or "").strip()
    if linked_jid and not jobid.startswith("manual_") and linked_jid != jobid:
        # Migrate state from duplicate to master
        dup_state = scout_state.get("jobs", {}).get(jobid, {})
        master_entry = scout_state["jobs"].setdefault(linked_jid, {})
        for k, v in dup_state.items():
            if k not in master_entry or not master_entry[k]:
                master_entry[k] = v
        # Dismiss the duplicate
        scout_state["jobs"].setdefault(jobid, {})["dismissed"] = True
        if save_state:
            save_state(STATE_PATH, scout_state)
        return True, f"Merged into {linked_jid}"

    # --- Write overrides to state (salary, workplace, score) ---
    overrides: dict = {}
    if "salary" in f and (f.get("salary") or "").strip():
        overrides["salary"] = f["salary"].strip()
    if "workplace" in f and (f.get("workplace") or "").strip():
        overrides["workplace"] = f["workplace"].strip()
    if "score" in f and str(f.get("score", "")).strip() != "":
        try:
            overrides["score"] = max(0, min(10, int(f["score"])))
        except Exception:
            pass
    state_updates: dict = {}
    if overrides:
        state_updates["overrides"] = overrides
    if "notes" in f and f["notes"] is not None:
        state_updates["manualNotes"] = f["notes"]
    if "appliedAt" in f and f["appliedAt"] is not None:
        state_updates["appliedAt"] = (f["appliedAt"] or "")[:10]
    if "followUpDate" in f:
        state_updates["followUpDate"] = (f.get("followUpDate") or "")[:10]
    if state_updates:
        update_job_state(jobid, state_updates, rebuild=False)

    # --- Mirror non-override edits into data.json ---
    dpath = DATA_JSON
    found_in_data = False
    if os.path.exists(dpath):
        try:
            with open(dpath) as fh:
                data = json.load(fh)
            for rec in data.get("jobs", []):
                if str(rec.get("id")) == jobid:
                    field_map = {
                        "title": "title", "company": "companyName",
                        "location": "locationRaw", "posted": "postedAt",
                        "reason": "scoreReason", "link": "linkedinUrl",
                        "description": "description",
                    }
                    for form_key, rec_key in field_map.items():
                        if form_key in f and f[form_key] is not None:
                            rec[rec_key] = f[form_key]
                    # Also update legacy keys for older records
                    for form_key, rec_key in [("title", "title"), ("company", "company"),
                                               ("location", "location"), ("posted", "posted"),
                                               ("reason", "why"), ("link", "link"),
                                               ("description", "description")]:
                        if form_key in f and f[form_key] is not None:
                            rec[rec_key] = f[form_key]
                    # Flag for re-scoring on next run when description added and score unset
                    if f.get("description") and not overrides.get("score") and not rec.get("score"):
                        rec["needsScore"] = True
                    found_in_data = True
                    break
            with open(dpath, "w") as fh:
                json.dump(data, fh, indent=1)
        except Exception:
            pass

    # If job isn't in data.json, fall back to xlsx edit for non-override fields
    if not found_in_data and os.path.exists(XLSX):
        colmap = {"Score": "score", "Title": "title", "Company": "company", "Workplace": "workplace",
                  "Location": "location", "Salary": "salary", "Posted": "posted", "Notes": "notes",
                  "Why": "reason", "Description": "description"}
        with _write_lock:
            try:
                shutil.copy2(XLSX, BACKUP)
            except Exception:
                pass
            wb = openpyxl.load_workbook(XLSX)
            ws = wb[SHEET]
            found = False
            for row in ws.iter_rows(min_row=2):
                cell = row[COL["JobID"] - 1]
                if cell.value is not None and str(cell.value).strip() == jobid:
                    r = cell.row
                    for colname, key in colmap.items():
                        if key in f and f[key] is not None:
                            val = f[key]
                            if colname == "Score":
                                try:
                                    val = max(0, min(10, int(val)))
                                except Exception:
                                    continue
                            ws.cell(row=r, column=COL[colname]).value = val if val != "" else None
                    if "link" in f and f["link"] is not None:
                        lc = ws.cell(row=r, column=COL["Link"])
                        if f["link"].strip():
                            lc.value = "Open ↗"; lc.hyperlink = f["link"].strip()
                        else:
                            lc.value = None; lc.hyperlink = None
                    if core:
                        wpv = ws.cell(row=r, column=COL["Workplace"]).value or ""
                        locv = ws.cell(row=r, column=COL["Location"]).value or ""
                        ws.cell(row=r, column=COL["GeoFit"]).value = "x" if core.geo_fit(wpv, locv) else None
                    found = True
                    break
            if not found:
                wb.close()
                return False, "job not found"
            wb.save(XLSX)
            wb.close()

    # Also keep legacy overrides file in sync for build compat
    if "salary" in overrides:
        _set_override(jobid, "salary", overrides["salary"])
    if "workplace" in overrides:
        _set_override(jobid, "workplace", overrides["workplace"])
    if "score" in overrides:
        _set_override(jobid, "score", overrides["score"])

    trigger_rebuild()
    return True, jobid


PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Job Scout</title>
<style>
  :root{
    --bg:#f6f7f9; --panel:#ffffff; --ink:#1a1d21; --muted:#6b7280; --line:#e6e8eb;
    --accent:#2563eb; --accent-soft:#eef3ff; --good:#16a34a; --warn:#b45309;
    --shadow:0 1px 2px rgba(16,24,40,.06),0 1px 3px rgba(16,24,40,.08);
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--ink);
    font:15px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;}
  a{color:var(--accent);text-decoration:none}
  a:hover{text-decoration:underline}
  header{position:sticky;top:0;z-index:10;background:rgba(246,247,249,.85);
    backdrop-filter:saturate(180%) blur(12px);border-bottom:1px solid var(--line);}
  .bar{max-width:980px;margin:0 auto;padding:14px 20px;}
  .titlerow{display:flex;align-items:baseline;gap:12px;flex-wrap:wrap}
  h1{font-size:19px;margin:0;font-weight:650;letter-spacing:-.01em}
  .built{color:var(--muted);font-size:12.5px}
  .controls{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-top:12px}
  .filtertoggle{display:inline-flex;align-items:center;gap:5px;font:inherit;font-size:13px;
    color:var(--muted);cursor:pointer;padding:5px 10px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);margin-left:auto;white-space:nowrap;}
  .filtertoggle:hover{background:#f8f9fb;}
  input[type=search],select{font:inherit;padding:8px 11px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);color:var(--ink);outline:none}
  input[type=search]{flex:1;min-width:200px}
  input[type=search]:focus,select:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .toggle{display:inline-flex;align-items:center;gap:7px;font-size:13.5px;color:var(--muted);
    cursor:pointer;user-select:none;padding:7px 10px;border:1px solid var(--line);border-radius:9px;background:var(--panel)}
  .toggle input{accent-color:var(--accent)}
  .count{max-width:980px;margin:8px auto 0;padding:0 20px;color:var(--muted);font-size:13px}
  main{max-width:980px;margin:0 auto;padding:16px 20px 80px}
  .card{background:var(--panel);border:1px solid var(--line);border-radius:13px;
    padding:16px 18px;margin-bottom:13px;box-shadow:var(--shadow);transition:opacity .15s;position:relative;}
  .xbtn{flex:none;margin-left:4px;background:none;border:none;cursor:pointer;
    font-size:16px;line-height:1;color:#d1d5db;padding:0 6px;transition:color .1s;}
  .xbtn:hover{color:#ef4444;}
  .xbtn.on{color:#d1d5db;}
  .xbtn.on:hover{color:#374151;}
  .card.dismissed{opacity:.55}
  .chead{display:flex;gap:12px;align-items:flex-start}
  .score{flex:none;width:40px;height:40px;border-radius:10px;display:grid;place-items:center;
    font-weight:700;font-size:16px;color:#fff;background:var(--muted)}
  select.score{appearance:none;-webkit-appearance:none;-moz-appearance:none;display:inline-block;
    line-height:40px;text-align:center;text-align-last:center;cursor:pointer;border:none;padding:0;outline:none}
  .s9,.s10{background:#15803d}.s7,.s8{background:#16a34a}.s5,.s6{background:#ca8a04}
  .s3,.s4{background:#ea580c}.s0,.s1,.s2{background:#9ca3af}
  .ctitle{flex:1;min-width:0}
  .ctitle h2{margin:0;font-size:16px;font-weight:620;letter-spacing:-.01em;line-height:1.35}
  .meta{color:var(--muted);font-size:13px;margin-top:3px;display:flex;gap:7px;flex-wrap:wrap;align-items:center}
  .wp{font-size:11.5px;font-weight:600;padding:2px 8px;border-radius:20px}
  .wp.Remote{background:#dcfce7;color:#166534}
  .wp.unv{background:#fef3c7;color:#92400e}
  .wp.Hybrid,.wp.Onsite{background:#ffe4cc;color:#9a3412}
  .wpsel{font:inherit;font-size:11.5px;font-weight:600;border:1px solid var(--line);border-radius:8px;
    padding:2px 4px 2px 7px;cursor:pointer;outline:none}
  .wpsel.Remote{background:#dcfce7;color:#166534}
  .wpsel.unv{background:#fef3c7;color:#92400e}
  .wpsel.Hybrid,.wpsel.Onsite{background:#ffe4cc;color:#9a3412}
  .applied{background:var(--accent-soft);color:var(--accent);font-size:11.5px;font-weight:600;padding:2px 8px;border-radius:20px}
  .statsel{font:inherit;font-size:11.5px;font-weight:600;border:1px solid var(--line);border-radius:8px;
    padding:2px 4px 2px 7px;cursor:pointer;outline:none;text-align:center;text-align-last:center}
  .statsel.st-new{background:#eef1f6;color:#475569}
  .statsel.st-applied{background:var(--accent-soft);color:var(--accent)}
  .statsel.st-interviewing{background:#e0f2fe;color:#075985}
  .statsel.st-rejected{background:#fee2e2;color:#991b1b}
  .statsel.st-drafting{background:#ede9fe;color:#5b21b6}
  .statsel.st-not-applying{background:#f1f5f9;color:#64748b}
  .grip{cursor:grab;color:var(--muted);font-size:17px;line-height:1;padding:0 2px;align-self:center;user-select:none}
  .grip:active{cursor:grabbing}
  .card.dragging{opacity:.45}
  .card.dragover{box-shadow:inset 0 3px 0 -1px var(--accent)}
  .why{margin:11px 0 0;font-size:14px;color:#374151}
  .kw{margin-top:9px;display:flex;gap:6px;flex-wrap:wrap}
  .kw span{font-size:11.5px;color:var(--muted);background:#f1f3f5;padding:2px 8px;border-radius:6px}
  .foot{margin-top:13px;display:flex;gap:10px;align-items:flex-end}
  /* wrapper so the salary input isn't a direct flex child (WebKit stretches and
     ignores height/max-height on form controls that are direct flex children) */
  .footact{flex:none;display:flex;align-items:flex-end;gap:10px}
  /* plain block wrapper: the input is NOT a flex item here, so its height sticks */
  .salwrap{flex:none;align-self:flex-end}
  .salwrap .salin{display:block}
  textarea{flex:1;font:inherit;font-size:13.5px;padding:8px 10px;line-height:19px;border:1px solid var(--line);
    border-radius:9px;background:#fcfcfd;resize:none;overflow:hidden;height:37px;min-height:37px;
    box-sizing:border-box;outline:none}
  textarea:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .salin{width:124px;height:37px;box-sizing:border-box;font:inherit;font-size:13px;
    padding:8px 10px;line-height:19px;border:1px solid var(--line);border-radius:9px;
    background:#fcfcfd;outline:none;overflow:hidden;white-space:nowrap}
  .salin:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .salin.ph{border-style:dashed;color:#9aa0a6}
  .star{flex:none;margin-left:8px;background:none;border:none;cursor:pointer;font-size:16px;
    line-height:1;color:#d1d5db;padding:0 2px}
  .star:hover{color:#f5b301}
  .star.on{color:#f5b301}
  .undo{font:inherit;font-size:13.5px;font-weight:550;padding:7px 12px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);color:var(--ink);cursor:pointer}
  .undo:hover:not(:disabled){background:#f8f9fb}
  .undo:disabled{opacity:.4;cursor:default}
  .addbtn{font:inherit;font-size:13.5px;font-weight:550;padding:7px 12px;border:1px solid var(--accent);
    border-radius:9px;background:var(--accent-soft);color:var(--accent);cursor:pointer}
  .addbtn:hover{background:#e3ecff}
  #addpanel{display:none;max-width:980px;margin:10px auto 0;padding:14px 16px;border:1px solid var(--line);
    border-radius:12px;background:var(--panel);box-shadow:var(--shadow)}
  .addrow{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px}
  #addpanel input,#addpanel select{font:inherit;font-size:13.5px;padding:8px 11px;border:1px solid var(--line);
    border-radius:9px;background:#fcfcfd;outline:none;flex:1;min-width:140px}
  #addpanel input:focus,#addpanel select:focus,#addpanel textarea:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  #addpanel textarea{width:100%;box-sizing:border-box;font:inherit;font-size:13.5px;padding:9px 11px;
    border:1px solid var(--line);border-radius:9px;background:#fcfcfd;resize:vertical;min-height:90px;outline:none;margin-bottom:10px}
  .addsubmit{font:inherit;font-size:13.5px;font-weight:600;padding:9px 16px;border:none;border-radius:9px;
    background:var(--accent);color:#fff;cursor:pointer}
  .addsubmit:hover{background:#1d4ed8}
  .editbtn{flex:none;margin-left:4px;background:none;border:none;cursor:pointer;font-size:16px;color:#c2c7cd;padding:0 2px;line-height:1}
  .editbtn:hover{color:var(--accent)}
  .expbtn{flex:none;margin-left:4px;background:none;border:none;cursor:pointer;font-size:16px;color:#c2c7cd;padding:0 2px;line-height:1}
  .expbtn:hover{color:var(--good)}
  .copybtn{flex:none;margin-left:4px;background:none;border:none;cursor:pointer;font-size:16px;color:#c2c7cd;padding:0 2px;line-height:1}
  .copybtn:hover{color:var(--accent)}
  .copybtn.copied{color:var(--good)}
  .editform{display:none;margin-top:12px;padding:12px;border:1px dashed var(--line);border-radius:10px;background:#fafbfc}
  .editform .addrow{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px}
  .editform input,.editform select,.editform textarea{font:inherit;font-size:13px;padding:7px 9px;border:1px solid var(--line);border-radius:8px;background:#fff;outline:none}
  .editform input,.editform select{flex:1;min-width:110px}
  .ef_merge_results{border:1px solid var(--line);border-radius:8px;background:#fff;margin-top:3px;overflow:hidden}
  .ef_merge_result:hover{background:var(--accent-soft)}
  .editform textarea{width:100%;box-sizing:border-box;resize:vertical;min-height:48px;margin-bottom:8px}
  .editform input:focus,.editform select:focus,.editform textarea:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .ef_cancel{font:inherit;font-size:13px;font-weight:500;padding:9px 14px;border:1px solid var(--line);border-radius:9px;background:#fff;color:var(--ink);cursor:pointer}
  .ef_cancel:hover{background:#f1f3f5}
  .btn{flex:none;font:inherit;font-size:13px;font-weight:550;padding:8px 13px;border-radius:9px;
    border:1px solid var(--line);background:var(--panel);color:var(--ink);cursor:pointer;white-space:nowrap}
  .btn:hover{background:#f8f9fb}
  .btn.on{background:#fee2e2;border-color:#fecaca;color:#b91c1c}
  .saved{font-size:12px;color:var(--good);opacity:0;transition:opacity .2s}
  .saved.show{opacity:1}
  .empty{text-align:center;color:var(--muted);padding:60px 0}
  .trackerbtn{font-size:13px;color:var(--accent);padding:5px 10px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);white-space:nowrap}
  .trackerbtn:hover{background:#f8f9fb;text-decoration:none}
  .dot{color:#d1d5db}
  .sel{flex:none;display:flex;align-items:center;padding-top:2px}
  .sel input{width:18px;height:18px;accent-color:var(--accent);cursor:pointer}
  .card.picked{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .selbar{position:fixed;left:50%;transform:translateX(-50%);bottom:22px;z-index:20;
    display:none;align-items:center;gap:14px;background:#1f2937;color:#fff;
    padding:11px 16px;border-radius:13px;box-shadow:0 8px 24px rgba(16,24,40,.28)}
  .selbar.show{display:flex}
  .selbar b{font-weight:650}
  .selbar button{font:inherit;font-size:13.5px;font-weight:550;padding:8px 14px;border-radius:9px;
    border:none;cursor:pointer}
  .selbar .dismiss{background:#ef4444;color:#fff}
  .selbar .dismiss:hover{background:#dc2626}
  .selbar .undismiss{background:#374151;color:#fff}
  .selbar .clear{background:transparent;color:#cbd5e1}
  .selbar .clear:hover{color:#fff}
  .selhint{color:var(--muted);font-size:12px;margin-left:8px}
</style>
</head>
<body>
<header>
  <div class="bar">
    <div class="titlerow">
      <h1>Job Scout</h1>
      <span class="built" id="built"></span>
      <a href="/tracker" class="trackerbtn">Application Tracker →</a>
      <button class="undo" id="restartbtn" title="Reload server code without leaving the browser">↺ Restart</button>
      <button class="filtertoggle" id="filtertoggle" title="Show/hide filters">Filters &#9662;</button>
    </div>
    <div class="controls">
      <input type="search" id="q" placeholder="Search title, company, keywords, reason…" autocomplete="off">
      <div id="filterextras" style="display:contents">
      <select id="workplace"><option value="">All workplaces</option></select>
      <select id="category"><option value="">All categories</option></select>
      <select id="minscore">
        <option value="0">Any score</option><option value="5">Score ≥ 5</option>
        <option value="7">Score ≥ 7</option><option value="8">Score ≥ 8</option><option value="9">Score ≥ 9</option>
      </select>
      <select id="sort">
        <option value="score">Sort: Score</option>
        <option value="posted">Sort: Newest</option>
        <option value="company">Sort: Company</option>
        <option value="manual">Sort: Manual (drag)</option>
      </select>
      <select id="geomode">
        <option value="fit">Remote or commutable only</option>
        <option value="all">All locations</option>
        <option value="out">Out-of-area only (not remote)</option>
      </select>
      <select id="datefilter" title="Filter by posted date">
        <option value="">Any date</option>
        <option value="1">Posted today</option>
        <option value="3">Last 3 days</option>
        <option value="7">Last 7 days</option>
        <option value="14">Last 14 days</option>
        <option value="30">Last 30 days</option>
      </select>
      <select id="starredonly" title="Filter by starred">
        <option value="all">★ All</option>
        <option value="only">★ Starred only</option>
        <option value="hide">★ Hide starred</option>
      </select>
      <label class="toggle"><input type="checkbox" id="startop" checked> ★ on top</label>
      <select id="statusfilter" title="Filter by application status">
        <option value="">All statuses</option>
        <option value="active">Active (applied+)</option>
        <option value="new">New</option>
        <option value="drafting">Drafting</option>
        <option value="applied">Applied</option>
        <option value="interviewing">Interviewing</option>
        <option value="rejected">Rejected</option>
        <option value="not-applying">Not Applying</option>
      </select>
      <label class="toggle"><input type="checkbox" id="showdismissed"> Show dismissed</label>
      <button class="addbtn" id="exportbtn" title="Export the current view to a spreadsheet (CSV)">⤓ Export view</button>
      <button class="addbtn" id="addbtn" title="Add a job the scraper missed">➕ Add job</button>
      <select id="minsalary" title="Filter by minimum annual salary">
        <option value="0">Any salary</option>
        <option value="80000">≥ $80K</option>
        <option value="100000">≥ $100K</option>
        <option value="120000">≥ $120K</option>
        <option value="140000">≥ $140K</option>
        <option value="160000">≥ $160K</option>
        <option value="180000">≥ $180K</option>
        <option value="200000">≥ $200K</option>
      </select>
      <select id="sourcefilter" title="Filter by source"><option value="">All sources</option></select>
      <button class="undo" id="clearfilters" title="Reset all filters to defaults">✕ Clear Filters</button>
      </div>
      <button class="undo" id="undo" disabled title="Undo last change (⌘Z)">↶ Undo</button>
    </div>
  </div>
  <div id="addpanel">
    <div class="addrow">
      <input id="ax_title" placeholder="Title *">
      <input id="ax_company" placeholder="Company *">
      <input id="ax_location" placeholder="Location (e.g. Remote, Irvine CA)">
    </div>
    <div class="addrow">
      <input id="ax_link" placeholder="LinkedIn / posting URL">
      <select id="ax_workplace"><option value="">Workplace: auto-detect</option><option>Remote</option><option>Hybrid</option><option>On-site</option></select>
      <input id="ax_salary" placeholder="Salary (optional — auto from description)">
      <input id="ax_score" type="number" min="0" max="10" placeholder="Score 0–10 (blank = AI scores it)">
      <input id="ax_posted" type="date" title="Posted date">
    </div>
    <textarea id="ax_desc" placeholder="Paste the full job description here (used to auto-derive salary/workplace/keywords, and for AI scoring on the next run)…"></textarea>
    <div class="addrow">
      <button id="ax_submit" class="addsubmit">Add to dashboard</button>
      <span id="ax_msg" class="built"></span>
    </div>
  </div>
  <div class="count" id="count"><span id="counttext"></span><span id="sizewarn" style="color:#b45309;font-weight:600"></span><span class="selhint">Tip: click a checkbox, then Shift-click another to select the range.</span></div>
</header>
<main id="list"></main>
<div class="selbar" id="selbar">
  <b><span id="selcount">0</span> selected</b>
  <button class="dismiss" id="bulkdismiss">Dismiss selected</button>
  <button class="undismiss" id="bulkundismiss">Un-dismiss</button>
  <button class="clear" id="bulkclear">Clear</button>
</div>

<script>
let JOBS=[];
let SIZE_MB=0;
let currentRows=[];
let ORDER=[];
let dragMode=false, dragId=null;
let selected=new Set();
let lastIdx=null;
const $=s=>document.querySelector(s);
function grow(t){t.style.height='37px';t.style.height=Math.max(37,t.scrollHeight)+'px';}
async function postJSON(url,body){try{const r=await fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});return (await r.json()).ok;}catch(e){return false;}}

function scoreClass(n){n=Math.round(Number(n)||0);return 's'+Math.max(0,Math.min(10,n));}
function relDate(s){
  const m=String(s||'').match(/^(\d{4})-(\d{2})-(\d{2})/);
  if(!m) return s||'';
  const d=new Date(+m[1],+m[2]-1,+m[3]);
  const now=new Date(), today=new Date(now.getFullYear(),now.getMonth(),now.getDate());
  const days=Math.round((today-d)/86400000);
  if(days<=0) return '0d';
  return days+'d';
}
function trunc(s,n){s=String(s||'');return s.length>n?s.slice(0,n)+'…':s;}
function compactSalary(s){
  if(!s) return '';
  return s.replace(/\$(\d{1,3}(?:,\d{3})+)/g,(_,n)=>'$'+Math.round(parseInt(n.replace(/,/g,''))/1000)+'K');
}
// Parse a raw salary string to a normalized annual integer (best-effort).
// Returns 0 if unparseable. Hourly rates × 2080; K-suffixed values expanded.
function parseSalaryNorm(s){
  if(!s) return 0;
  // Extract all dollar amounts in order
  const nums=[];
  const re=/\$\s*([\d,]+(?:\.\d+)?)\s*[Kk]?/g;
  let m;
  while((m=re.exec(s))!==null){
    let v=parseFloat(m[0].replace(/[$,\s]/g,''));
    if(/[Kk]/.test(m[0])) v*=1000;
    nums.push(v);
  }
  if(!nums.length) return 0;
  // Use the max value (top of range)
  let val=Math.max(...nums);
  // Hourly detection: value ≤ 200 and string contains /hr /hour per hour
  const hourly=/\bhr\b|\bhour\b|per hour|\/hr/i.test(s) || val<=200;
  if(hourly) val=val*2080;
  return Math.round(val);
}
function fmtShortDate(s){
  const m=String(s||'').match(/^(\d{4})-(\d{2})-(\d{2})/);
  if(!m) return s||'';
  const months=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  return months[+m[2]-1]+' '+String(+m[3]);
}
function postedColor(s){
  const m=String(s||'').match(/^(\d{4})-(\d{2})-(\d{2})/);
  if(!m) return 'var(--muted)';
  const d=new Date(+m[1],+m[2]-1,+m[3]);
  const today=new Date(); today.setHours(0,0,0,0);
  const days=Math.round((today-d)/86400000);
  if(days<=2)  return '#16a34a'; // green  — fresh, get in early
  if(days<=7)  return '#b45309'; // amber  — still competitive
  if(days<=14) return '#ea580c'; // orange — pipeline filling
  if(days<=30) return '#dc2626'; // red    — apply now or pass
  return 'var(--muted)';         // grey   — likely stale
}
function esc(s){return String(s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));}
function wpClass(w){if(w==='Remote')return'Remote';if(/unverified/i.test(w))return'unv';if(/hybrid/i.test(w))return'Hybrid';if(/site/i.test(w))return'Onsite';return'';}

const SOURCE_META={
  linkedin:{label:'LinkedIn',color:'#0a66c2',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#0a66c2"/><path d="M4.5 6.5H3V13H4.5V6.5ZM3.75 5.75A.875.875 0 1 0 3.75 4a.875.875 0 0 0 0 1.75ZM13 9.2C13 7.4 12.1 6.5 10.7 6.5c-.7 0-1.3.3-1.7.8V6.5H7.5V13H9V9.5c0-.8.4-1.3 1.1-1.3.7 0 1.1.5 1.1 1.3V13H12.7V9.2H13Z" fill="#fff"/></svg>`},
  indeed:{label:'Indeed',color:'#2164f3',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#2164f3"/><text x="4" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">in</text></svg>`},
  glassdoor:{label:'Glassdoor',color:'#0caa41',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#0caa41"/><text x="3.5" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">G</text></svg>`},
  greenhouse:{label:'Greenhouse',color:'#3d9970',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#3d9970"/><text x="3" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">gh</text></svg>`},
  lever:{label:'Lever',color:'#3c4fe0',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#3c4fe0"/><text x="3.5" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">L</text></svg>`},
  workday:{label:'Workday',color:'#f8783a',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#f8783a"/><text x="3" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">W</text></svg>`},
  ashby:{label:'Ashby',color:'#6c47ff',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#6c47ff"/><text x="3.5" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">A</text></svg>`},
  employer:{label:'Direct',color:'#64748b',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#64748b"/><path d="M3 11V6l5-3 5 3v5H9.5V8.5h-3V11H3Z" fill="#fff"/></svg>`},
  manual:{label:'Manual',color:'#94a3b8',svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#94a3b8"/><path d="M4 4h5l3 3v5H4V4Z" fill="#fff" opacity=".9"/><path d="M9 4v3h3" fill="none" stroke="#94a3b8" stroke-width="1"/><path d="M6 9h4M6 11h3" stroke="#64748b" stroke-width="1" stroke-linecap="round"/></svg>`},
};
function sourceBadge(source){
  if(!source)return'';
  const m=SOURCE_META[source];
  if(!m)return`<span style="display:inline-flex;align-items:center;font-size:11px;font-weight:600;padding:1px 6px;border-radius:4px;background:#f1f5f9;color:#64748b;white-space:nowrap">${esc(source)}</span>`;
  return`<span title="${m.label}" style="display:inline-flex;align-items:center;line-height:1">${m.svg}</span>`;
}
function targetBadge(j){
  if(!j.targetEmployer) return '';
  const label=j.targetEmployerName||j.company||'Target employer';
  return `<span title="Target employer: ${esc(label)}" style="display:inline-flex;align-items:center;font-size:11px;font-weight:600;padding:1px 6px;border-radius:4px;background:#fef3c7;color:#92400e;white-space:nowrap">◎ Target</span>`;
}
const INCENTIVE_COLORS={
  equity:  {bg:'#dcfce7',fg:'#166534'},
  bonus:   {bg:'#dbeafe',fg:'#1e40af'},
  ote:     {bg:'#ffedd5',fg:'#9a3412'},
  profit:  {bg:'#ccfbf1',fg:'#115e59'},
  retirement:{bg:'#f1f5f9',fg:'#475569'},
  signing: {bg:'#ede9fe',fg:'#5b21b6'},
};
const INCENTIVE_CAT={
  'RSU':'equity','Stock options':'equity','ESPP':'equity','ESOP':'equity','Equity':'equity',
  'Bonus':'bonus','STI':'bonus',
  'OTE':'ote',
  'Profit share':'profit',
  '401k match':'retirement',
  'Sign-on':'signing',
};
function incentivePills(incentives){
  if(!incentives||!incentives.length) return '';
  return incentives.map(label=>{
    const cat=INCENTIVE_CAT[label]||'bonus';
    const c=INCENTIVE_COLORS[cat]||INCENTIVE_COLORS.bonus;
    return `<span title="${esc(label)}" style="display:inline-flex;align-items:center;font-size:11px;font-weight:600;padding:1px 6px;border-radius:4px;background:${c.bg};color:${c.fg};white-space:nowrap">${esc(label)}</span>`;
  }).join('');
}

// Orange County + ~25-mile commutable ring (mirrors jobscout_core.is_commutable_to_oc).
const COMMUTE=/orange county|aliso viejo|anaheim|brea|buena park|costa mesa|cypress|dana point|fountain valley|fullerton|garden grove|huntington beach|irvine|la habra|la palma|laguna|los alamitos|mission viejo|newport beach|placentia|rancho santa margarita|san juan capistrano|santa ana|seal beach|stanton|tustin|villa park|westminster|yorba linda|long beach|lakewood|cerritos|artesia|norwalk|la mirada|whittier|downey|bellflower|hawaiian gardens|signal hill|diamond bar|walnut|rowland heights|hacienda heights|chino hills|chino|pomona|montclair|corona|norco|pico rivera|santa fe springs|la puente|carson|west covina|covina/i;
function isCommutable(loc){
  loc=(loc||'').trim().toLowerCase();
  if(!loc) return false;
  const states=[...loc.matchAll(/,\s*([a-z]{2})\b/g)].map(m=>m[1].toUpperCase());
  if(states.length && !states.includes('CA') && !loc.includes('california')) return false;
  return COMMUTE.test(loc);
}
// Verified-remote anywhere, or hybrid/on-site commutable to OC; else not a geo fit.
function geoFit(j){ return j.workplace==='Remote' || isCommutable(j.location); }

async function load(){
  const r=await fetch('/api/jobs'); const d=await r.json();
  JOBS=d.jobs; SIZE_MB=d.sizeMB||0; ORDER=d.order||[]; $('#built').textContent=d.built||'';
  const wp=[...new Set(JOBS.map(j=>j.workplace).filter(Boolean))].sort();
  const cat=[...new Set(JOBS.map(j=>j.category).filter(Boolean))].sort();
  const src=[...new Set(JOBS.map(j=>j.source).filter(Boolean))].sort();
  fill('#workplace',wp); fill('#category',cat); fill('#sourcefilter',src);
  restoreFilters(DYNAMIC_FILTERS);
  render();
}
function fill(sel,vals){const el=$(sel);vals.forEach(v=>{const o=document.createElement('option');o.value=v;o.textContent=v;el.appendChild(o);});}

function render(){
  const q=$('#q').value.toLowerCase().trim();
  const wp=$('#workplace').value, cat=$('#category').value, srcf=$('#sourcefilter').value;
  const ms=Number($('#minscore').value), sort=$('#sort').value;
  const showD=$('#showdismissed').checked;
  const geomode=$('#geomode').value;
  const statusf=$('#statusfilter').value;
  const starredOnly=$('#starredonly').value;
  const dateDays=Number($('#datefilter').value)||0;
  const dateThresh=dateDays?new Date(Date.now()-dateDays*864e5).toISOString().slice(0,10):'';
  const minSal=Number($('#minsalary').value)||0;
  let rows=JOBS.filter(j=>{
    if(j.starred) return false;
    // An active search OVERRIDES every other filter (incl. dismissed) — text match only.
    if(q){const hay=(j.title+' '+j.company+' '+j.location+' '+j.why+' '+j.keywords.join(' ')).toLowerCase();
      return hay.includes(q);}
    if(!showD && j.dismissed) return false;
    if(geomode==='fit' && !geoFit(j)) return false;
    if(geomode==='out' && geoFit(j)) return false;
    if(statusf){const st=j.status||'new';
      if(statusf==='active'){if(st==='new'||st==='not-applying') return false;}
      else if(st!==statusf) return false;}
    if(starredOnly==='only' && !j.starred) return false;
    if(starredOnly==='hide' && j.starred) return false;
    if(dateThresh && j.posted && j.posted < dateThresh) return false;
    if(wp && j.workplace!==wp) return false;
    if(cat && j.category!==cat) return false;
    if(srcf && j.source!==srcf) return false;
    if((Number(j.score)||0)<ms) return false;
    if(minSal){const sn=j.salaryNormAnnual||parseSalaryNorm(j.salary);if(sn>0&&sn<minSal) return false;}
    return true;
  });
  const manual=sort==='manual';
  dragMode=manual;
  const starTop=$('#startop').checked && !manual;
  const opos=manual?new Map(ORDER.map((id,i)=>[String(id),i])):null;
  rows.sort((a,b)=>{
    if(starTop){const d=(b.starred?1:0)-(a.starred?1:0); if(d) return d;}
    if(manual){
      const pa=opos.has(a.jobid)?opos.get(a.jobid):Infinity;
      const pb=opos.has(b.jobid)?opos.get(b.jobid):Infinity;
      if(pa!==pb) return pa-pb;
      return (Number(b.score)||0)-(Number(a.score)||0);
    }
    if(sort==='company') return String(a.company).localeCompare(b.company);
    if(sort==='posted') return String(b.posted).localeCompare(String(a.posted));
    return (Number(b.score)||0)-(Number(a.score)||0) || String(b.posted).localeCompare(String(a.posted));
  });
  const dCount=JOBS.filter(j=>j.dismissed).length;
  $('#counttext').textContent=`${rows.length} shown · ${JOBS.length} total · ${dCount} dismissed · ${SIZE_MB} MB`;
  const big=(SIZE_MB>=8)||(JOBS.length>2500);
  const w=$('#sizewarn'); if(w) w.textContent = big ? '  ⚠ file getting large — time to archive/trim' : '';
  currentRows=rows;
  // drop selections that are no longer visible
  const visible=new Set(rows.map(j=>j.jobid));
  selected.forEach(id=>{if(!visible.has(id))selected.delete(id);});
  const list=$('#list');
  if(!rows.length){list.innerHTML='<div class="empty">No jobs match these filters.</div>';updateSelBar();return;}
  list.innerHTML='';
  rows.forEach((j,i)=>list.appendChild(card(j,i)));
  updateSelBar();
  list.querySelectorAll('textarea').forEach(grow);
}

function updateSelBar(){
  const bar=$('#selbar');
  $('#selcount').textContent=selected.size;
  bar.classList.toggle('show',selected.size>0);
}

function toggleSelect(jobid,idx,shift){
  if(shift && lastIdx!==null){
    const [a,b]=[Math.min(lastIdx,idx),Math.max(lastIdx,idx)];
    const turnOn=!selected.has(jobid);
    for(let k=a;k<=b;k++){const id=currentRows[k].jobid; if(turnOn)selected.add(id); else selected.delete(id);}
  }else{
    if(selected.has(jobid))selected.delete(jobid); else selected.add(jobid);
    lastIdx=idx;
  }
  // refresh picked styling without full re-render
  document.querySelectorAll('.card').forEach(c=>{
    const id=c.getAttribute('data-id');
    const on=selected.has(id);
    c.classList.toggle('picked',on);
    const cb=c.querySelector('.sel input'); if(cb)cb.checked=on;
  });
  updateSelBar();
}

async function bulkSet(ids,dismissed){
  const r=await fetch('/api/bulk',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({jobids:ids,dismissed})});
  return (await r.json()).ok;
}
async function bulkDismiss(dismissed){
  const ids=[...selected];
  if(!ids.length)return;
  const prev=ids.map(id=>{const j=JOBS.find(x=>x.jobid===id); return {id, d:j?!!j.dismissed:false};});
  try{
    if(await bulkSet(ids,dismissed)){
      const idset=new Set(ids);
      JOBS.forEach(j=>{if(idset.has(j.jobid))j.dismissed=dismissed;});
      selected.clear(); lastIdx=null;
      pushUndo(`${dismissed?'dismiss':'un-dismiss'} ${ids.length}`,async()=>{
        const on=prev.filter(p=>p.d).map(p=>p.id), off=prev.filter(p=>!p.d).map(p=>p.id);
        if(on.length) await bulkSet(on,true);
        if(off.length) await bulkSet(off,false);
        const m=new Map(prev.map(p=>[p.id,p.d])); JOBS.forEach(j=>{if(m.has(j.jobid))j.dismissed=m.get(j.jobid);});
        render();
      });
      render();
    }else{alert('Bulk update failed.');}
  }catch(e){alert('Bulk update failed: '+e);}
}

async function postOrder(){
  try{await fetch('/api/order',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({order:ORDER})});}catch(e){}
}
function reorderTo(draggedId,targetId){
  const prev=ORDER.slice();
  const vis=currentRows.map(j=>j.jobid);
  const newVis=vis.filter(id=>id!==draggedId);
  let ti=newVis.indexOf(targetId); if(ti<0) ti=newVis.length;
  newVis.splice(ti,0,draggedId);                 // dragged lands just above the drop target
  let full=ORDER.slice();
  vis.forEach(id=>{if(!full.includes(id)) full.push(id);});   // make sure visible ids are represented
  const visSet=new Set(vis), slots=[];
  full.forEach((id,idx)=>{if(visSet.has(id)) slots.push(idx);});
  newVis.forEach((id,k)=>{full[slots[k]]=id;});  // rewrite only the visible slots, keep hidden ones fixed
  ORDER=full; postOrder();
  pushUndo('reorder',async()=>{ORDER=prev; await postOrder(); render();});
  render();
}
function card(j,idx){
  const el=document.createElement('div');
  el.className='card'+(j.dismissed?' dismissed':'')+(selected.has(j.jobid)?' picked':'');
  el.setAttribute('data-id',j.jobid);
  const wpc=wpClass(j.workplace);
  const meta=[];
  meta.push(`<span title="${esc(j.company)}">${esc(trunc(j.company,22))}</span>`);
  if(j.workplace) meta.push(`<select class="wpsel ${wpc}" title="Set workplace">${['Remote','Remote — unverified','Hybrid','On-site'].map(w=>`<option ${j.workplace===w?'selected':''}>${w}</option>`).join('')}</select>`);
  const loc=(j.location||'').trim();
  if(loc && !/^(united states|usa|us)$/i.test(loc)) meta.push(`<span title="${esc(loc)}">${esc(trunc(loc,20))}</span>`);
  if(j.salary) meta.push(esc(compactSalary(j.salary)));
  if(j.incentives&&j.incentives.length) meta.push(incentivePills(j.incentives));
  if(j.posted) meta.push(`<span title="Posted ${esc(j.posted)}" style="color:${postedColor(j.posted)};font-weight:600">${esc(relDate(j.posted))}</span>`);
  if(j.addedAt) meta.push(`<span title="Scraped ${esc(j.addedAt)}">Scraped ${esc(fmtShortDate(j.addedAt))}</span>`);
  const st=j.status||'new';
  meta.push(`<select class="statsel st-${st}" title="Application status">${['new','drafting','applied','interviewing','rejected','not-applying'].map(s=>`<option value="${s}" ${st===s?'selected':''}>${s==='not-applying'?'Not Applying':s.charAt(0).toUpperCase()+s.slice(1)}</option>`).join('')}</select>`);
  const thirdPartyBadge=j.thirdParty?`<span title="Posted by recruiter or aggregator — not the direct employer" style="font-size:11px;margin-right:4px;vertical-align:middle;position:relative;top:-1px;cursor:default">🔶</span>`:'';
  const titleHtml=thirdPartyBadge+(j.link?`<a class="joblink" href="${esc(j.link)}" target="_blank" rel="noopener noreferrer">${esc(j.title)} ↗</a>`:esc(j.title));
  const srcBadgeHtml=sourceBadge(j.source);
  const tgtBadgeHtml=targetBadge(j);
  el.innerHTML=`
    <div class="chead">
      ${dragMode?'<span class="grip" draggable="true" title="Drag to rerank">⠿</span>':''}
      <div class="sel"><input type="checkbox" ${selected.has(j.jobid)?'checked':''}></div>
      <select class="score ${scoreClass(j.score)}" title="Set score (overrides the auto-score)">${Array.from({length:11},(_,n)=>`<option ${Number(j.score)===n?'selected':''}>${n}</option>`).join('')}</select>
      <div class="ctitle">
        <h2>${titleHtml}</h2>
        <div class="meta">${srcBadgeHtml?srcBadgeHtml+'<span class="dot">·</span>':''}${tgtBadgeHtml?tgtBadgeHtml+'<span class="dot">·</span>':''}${meta.join('<span class="dot">·</span>')}</div>
      </div>
      <button class="star${j.starred?' on':''}" title="Star for follow-up">${j.starred?'★':'☆'}</button>
      ${j.link?`<button class="copybtn" title="Copy job URL">⧉</button>`:''}
      <button class="editbtn" title="Edit fields">✎</button>
      <button class="expbtn" title="Save job info to file">⤓</button>
      <button class="xbtn dbtn${j.dismissed?' on':''}" title="${j.dismissed?'Un-dismiss':'Dismiss'}">${j.dismissed?'↩':'&times;'}</button>
    </div>
    ${j.why?`<p class="why">${esc(j.why)}</p>`:''}
    ${j.keywords.length?`<div class="kw">${j.keywords.map(k=>`<span>${esc(k)}</span>`).join('')}</div>`:''}
    <div class="foot">
      <textarea placeholder="Notes…">${esc(j.notes)}</textarea>
      <div class="footact">
        ${j.salary?'':`<div class="salwrap"><div class="salin ph" contenteditable="true">Salary…</div></div>`}
      </div>
      <span class="saved">Saved</span>
    </div>
    <div class="editform">
      <div class="addrow">
        <input class="ef_title" value="${esc(j.title)}" placeholder="Title">
        <input class="ef_company" value="${esc(j.company)}" placeholder="Company">
        <input class="ef_location" value="${esc(j.location||'')}" placeholder="Location">
      </div>
      <div class="addrow">
        <select class="ef_workplace">${['Remote','Remote — unverified','Hybrid','On-site'].map(w=>`<option ${j.workplace===w?'selected':''}>${w}</option>`).join('')}</select>
        <input class="ef_salary" value="${esc(j.salary||'')}" placeholder="Salary">
        <input class="ef_score" type="number" min="0" max="10" value="${esc(j.score)}" placeholder="Score">
        <input class="ef_posted" type="date" value="${esc(j.posted||'')}" title="Posted date">
      </div>
      <input class="ef_link" value="${esc(j.link||'')}" placeholder="Link URL">
      <textarea class="ef_reason" placeholder="Why / fit notes">${esc(j.why||'')}</textarea>
      <textarea class="ef_description" placeholder="Full job description (used for AI scoring on next scheduled run)…">${esc(j.description||'')}</textarea>
      <div style="border-top:1px solid var(--line);padding-top:8px;margin-top:4px">
        <div style="font-size:11px;color:var(--muted);margin-bottom:4px">Merge into master record (marks this job as duplicate)</div>
        <input class="ef_merge_search" placeholder="Search for the master job record to merge into…" autocomplete="off"
          style="width:100%;box-sizing:border-box;font:inherit;font-size:13px;padding:7px 9px;border:1px solid var(--line);border-radius:8px;background:#fff;outline:none">
        <div class="ef_merge_results" style="display:none"></div>
        <div class="ef_merge_display" style="display:none;margin-top:6px;font-size:12px;color:var(--ink);align-items:center;gap:6px">
          <span class="ef_merge_label" style="background:var(--accent-soft);color:var(--accent);padding:2px 8px;border-radius:5px;font-weight:500"></span>
          <button class="ef_merge_unlink" style="font:inherit;font-size:11px;padding:2px 7px;border:1px solid var(--line);border-radius:5px;background:none;cursor:pointer;color:var(--muted)">✕ Clear</button>
        </div>
        <input class="ef_merge_jobid" type="hidden" value="">
      </div>
      <div class="addrow">
        <button class="ef_save addsubmit">Save changes</button>
        <button class="ef_cancel">Cancel</button>
      </div>
    </div>`;
  const cb=el.querySelector('.sel input');
  cb.addEventListener('click',e=>{toggleSelect(j.jobid,idx,e.shiftKey);});
  const grip=el.querySelector('.grip');
  if(grip){
    grip.addEventListener('dragstart',e=>{dragId=j.jobid; el.classList.add('dragging');
      e.dataTransfer.effectAllowed='move'; try{e.dataTransfer.setData('text/plain',j.jobid); e.dataTransfer.setDragImage(el,12,12);}catch(_){}});
    grip.addEventListener('dragend',()=>{dragId=null; el.classList.remove('dragging');
      document.querySelectorAll('.card.dragover').forEach(c=>c.classList.remove('dragover'));});
    el.addEventListener('dragover',e=>{if(!dragId)return; e.preventDefault(); e.dataTransfer.dropEffect='move'; if(dragId!==j.jobid)el.classList.add('dragover');});
    el.addEventListener('dragleave',()=>el.classList.remove('dragover'));
    el.addEventListener('drop',e=>{if(!dragId)return; e.preventDefault(); el.classList.remove('dragover');
      if(dragId!==j.jobid) reorderTo(dragId,j.jobid);});
  }
  const ta=el.querySelector('textarea');
  const salin=el.querySelector('.salin');
  if(salin){
    salin.style.setProperty('height','37px','important');
    salin.addEventListener('focus',()=>{if(salin.classList.contains('ph')){salin.textContent='';salin.classList.remove('ph');}});
  }
  const btn=el.querySelector('.dbtn');
  const saved=el.querySelector('.saved');
  const star=el.querySelector('.star');
  let timer=null, noteAtFocus=j.notes||'';
  function flash(){saved.classList.add('show');setTimeout(()=>saved.classList.remove('show'),1100);}
  async function save(){
    const ok=await post(j.jobid,j.dismissed,ta.value);
    if(ok){j.notes=ta.value; flash();}
  }
  ta.addEventListener('focus',()=>{noteAtFocus=j.notes||'';});
  ta.addEventListener('input',()=>{grow(ta);clearTimeout(timer);timer=setTimeout(save,700);});
  ta.addEventListener('keydown',e=>{if((e.metaKey||e.ctrlKey)&&e.key==='Enter'){e.preventDefault();btn.click();}});
  ta.addEventListener('blur',async()=>{clearTimeout(timer); const before=noteAtFocus; await save();
    if((j.notes||'')!==before){ pushUndo('note',async()=>{await post(j.jobid,j.dismissed,before,j.salary); j.notes=before; render();}); }});
  async function saveSalary(){
    let v=salin.classList.contains('ph')?'':salin.textContent.trim();
    const prev=j.salary||'';
    if(!v && !salin.classList.contains('ph')){salin.textContent='Salary…';salin.classList.add('ph');}
    if(v===prev) return;
    const ok=await post(j.jobid,j.dismissed,ta.value,v);
    if(ok){j.salary=v; flash();
      pushUndo('salary',async()=>{await post(j.jobid,j.dismissed,j.notes,prev); j.salary=prev; render();});
      render();}
  }
  if(salin){
    salin.addEventListener('blur',saveSalary);
    salin.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();salin.blur();}});
  }
  btn.addEventListener('click',async()=>{
    const prev=j.dismissed, nd=!j.dismissed;
    const ok=await post(j.jobid,nd,ta.value);
    if(ok){j.dismissed=nd; flash();
      pushUndo(nd?'dismiss':'un-dismiss',async()=>{await post(j.jobid,prev,j.notes,j.salary); j.dismissed=prev; render();});
      if(nd && !$('#showdismissed').checked){el.style.opacity=0;setTimeout(render,160);}
      else render();}
  });
  star.addEventListener('click',async()=>{
    const prev=!!j.starred, nv=!prev;
    if(await postStar(j.jobid,nv)){j.starred=nv; flash();
      pushUndo(nv?'star':'unstar',async()=>{await postStar(j.jobid,prev); j.starred=prev; render();});
      render();}
  });
  const wpsel=el.querySelector('.wpsel');
  if(wpsel) wpsel.addEventListener('change',async()=>{
    const nw=wpsel.value, prev=j.workplace;
    if(await postJSON('/api/editjob',{jobid:j.jobid,workplace:nw})){
      j.workplace=nw; wpsel.className='wpsel '+wpClass(nw); flash();
      pushUndo('workplace',async()=>{await postJSON('/api/editjob',{jobid:j.jobid,workplace:prev}); j.workplace=prev; render();});
    }else{wpsel.value=prev; alert('Could not update workplace.');}
  });
  const scoresel=el.querySelector('select.score');
  if(scoresel) scoresel.addEventListener('change',async()=>{
    const nv=parseInt(scoresel.value), prev=Number(j.score)||0;
    if(nv===prev) return;
    if(await postJSON('/api/editjob',{jobid:j.jobid,score:nv})){
      j.score=nv; j.manualScore=true;
      pushUndo('score',async()=>{await postJSON('/api/editjob',{jobid:j.jobid,score:prev}); j.score=prev; render();});
      render();
    }else{scoresel.value=prev; alert('Could not update score.');}
  });
  const statsel=el.querySelector('.statsel');
  if(statsel) statsel.addEventListener('change',async()=>{
    const ns=statsel.value, prev=j.status||'new';
    if(await postJSON('/api/status',{jobid:j.jobid,status:ns})){
      j.status=ns; j.applied=(ns!=='new'); statsel.className='statsel st-'+ns; flash();
      pushUndo('status',async()=>{await postJSON('/api/status',{jobid:j.jobid,status:prev}); j.status=prev; j.applied=(prev!=='new'); render();});
      if($('#statusfilter').value) render();
    }else{statsel.value=prev; alert('Could not update status.');}
  });
  const copybtn=el.querySelector('.copybtn');
  if(copybtn) copybtn.addEventListener('click',()=>{
    navigator.clipboard.writeText(j.link).then(()=>{
      copybtn.classList.add('copied'); copybtn.textContent='✓';
      setTimeout(()=>{copybtn.classList.remove('copied'); copybtn.textContent='⧉';},1500);
    });
  });
  const editbtn=el.querySelector('.editbtn'), editform=el.querySelector('.editform');
  editbtn.addEventListener('click',()=>{editform.style.display=editform.style.display==='block'?'none':'block';});
  el.querySelector('.ef_cancel').addEventListener('click',()=>{editform.style.display='none';});
  el.querySelector('.ef_save').addEventListener('click',async()=>{
    const gv=c=>el.querySelector('.'+c).value.trim();
    const mergeJobId=el.querySelector('.ef_merge_jobid').value||'';
    const body={jobid:j.jobid,title:gv('ef_title'),company:gv('ef_company'),location:gv('ef_location'),
      workplace:el.querySelector('.ef_workplace').value,salary:gv('ef_salary'),score:gv('ef_score'),
      posted:gv('ef_posted'),link:gv('ef_link'),reason:gv('ef_reason'),description:gv('ef_description'),
      ...(mergeJobId&&{linkedJobId:mergeJobId})};
    if(mergeJobId&&!confirm(`Merge this job into the selected master record? This job will be dismissed as a duplicate.`))return;
    if(await postJSON('/api/editjob',body)){
      if(mergeJobId){render();}
      else{
        Object.assign(j,{title:body.title,company:body.company,location:body.location,workplace:body.workplace,
          salary:body.salary,score:parseInt(body.score)||0,posted:body.posted,link:body.link,
          why:body.reason,description:body.description});
        render();
      }
    }else{alert('Could not save changes.');}
  });
  // Merge search wiring
  (()=>{
    const inp=el.querySelector('.ef_merge_search');
    const results=el.querySelector('.ef_merge_results');
    const display=el.querySelector('.ef_merge_display');
    const label=el.querySelector('.ef_merge_label');
    const hidden=el.querySelector('.ef_merge_jobid');
    el.querySelector('.ef_merge_unlink').addEventListener('click',()=>{
      hidden.value=''; display.style.display='none'; inp.value=''; inp.style.display='';
    });
    let _mTimer=null;
    inp.addEventListener('input',()=>{
      clearTimeout(_mTimer);
      const q=inp.value.trim();
      if(q.length<2){results.style.display='none';results.innerHTML='';return;}
      _mTimer=setTimeout(async()=>{
        try{
          const d=await(await fetch('/api/jobs/search?q='+encodeURIComponent(q))).json();
          const filtered=d.results.filter(r=>r.jobid!==j.jobid);
          if(!filtered.length){results.style.display='none';results.innerHTML='';return;}
          results.innerHTML=filtered.map(r=>`<div class="ef_merge_result" data-jid="${esc(r.jobid)}" data-company="${esc(r.company)}" data-title="${esc(r.title)}"
            style="padding:6px 10px;cursor:pointer;font-size:13px;border-bottom:1px solid var(--line)">${esc(r.company)} — ${esc(r.title)}</div>`).join('');
          results.style.display='block';
        }catch(_){}
      },220);
    });
    results.addEventListener('click',e=>{
      const row=e.target.closest('.ef_merge_result');
      if(!row)return;
      hidden.value=row.dataset.jid;
      label.textContent=row.dataset.company+' — '+row.dataset.title;
      display.style.display='flex';
      inp.style.display='none';
      results.style.display='none'; results.innerHTML='';
    });
    document.addEventListener('click',e=>{if(!el.contains(e.target))results.style.display='none';});
  })();
  const expbtn=el.querySelector('.expbtn');
  if(expbtn) expbtn.addEventListener('click',async()=>{
    expbtn.style.pointerEvents='none';
    try{
      const r=await fetch('/api/savejobmd',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({jobid:j.jobid})});
      const d=await r.json();
      if(d.ok){
        saved.textContent='Saved ✓'; saved.classList.add('show');
        expbtn.title='Saved!';
        setTimeout(()=>{saved.classList.remove('show');saved.textContent='Saved';expbtn.title='Save job info to file';},2000);
      }else{alert('Save failed — '+(d.error||'unknown error'));}
    }catch(e){alert('Save failed — '+e);}
    finally{expbtn.style.pointerEvents='';}
  });
  return el;
}

async function post(jobid,dismissed,notes,salary){
  try{
    const body={jobid,dismissed,notes};
    if(salary!==undefined) body.salary=salary;
    const r=await fetch('/api/update',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify(body)});
    const d=await r.json(); return d.ok;
  }catch(e){alert('Save failed: '+e);return false;}
}
async function postStar(jobid,starred){
  try{
    const r=await fetch('/api/star',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jobid,starred})});
    return (await r.json()).ok;
  }catch(e){alert('Star failed: '+e);return false;}
}

// --- Undo stack: each entry is a function that reverts the last change ---
let undoStack=[];
function updateUndoBtn(){const b=$('#undo'); if(b) b.disabled=undoStack.length===0;}
function pushUndo(label,fn){undoStack.push(fn); if(undoStack.length>100) undoStack.shift(); updateUndoBtn();}
async function doUndo(){const fn=undoStack.pop(); updateUndoBtn(); if(fn){try{await fn();}catch(e){alert('Undo failed: '+e);}}}

const STATIC_FILTERS=['#q','#minscore','#sort','#showdismissed','#geomode','#statusfilter','#starredonly','#startop','#datefilter','#minsalary'];
const DYNAMIC_FILTERS=['#workplace','#category','#sourcefilter'];
function saveFilters(){
  const state={};
  [...STATIC_FILTERS,...DYNAMIC_FILTERS].forEach(s=>{const el=$(s);state[s]=el.type==='checkbox'?el.checked:el.value;});
  localStorage.setItem('filterState',JSON.stringify(state));
}
function restoreFilters(ids){
  try{const state=JSON.parse(localStorage.getItem('filterState')||'{}');
    ids.forEach(s=>{if(state[s]===undefined)return;const el=$(s);if(el.type==='checkbox')el.checked=state[s];else el.value=state[s];});
  }catch(e){}
}
restoreFilters(STATIC_FILTERS);
[...STATIC_FILTERS,...DYNAMIC_FILTERS].forEach(s=>{
  $(s).addEventListener('input',()=>{render();saveFilters();});
});
$('#bulkdismiss').addEventListener('click',()=>bulkDismiss(true));
$('#bulkundismiss').addEventListener('click',()=>bulkDismiss(false));
$('#bulkclear').addEventListener('click',()=>{selected.clear();lastIdx=null;render();});
$('#undo').addEventListener('click',doUndo);
document.getElementById('clearfilters').addEventListener('click',()=>{
  $('#q').value='';
  $('#minscore').value='0';
  $('#sort').value='score';
  $('#showdismissed').checked=false;
  $('#geomode').value='fit';
  $('#statusfilter').value='';
  $('#starredonly').value='all';
  $('#startop').checked=true;
  $('#datefilter').value='';
  $('#minsalary').value='0';
  const wp=$('#workplace'); if(wp) wp.value='';
  const cat=$('#category'); if(cat) cat.value='';
  const src=$('#sourcefilter'); if(src) src.value='';
  render(); saveFilters();
});
document.addEventListener('keydown',e=>{if((e.metaKey||e.ctrlKey)&&e.key.toLowerCase()==='z'){e.preventDefault();doUndo();}});
(function(){
  const btn=$('#restartbtn');
  if(!btn)return;
  btn.addEventListener('click',async()=>{
    btn.textContent='Restarting…';btn.disabled=true;
    try{await fetch('/api/restart',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'})}catch(e){}
    const t=Date.now();
    while(Date.now()-t<15000){
      await new Promise(r=>setTimeout(r,400));
      try{const r=await fetch('/api/ping');if(r.ok){location.reload(true);return;}}catch(e){}
    }
    btn.textContent='↺ Restart';btn.disabled=false;alert('Server did not come back up.');
  });
})();
$('#addbtn').addEventListener('click',()=>{const p=$('#addpanel');const show=p.style.display!=='block';p.style.display=show?'block':'none';if(show)$('#ax_title').focus();});
$('#exportbtn').addEventListener('click',exportView);
async function exportView(){
  const btn=$('#exportbtn'), label=btn.textContent;
  btn.disabled=true; btn.textContent='Exporting…';
  try{
    const jobids=currentRows.map(j=>j.jobid);
    const r=await fetch('/api/export',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({jobids})});
    if(!r.ok) throw new Error('export failed');
    const csv=await r.text();
    const blob=new Blob([csv],{type:'text/csv;charset=utf-8'});
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    a.download='job-scout-view-'+new Date().toISOString().slice(0,10)+'.csv';
    document.body.appendChild(a);a.click();
    setTimeout(()=>{URL.revokeObjectURL(a.href);a.remove();},0);
  }catch(e){alert('Could not export: '+e);}
  finally{btn.disabled=false; btn.textContent=label;}
}
$('#ax_submit').addEventListener('click',addJob);
async function addJob(){
  const g=id=>$('#'+id).value.trim();
  const body={title:g('ax_title'),company:g('ax_company'),location:g('ax_location'),link:g('ax_link'),
    workplace:$('#ax_workplace').value,salary:g('ax_salary'),score:g('ax_score'),
    posted:g('ax_posted'),description:g('ax_desc')};
  const msg=$('#ax_msg');
  if(!body.title||!body.company){msg.textContent='Title and Company are required.';msg.style.color='#b91c1c';return;}
  msg.textContent='Adding…';msg.style.color='var(--muted)';
  try{
    const r=await fetch('/api/addjob',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
    const d=await r.json();
    if(d.ok){
      ['ax_title','ax_company','ax_location','ax_link','ax_salary','ax_score','ax_posted','ax_desc'].forEach(id=>$('#'+id).value='');
      $('#ax_workplace').value='';
      msg.textContent='Added ✓';msg.style.color='var(--good)';
      await load();
      setTimeout(()=>{$('#addpanel').style.display='none';msg.textContent='';},1000);
    }else{msg.textContent=d.message||'Could not add.';msg.style.color='#b91c1c';}
  }catch(e){msg.textContent='Error: '+e;msg.style.color='#b91c1c';}
}
load();
(function(){
  const btn=document.getElementById('filtertoggle');
  const extras=document.getElementById('filterextras');
  if(localStorage.getItem('filtersCollapsed')==='1'){
    extras.style.display='none';btn.innerHTML='Filters &#9656;';
  }
  btn.addEventListener('click',()=>{
    const collapsed=extras.style.display==='none';
    extras.style.display=collapsed?'contents':'none';
    btn.innerHTML=collapsed?'Filters &#9662;':'Filters &#9656;';
    localStorage.setItem('filtersCollapsed',collapsed?'0':'1');
  });
})();
</script>
</body>
</html>"""


def api_applied() -> str:
    """Merge dashboard applied/interviewing/rejected jobs with job-scout-applied.json.

    Returns enriched entries with job posting links, Gmail search URLs, and local PDF paths.
    """
    applied_path = os.path.join(HERE, "job-scout-applied.json")
    last_updated, manual_entries = "", []
    try:
        with open(applied_path) as f:
            raw = json.load(f)
        manual_entries = raw.get("appliedRoles", [])
        last_updated = raw.get("lastUpdated", "")
    except Exception:
        pass

    TARGET = {"applied", "drafting", "interviewing", "rejected", "not-applying"}
    try:
        all_jobs = read_jobs_from_data()
    except Exception:
        all_jobs = []

    result = []
    seen: set = set()           # exact (norm_company, norm_role) pairs from Source 1
    seen_companies: set = set() # norm_company values from Source 1 (for empty-role dedup)

    # Source 1: dashboard jobs with a terminal status
    for j in all_jobs:
        if j.get("status") not in TARGET:
            continue
        company = j.get("company") or ""
        role = j.get("title") or ""
        pdfs = _get_company_pdfs(company, role)
        _co_folder = pdfs.get("folder")
        _todo_md = os.path.join(_co_folder, "TODO.md") if _co_folder and os.path.isfile(os.path.join(_co_folder, "TODO.md")) else None
        result.append({
            "source": "dashboard",
            "jobid": j.get("jobid") or "",
            "company": company,
            "role": role,
            "posted": j.get("posted") or "",
            "addedAt": (j.get("addedAt") or "")[:10],
            "appliedAt": j.get("appliedAt") or j.get("addedAt") or "",
            "targetEmployer": j.get("targetEmployer") or False,
            "targetEmployerName": j.get("targetEmployerName") or "",
            "incentives": j.get("incentives") or [],
            "status": j.get("status", "applied"),
            "link": j.get("link") or "",
            "score": j.get("score"),
            "why": j.get("why") or "",
            "source": j.get("source") or "",
            "salary": j.get("salary") or "",
            "workplace": j.get("workplace") or "",
            "location": j.get("location") or "",
            "notes": j.get("notes") or "",
            "followUpDate": scout_state.get("jobs", {}).get(str(j.get("jobid") or ""), {}).get("followUpDate") or "",
            "gmailSearch": "https://mail.google.com/mail/u/0/#search/" + quote_plus('"' + company + '" applied'),
            "resume": pdfs["resume"],
            "coverLetter": pdfs["coverLetter"],
            "todoMd": _todo_md,
            "companyFolder": _co_folder,
        })
        seen.add((_norm_company(company), _norm_company(role)))
        seen_companies.add(_norm_company(company))

    # Source 2: manual / email-detected entries not already covered
    for m in manual_entries:
        company = m.get("company") or ""
        role = m.get("role") or ""
        nc = _norm_company(company)
        nr = _norm_company(role)
        pair = (nc, nr)
        if pair in seen:
            continue
        # Empty role in applied.json: skip if any Source 1 entry exists for this company
        if not nr and nc in seen_companies:
            continue
        # Role mismatch: skip if Source 1 already has an entry for this company where
        # one role is a substring of the other (handles "Content Strategist" vs
        # "Content Strategist (contractor via Intelliswift)")
        if nr and nc in seen_companies:
            if any(nr in s_nr or s_nr in nr for (s_nc, s_nr) in seen if s_nc == nc and s_nr):
                continue
        jobid = _manual_jobid(company, role)
        state_entry = scout_state.get("jobs", {}).get(jobid, {})
        status = state_entry.get("reviewStatus") or "applied"
        if status not in STATUSES:
            status = "applied"
        pdfs = _get_company_pdfs(company, role)
        _co_folder = pdfs.get("folder")
        _todo_md = os.path.join(_co_folder, "TODO.md") if _co_folder and os.path.isfile(os.path.join(_co_folder, "TODO.md")) else None
        overrides = state_entry.get("overrides") or {}
        result.append({
            "source": "manual",
            "jobid": jobid,
            "company": company,
            "role": role,
            "posted": m.get("posted") or "",
            "addedAt": (m.get("addedAt") or m.get("appliedAt") or "")[:10],
            "appliedAt": m.get("appliedAt") or "",
            "targetEmployer": False,
            "targetEmployerName": "",
            "incentives": [],
            "status": status,
            "link": m.get("link") or "",
            "score": overrides.get("score") or None,
            "why": "",
            "salary": overrides.get("salary") or m.get("salary") or "",
            "workplace": overrides.get("workplace") or m.get("workplace") or "",
            "location": m.get("location") or "",
            "notes": state_entry.get("manualNotes") or "",
            "followUpDate": state_entry.get("followUpDate") or "",
            "gmailSearch": "https://mail.google.com/mail/u/0/#search/" + quote_plus('"' + company + '" applied'),
            "resume": pdfs["resume"],
            "coverLetter": pdfs["coverLetter"],
            "todoMd": _todo_md,
            "companyFolder": _co_folder,
        })
        seen.add(pair)

    result.sort(key=lambda x: x.get("appliedAt") or "", reverse=True)
    return json.dumps({"roles": result, "lastUpdated": last_updated})


TRACKER_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Application Tracker</title>
<style>
  :root{--bg:#f6f7f9;--panel:#fff;--ink:#1a1d21;--muted:#6b7280;--line:#e6e8eb;
    --accent:#2563eb;--accent-soft:#eef3ff;--good:#16a34a;
    --shadow:0 1px 2px rgba(16,24,40,.06),0 1px 3px rgba(16,24,40,.08);}
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--ink);
    font:15px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;}
  a{color:var(--accent);text-decoration:none}
  a:hover{text-decoration:underline}
  /* Header */
  header{position:sticky;top:0;z-index:10;background:rgba(246,247,249,.88);
    backdrop-filter:saturate(180%) blur(12px);border-bottom:1px solid var(--line);}
  .bar{max-width:1040px;margin:0 auto;padding:14px 20px}
  .titlerow{display:flex;align-items:center;gap:12px}
  h1{font-size:19px;margin:0;font-weight:650;letter-spacing:-.01em}
  .back{font-size:13px;color:var(--muted);padding:5px 10px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);white-space:nowrap}
  .back:hover{background:#f3f4f6;text-decoration:none}
  #tracker-undo{font:inherit;font-size:13px;font-weight:550;padding:5px 12px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);cursor:pointer;white-space:nowrap;color:var(--ink)}
  #tracker-undo:hover:not(:disabled){background:#f3f4f6}
  #tracker-undo:disabled{opacity:.4;cursor:default}
  /* Main layout */
  main{max-width:1040px;margin:0 auto;padding:24px 20px 80px}
  section{margin-bottom:40px}
  .sec-head{display:flex;align-items:baseline;gap:10px;margin-bottom:14px}
  h2{font-size:15px;font-weight:650;letter-spacing:-.01em;margin:0}
  .scount{font-size:13px;color:var(--muted);font-weight:400;display:inline-flex;align-items:center;gap:6px;flex-wrap:wrap}
  /* To-do cards */
  .card{background:var(--panel);border:1px solid var(--line);border-radius:13px;
    padding:14px 16px;margin-bottom:10px;box-shadow:var(--shadow)}
  .chead{display:flex;gap:12px;align-items:flex-start}
  .score{flex:none;width:36px;height:36px;border-radius:9px;display:grid;place-items:center;
    font-weight:700;font-size:14px;color:#fff;background:var(--muted)}
  .s9,.s10{background:#15803d}.s7,.s8{background:#16a34a}.s5,.s6{background:#ca8a04}
  .s3,.s4{background:#ea580c}.s0,.s1,.s2{background:#9ca3af}
  .ctitle{flex:1;min-width:0}
  .ctitle h3{margin:0;font-size:15px;font-weight:620;letter-spacing:-.01em;line-height:1.35}
  .meta{color:var(--muted);font-size:13px;margin-top:3px;display:flex;gap:7px;flex-wrap:wrap;align-items:center}
  .dot{color:#d1d5db}
  .wp{font-size:11.5px;font-weight:600;padding:2px 8px;border-radius:20px}
  .wp.Remote{background:#dcfce7;color:#166534}
  .wp.Unv{background:#fef3c7;color:#92400e}
  .wp.Hybrid,.wp.Onsite{background:#ffe4cc;color:#9a3412}
  .why{margin:9px 0 0;font-size:13.5px;color:#374151}
  /* Application filters bar */
  .filters{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:14px}
  input[type=search]{font:inherit;padding:7px 11px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);color:var(--ink);outline:none;min-width:200px;flex:1}
  input[type=search]:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  select{font:inherit;padding:7px 11px;border:1px solid var(--line);
    border-radius:9px;background:var(--panel);color:var(--ink);outline:none;cursor:pointer}
  select:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  /* Status filter buttons */
  .sf-btn{font:inherit;font-size:12px;font-weight:600;padding:4px 10px;border-radius:20px;
    border:1.5px solid transparent;cursor:pointer;white-space:nowrap;transition:opacity .15s,filter .15s}
  .sf-btn.inactive{opacity:.35;filter:grayscale(.4)}
  .sf-btn.st-interviewing{background:#e0f2fe;color:#075985;border-color:#bae6fd}
  .sf-btn.st-applied{background:var(--accent-soft);color:var(--accent);border-color:#c7d7fd}
  .sf-btn.st-drafting{background:#ede9fe;color:#5b21b6;border-color:#ddd6fe}
  .sf-btn.st-rejected{background:#fee2e2;color:#991b1b;border-color:#fecaca}
  .sf-btn.st-not-applying{background:#f1f5f9;color:#64748b;border-color:#e2e8f0}
  /* Applications table */
  .app-list{display:flex;flex-direction:column;gap:4px}
  .app-row{display:flex;flex-direction:column;gap:5px;padding:9px 14px 8px}
  .app-row-top{display:flex;align-items:baseline;gap:9px;min-width:0}
  .app-row-bot{display:flex;align-items:center;gap:6px;flex-wrap:wrap;min-width:0}
  .app-score{flex:none;width:22px;height:22px;border-radius:5px;display:grid;place-items:center;font-weight:700;font-size:11px;color:#fff;align-self:center;box-sizing:border-box}
  .app-score b{font:inherit;display:block}
  .app-meta{display:flex;align-items:center;gap:5px;white-space:nowrap;margin-left:auto;flex:none}
  .app-age{font-size:12px;color:var(--muted);white-space:nowrap}
  .app-salary{font-size:12px;color:var(--muted);white-space:nowrap}
  .app-row-wrap{background:var(--panel);border:1px solid var(--line);border-radius:10px;
    overflow:hidden;position:relative}
  .app-row-wrap .app-row{border-bottom:none}
  .app-row:hover{background:#fafbfc}
  /* Left status bar — shown for applied/rejected/not-applying (replaces dropdown) */
  .app-row-wrap::before{content:'';position:absolute;left:0;top:0;bottom:0;width:3px;pointer-events:none}
  .app-row-wrap[data-status="applied"]::before{background:var(--accent)}
  .app-row-wrap[data-status="interviewing"]::before{background:#0284c7}
  .app-row-wrap[data-status="drafting"]::before{background:#7c3aed}
  .app-row-wrap[data-status="rejected"]::before{background:#dc2626}
  .app-row-wrap[data-status="not-applying"]::before{background:#94a3b8}
  /* Hide status dropdown on collapsed rows for terminal statuses — bar is the visual indicator.
     When expanded (collapsed class removed), the dropdown reappears so status can be changed. */
  .app-row-wrap.collapsed[data-status="applied"] .statsel,
  .app-row-wrap.collapsed[data-status="rejected"] .statsel,
  .app-row-wrap.collapsed[data-status="not-applying"] .statsel{display:none}
  /* Status badges — match main dashboard palette */
  .st-badge{display:inline-block;font-size:11.5px;font-weight:600;
    padding:2px 9px;border-radius:20px;white-space:nowrap}
  .wpsel{font:inherit;font-size:11.5px;font-weight:600;border:1px solid var(--line);border-radius:8px;
    padding:2px 4px 2px 7px;cursor:pointer;outline:none}
  .wpsel.Remote{background:#dcfce7;color:#166534;border-color:#bbf7d0}
  .wpsel.unv{background:#fef3c7;color:#92400e}
  .wpsel.Hybrid,.wpsel.Onsite{background:#ffe4cc;color:#9a3412}
  .statsel{font:inherit;font-size:11.5px;font-weight:600;border:1px solid var(--line);border-radius:8px;
    padding:2px 4px 2px 7px;cursor:pointer;outline:none;text-align:center;text-align-last:center}
  .statsel:focus{border-color:var(--accent);box-shadow:0 0 0 2px var(--accent-soft)}
  .statsel.st-new{background:#eef1f6;color:#475569}
  .statsel.st-applied{background:var(--accent-soft);color:var(--accent)}
  .statsel.st-interviewing{background:#e0f2fe;color:#075985}
  .statsel.st-drafting{background:#ede9fe;color:#5b21b6}
  .statsel.st-rejected{background:#fee2e2;color:#991b1b}
  .statsel.st-not-applying{background:#f1f5f9;color:#64748b}
  .st-applied{background:var(--accent-soft);color:var(--accent)}
  .st-interviewing{background:#e0f2fe;color:#075985}
  .st-drafting{background:#ede9fe;color:#5b21b6}
  .st-rejected{background:#fee2e2;color:#991b1b}
  .st-not-applying{background:#f1f5f9;color:#64748b}
  .app-company{font-weight:650;font-size:14px;white-space:nowrap;flex:none;max-width:220px;
    overflow:hidden;text-overflow:ellipsis}
  .app-role{font-size:13px;color:#374151;min-width:0;flex:1;
    overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .app-date{font-size:12.5px;color:var(--muted);white-space:nowrap;text-align:right}
  .app-actions{display:flex;gap:5px;align-items:center;flex-wrap:nowrap;margin-left:auto}
  .abtn{font:inherit;font-size:11.5px;font-weight:600;padding:3px 9px;
    border-radius:7px;border:1px solid var(--line);background:var(--panel);
    color:var(--ink);cursor:pointer;white-space:nowrap;text-decoration:none;
    display:inline-flex;align-items:center}
  .abtn:hover{background:#f3f4f6;text-decoration:none}
  .abtn.gmail{color:var(--accent);border-color:#c7d7fd}
  .abtn.res{color:#166534;border-color:#bbf7d0}
  .abtn.cl{color:#92400e;border-color:#fde68a}
  .app-editbtn{background:none;border:none;cursor:pointer;font-size:15px;color:#c2c7cd;
    padding:3px 5px;border-radius:6px;line-height:1;flex:none}
  .app-editbtn:hover{color:var(--accent);background:#f3f4f6}
  .app-editform{padding:10px 12px 12px;border-top:1px solid var(--line);background:#fafbfc;display:flex;flex-direction:column;gap:8px}
  .app-editform .addrow{display:flex;gap:8px;flex-wrap:wrap}
  .app-editform input,.app-editform select,.app-editform textarea{font:inherit;font-size:13px;padding:7px 9px;border:1px solid var(--line);border-radius:8px;background:#fff;outline:none;flex:1;min-width:110px;box-sizing:border-box}
  .app-editform input:focus,.app-editform select:focus,.app-editform textarea:focus{border-color:var(--accent);box-shadow:0 0 0 2px var(--accent-soft)}
  .ae_cancel{font:inherit;font-size:12px;padding:5px 10px;border:1px solid var(--line);border-radius:7px;background:none;cursor:pointer;color:var(--ink)}
  /* To-Do card edit form */
  .editbtn{flex:none;background:none;border:none;cursor:pointer;font-size:16px;color:#c2c7cd;padding:0 2px;line-height:1}
  .editbtn:hover{color:var(--accent)}
  .editform{display:none;margin-top:12px;padding:12px;border:1px dashed var(--line);border-radius:10px;background:#fafbfc}
  .editform .addrow{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px}
  .editform input,.editform textarea{font:inherit;font-size:13px;padding:7px 9px;border:1px solid var(--line);border-radius:8px;background:#fff;outline:none;flex:1;min-width:110px;box-sizing:border-box}
  .editform textarea{width:100%;resize:vertical;min-height:60px;margin-bottom:8px}
  .editform input:focus,.editform textarea:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .addsubmit{font:inherit;font-size:13px;font-weight:600;padding:8px 16px;border:none;border-radius:9px;background:var(--accent);color:#fff;cursor:pointer}
  .addsubmit:hover{background:#1d4ed8}
  .ef_cancel{font:inherit;font-size:13px;padding:8px 14px;border:1px solid var(--line);border-radius:9px;background:none;cursor:pointer;color:var(--ink)}
  .ef_cancel:hover{background:#f1f3f5}
  .empty{text-align:center;color:var(--muted);padding:44px 0;font-size:14px}
  .card-actions{display:flex;align-items:center;gap:5px;flex:none;align-self:flex-start;padding-top:1px}
  .cpbtn{background:none;border:1px solid transparent;border-radius:7px;cursor:pointer;
    font-size:19px;color:#c2c7cd;padding:2px 6px;line-height:1}
  .cpbtn:hover{color:var(--accent);border-color:var(--line);background:#f3f4f6}
  .refreshbtn{font:inherit;font-size:13px;font-weight:550;padding:5px 11px;
    border:1px solid var(--line);border-radius:9px;background:var(--panel);
    color:var(--muted);cursor:pointer;white-space:nowrap}
  .refreshbtn:hover{background:#f3f4f6;color:var(--ink)}
  .app-row-sub{padding:3px 14px 9px;font-size:12.5px;display:flex;gap:12px;flex-wrap:wrap;align-items:center}
  .sub-notes{color:#374151;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .inline-notes{display:block;width:calc(100% - 28px);margin:0 14px 10px;box-sizing:border-box;
    font:inherit;font-size:13.5px;color:#374151;padding:8px 10px;line-height:19px;
    border:1px solid var(--line);border-radius:9px;background:#fcfcfd;
    resize:none;overflow:hidden;height:37px;min-height:37px;outline:none}
  .inline-notes:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-soft)}
  .inline-notes::placeholder{color:#9ca3af}
  /* Compact/Cozy toggle */
  #app-list.compact .app-row{gap:3px;padding:7px 14px 6px}
  #app-list.compact .app-row-top{gap:7px}
  #app-list.compact .app-row-bot{gap:4px}
  #app-list.compact .app-wp-col{display:none}
  #app-list.compact .app-editbtn{display:none}
  #app-list.compact .abtn:not(.todo-btn){display:none}
  .followup-pill{flex:none;display:inline-flex;align-items:center;gap:3px;font-size:11.5px;font-weight:600;padding:2px 9px;border-radius:20px;white-space:nowrap;cursor:default}
  .followup-pill.past{background:#fee2e2;color:#991b1b}
  .followup-pill.soon{background:#fef3c7;color:#92400e}
  .followup-pill.future{background:#dcfce7;color:#166534}
  .ae-label-row{font-size:11px;color:var(--muted);display:flex;gap:8px;margin-bottom:2px}
  .ae-label-row span{flex:1;min-width:0}
  .abtn.todo-btn{color:var(--muted);border-color:var(--line)}
  .abtn.todo-btn.has-todo{color:#166534;border-color:#bbf7d0;background:#f0fdf4}
  .todo-panel{padding:12px 16px 14px;border-top:1px solid var(--line);background:#fafbfc}
  .todo-toolbar{display:flex;align-items:center;gap:8px;margin-bottom:8px}
  .todo-fname{font-size:12px;font-weight:600;color:var(--muted);flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .todo-saved{font-size:12px;color:var(--good);transition:opacity .2s}
  .todo-textarea{width:100%;box-sizing:border-box;font:13.5px/1.6 ui-monospace,Menlo,Monaco,"Cascadia Code",monospace;
    padding:9px 11px;border:1px solid var(--line);border-radius:8px;background:#fff;
    outline:none;resize:vertical;min-height:180px;color:var(--ink)}
  .todo-textarea:focus{border-color:var(--accent);box-shadow:0 0 0 2px var(--accent-soft)}
  .todo-close{font:inherit;font-size:12px;padding:5px 12px;border:1px solid var(--line);border-radius:7px;background:none;cursor:pointer;color:var(--ink)}
  .todo-close:hover{background:#f1f3f5}
  .star{flex:none;background:none;border:none;cursor:pointer;font-size:16px;line-height:1;color:#d1d5db;padding:0 2px}
  .star:hover{color:#f5b301}
  .star.on{color:#f5b301}
  .row-body{}
  .collapsed .row-body{display:none}
  .chevron{background:none;border:none;cursor:pointer;font-size:14px;color:#9ca3af;padding:2px 4px;line-height:1;transition:transform .15s;flex:none;display:inline-flex;align-items:center}
  .collapsed .chevron{transform:rotate(-90deg)}
  .sec-toggle{font:inherit;font-size:12px;padding:3px 10px;border:1px solid var(--line);border-radius:7px;background:none;cursor:pointer;color:var(--muted);white-space:nowrap}
  .sec-toggle:hover{background:#f3f4f6;color:var(--ink)}
</style>
</head>
<body>
<header>
  <div class="bar">
    <div class="titlerow">
      <a href="/" class="back">← Dashboard</a>
      <h1>Application Tracker</h1>
      <button class="undo" id="tracker-undo" disabled title="Undo last change (⌘Z)">↶ Undo</button>
      <button id="tracker-restart" title="Reload server code without leaving the browser" style="font:inherit;font-size:13px;font-weight:550;padding:5px 12px;border:1px solid var(--line);border-radius:9px;background:var(--panel);cursor:pointer;white-space:nowrap;color:var(--ink)">↺ Restart</button>
    </div>
    <div style="margin-top:10px">
      <input type="search" id="appq" placeholder="Search all jobs…" autocomplete="off" style="width:100%;max-width:440px">
    </div>
  </div>
</header>
<main>
  <!-- Drafting -->
  <section id="drafting-section" style="display:none">
    <div class="sec-head">
      <h2>✏️ Drafting</h2><span class="scount" id="drafting-count"></span>
      <button class="sec-toggle" data-section="drafting">Collapse all</button>
    </div>
    <div id="drafting-list"><p class="empty">Loading…</p></div>
  </section>

  <!-- To-Do -->
  <section>
    <div class="sec-head">
      <h2>⭐ To-Do</h2><span class="scount" id="todo-count"></span>
      <button class="sec-toggle" data-section="todo">Collapse all</button>
    </div>
    <div id="todo-list"><p class="empty">Loading…</p></div>
  </section>

  <!-- Applications -->
  <section>
    <div class="sec-head">
      <h2>Applications</h2><span class="scount" id="app-count"></span>
      <button class="refreshbtn" id="refreshbtn" title="Re-read data from disk (picks up changes from the last scheduled run)">↻ Refresh</button>
      <button class="sec-toggle" data-section="apps">Expand all</button>
      <button class="sec-toggle" id="apps-density-toggle">Compact</button>
    </div>
    <div class="filters">
      <div id="sf-btns" style="display:flex;gap:5px;flex-wrap:wrap;align-items:center"></div>
      <select id="ss">
        <option value="date-desc">Applied: newest first</option>
        <option value="date-asc">Applied: oldest first</option>
        <option value="posted-desc">Posted: newest first</option>
        <option value="posted-asc">Posted: oldest first</option>
        <option value="company">Company A–Z</option>
        <option value="status">By status</option>
      </select>
    </div>
    <div id="app-list"><p class="empty">Loading…</p></div>
  </section>
</main>
<script>
let ROLES=[], TODO=[];

function esc(s){const d=document.createElement('div');d.textContent=String(s||'');return d.innerHTML;}
function grow(t){t.style.height='37px';t.style.height=Math.max(37,t.scrollHeight)+'px';}
function scoreClass(n){n=Math.round(Number(n)||0);return 's'+Math.max(0,Math.min(10,n));}
function relDate(s){
  const m=String(s||'').match(/^(\d{4})-(\d{2})-(\d{2})/);
  if(!m)return s||'';
  const d=new Date(+m[1],+m[2]-1,+m[3]);
  const now=new Date();now.setHours(0,0,0,0);
  const diff=Math.round((now-d)/86400000);
  if(diff===0)return 'today';
  if(diff===1)return 'yesterday';
  if(diff<7)return diff+'d ago';
  if(diff<30)return Math.round(diff/7)+'w ago';
  return Math.round(diff/30)+'mo ago';
}
function appliedColor(s){
  const m=String(s||'').match(/^(\d{4})-(\d{2})-(\d{2})/);
  if(!m)return 'var(--muted)';
  const d=new Date(+m[1],+m[2]-1,+m[3]);
  const now=new Date();now.setHours(0,0,0,0);
  const days=Math.round((now-d)/86400000);
  if(days<=7)  return '#16a34a'; // green  — just applied
  if(days<=14) return '#b45309'; // amber  — 1–2 weeks
  if(days<=30) return '#ea580c'; // orange — 2–4 weeks, consider following up
  if(days<=60) return '#dc2626'; // red    — over a month
  return 'var(--muted)';         // grey   — stale
}
function postedColor(s){
  const m=String(s||'').match(/^(\d{4})-(\d{2})-(\d{2})/);
  if(!m)return 'var(--muted)';
  const d=new Date(+m[1],+m[2]-1,+m[3]);
  const now=new Date();now.setHours(0,0,0,0);
  const days=Math.round((now-d)/86400000);
  if(days<=3)  return '#16a34a'; // green  — very fresh
  if(days<=7)  return '#b45309'; // amber  — within a week
  if(days<=14) return '#ea580c'; // orange — 1–2 weeks old
  if(days<=30) return '#dc2626'; // red    — getting stale
  return 'var(--muted)';         // grey   — old
}
function targetBadge(j){
  if(!j.targetEmployer) return '';
  const label=j.targetEmployerName||j.company||'Target employer';
  return `<span title="Target employer: ${esc(label)}" style="display:inline-flex;align-items:center;font-size:11px;font-weight:600;padding:1px 6px;border-radius:4px;background:#fef3c7;color:#92400e;white-space:nowrap">◎ Target</span>`;
}
const INCENTIVE_COLORS={
  equity:  {bg:'#dcfce7',fg:'#166534'},
  bonus:   {bg:'#dbeafe',fg:'#1e40af'},
  ote:     {bg:'#ffedd5',fg:'#9a3412'},
  profit:  {bg:'#ccfbf1',fg:'#115e59'},
  retirement:{bg:'#f1f5f9',fg:'#475569'},
  signing: {bg:'#ede9fe',fg:'#5b21b6'},
};
const INCENTIVE_CAT={
  'RSU':'equity','Stock options':'equity','ESPP':'equity','ESOP':'equity','Equity':'equity',
  'Bonus':'bonus','STI':'bonus',
  'OTE':'ote',
  'Profit share':'profit',
  '401k match':'retirement',
  'Sign-on':'signing',
};
function incentivePills(incentives){
  if(!incentives||!incentives.length) return '';
  return incentives.map(label=>{
    const cat=INCENTIVE_CAT[label]||'bonus';
    const c=INCENTIVE_COLORS[cat]||INCENTIVE_COLORS.bonus;
    return `<span title="${esc(label)}" style="display:inline-flex;align-items:center;font-size:11px;font-weight:600;padding:1px 6px;border-radius:4px;background:${c.bg};color:${c.fg};white-space:nowrap">${esc(label)}</span>`;
  }).join('');
}
function wpInfo(w){
  const v=(w||'').toLowerCase();
  if(v.includes('remote')&&(v.includes('unverif')||v.includes('unv')))return{cls:'Unv',label:'Remote?'};
  if(v.includes('remote'))return{cls:'Remote',label:'Remote'};
  if(v.includes('hybrid'))return{cls:'Hybrid',label:'Hybrid'};
  if(w)return{cls:'Onsite',label:w};
  return null;
}
function wpClass(w){
  if(!w)return'';
  const v=w.toLowerCase();
  if(v.includes('remote')&&(v.includes('unverif')||v.includes('unv')))return'unv';
  if(v.includes('remote'))return'Remote';
  if(v.includes('hybrid'))return'Hybrid';
  return'Onsite';
}
const WP_OPTS=['Remote','Remote — unverified','Hybrid','On-site'];
const ST_OPTS=['new','drafting','applied','interviewing','rejected','not-applying'];
const ST_LABELS={new:'New',drafting:'Drafting',applied:'Applied',interviewing:'Interviewing',rejected:'Rejected','not-applying':'Not Applying'};
function followupPill(dateStr){
  if(!dateStr)return'';
  const m=dateStr.match(/^(\d{4})-(\d{2})-(\d{2})/);
  if(!m)return'';
  const d=new Date(+m[1],+m[2]-1,+m[3]);
  const now=new Date();now.setHours(0,0,0,0);
  const diff=Math.round((d-now)/86400000);
  const mo=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  let cls,label;
  if(diff<0){cls='past';label=`Follow up: ${Math.abs(diff)}d overdue`;}
  else if(diff===0){cls='soon';label='Follow up: today';}
  else if(diff<=7){cls='soon';label=`Follow up in ${diff}d`;}
  else{cls='future';label=`Follow up: ${mo[+m[2]-1]} ${+m[3]}`;}
  return`<span class="followup-pill ${cls}" title="${dateStr}">🗓 ${label}</span>`;
}
const ST_LABEL={drafting:'Drafting',applied:'Applied',interviewing:'Interviewing',rejected:'Rejected','not-applying':'Not Applying'};
const ST_ORDER={interviewing:0,applied:1,drafting:2,rejected:3,'not-applying':4};

// Source badge registry — add entries here as new sources are wired up.
// svg: inline SVG icon (16×16 viewBox), color: brand color for fallback pill.
const SOURCE_META={
  linkedin:{
    label:'LinkedIn',color:'#0a66c2',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#0a66c2"/><path d="M4.5 6.5H3V13H4.5V6.5ZM3.75 5.75A.875.875 0 1 0 3.75 4a.875.875 0 0 0 0 1.75ZM13 9.2C13 7.4 12.1 6.5 10.7 6.5c-.7 0-1.3.3-1.7.8V6.5H7.5V13H9V9.5c0-.8.4-1.3 1.1-1.3.7 0 1.1.5 1.1 1.3V13H12.7V9.2H13Z" fill="#fff"/></svg>`
  },
  indeed:{
    label:'Indeed',color:'#2164f3',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#2164f3"/><text x="4" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">in</text></svg>`
  },
  glassdoor:{
    label:'Glassdoor',color:'#0caa41',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#0caa41"/><text x="3.5" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">G</text></svg>`
  },
  greenhouse:{
    label:'Greenhouse',color:'#3d9970',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#3d9970"/><text x="3" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">gh</text></svg>`
  },
  lever:{
    label:'Lever',color:'#3c4fe0',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#3c4fe0"/><text x="3.5" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">L</text></svg>`
  },
  workday:{
    label:'Workday',color:'#f8783a',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#f8783a"/><text x="3" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">W</text></svg>`
  },
  ashby:{
    label:'Ashby',color:'#6c47ff',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#6c47ff"/><text x="3.5" y="12" font-size="10" font-weight="700" fill="#fff" font-family="sans-serif">A</text></svg>`
  },
  employer:{
    label:'Direct',color:'#64748b',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#64748b"/><path d="M3 11V6l5-3 5 3v5H9.5V8.5h-3V11H3Z" fill="#fff"/></svg>`
  },
  manual:{
    label:'Manual',color:'#94a3b8',
    svg:`<svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg" width="14" height="14"><rect width="16" height="16" rx="3" fill="#94a3b8"/><path d="M4 4h5l3 3v5H4V4Z" fill="#fff" opacity=".9"/><path d="M9 4v3h3" fill="none" stroke="#94a3b8" stroke-width="1"/><path d="M6 9h4M6 11h3" stroke="#64748b" stroke-width="1" stroke-linecap="round"/></svg>`
  },
};
function sourceBadge(source){
  if(!source)return'';
  const m=SOURCE_META[source];
  if(!m)return`<span style="display:inline-flex;align-items:center;gap:3px;font-size:11px;font-weight:600;padding:1px 6px;border-radius:4px;background:#f1f5f9;color:#64748b;white-space:nowrap">${esc(source)}</span>`;
  return`<span title="${m.label}" style="display:inline-flex;align-items:center;line-height:1">${m.svg}</span>`;
}

async function openFile(path){
  try{await fetch('/api/open',{method:'POST',
    headers:{'Content-Type':'application/json'},body:JSON.stringify({path})});}
  catch(e){}
}

// Undo stack
let undoStack=[];
function updateUndoBtn(){const b=document.getElementById('tracker-undo');if(b)b.disabled=undoStack.length===0;}
function pushUndo(fn){undoStack.push(fn);if(undoStack.length>50)undoStack.shift();updateUndoBtn();}
async function doUndo(){const fn=undoStack.pop();updateUndoBtn();if(fn){try{await fn();}catch(err){alert('Undo failed: '+err);}}}
document.getElementById('tracker-undo').addEventListener('click',doUndo);
(function(){
  const btn=document.getElementById('tracker-restart');
  btn.addEventListener('click',async()=>{
    btn.textContent='Restarting…';btn.disabled=true;
    try{await fetch('/api/restart',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'})}catch(e){}
    const t=Date.now();
    while(Date.now()-t<15000){
      await new Promise(r=>setTimeout(r,400));
      try{const r=await fetch('/api/ping');if(r.ok){location.reload(true);return;}}catch(e){}
    }
    btn.textContent='↺ Restart';btn.disabled=false;alert('Server did not come back up.');
  });
})();
document.addEventListener('keydown',e=>{
  if((e.metaKey||e.ctrlKey)&&e.key.toLowerCase()==='z'&&!e.target.closest('input,textarea,select')){
    e.preventDefault();doUndo();
  }
});

async function postStar(jobid,starred){
  try{const r=await fetch('/api/star',{method:'POST',
    headers:{'Content-Type':'application/json'},body:JSON.stringify({jobid,starred})});
    return(await r.json()).ok;}
  catch(e){return false;}
}

function renderTodo(){
  const el=document.getElementById('todo-list');
  const q=document.getElementById('appq').value.trim().toLowerCase();
  const filtered=q?TODO.filter(j=>(j.title+' '+j.company).toLowerCase().includes(q)):TODO;
  document.getElementById('todo-count').textContent='('+filtered.length+')';
  if(!filtered.length){el.innerHTML='<p class="empty">No starred new jobs right now.</p>';return;}
  el.innerHTML=filtered.map(j=>{
    const sc=scoreClass(j.score);
    const wp=wpInfo(j.workplace);
    const titleHtml=j.link
      ?`<a href="${esc(j.link)}" target="_blank" rel="noopener noreferrer">${esc(j.title)} ↗</a>`
      :esc(j.title);
    const todoGmailUrl='https://mail.google.com/mail/u/0/#search/'+encodeURIComponent('"'+j.company+'"');
    const gmailBtnHtml=`<a class="abtn gmail" href="${esc(todoGmailUrl)}" target="_blank" rel="noopener noreferrer" title="Search Gmail for ${esc(j.company)}">Gmail</a>`;
    const resBtnHtml=j.resume?`<button class="abtn res" data-path="${esc(j.resume)}" onclick="openFile(this.dataset.path)" title="Open resume PDF">Resume</button>`:'';
    const clBtnHtml=j.coverLetter?`<button class="abtn cl" data-path="${esc(j.coverLetter)}" onclick="openFile(this.dataset.path)" title="Open cover letter PDF">CL</button>`:'';
    const todoBtnHtml=`<button class="abtn todo-btn${j.todoMd?' has-todo':''}" data-folder="${esc(j.companyFolder||'')}" data-path="${esc(j.todoMd||'')}" data-company="${esc(j.company||'')}" data-role="${esc(j.role||j.title||'')}" title="${j.todoMd?'Edit TODO.md':'New TODO.md'}">📋</button>`;
    const todoPanelHtml=`<div class="todo-panel" data-path="${esc(j.todoMd||'')}" data-folder="${esc(j.companyFolder||'')}" data-company="${esc(j.company||'')}" data-role="${esc(j.role||j.title||'')}" style="display:none"><div class="todo-toolbar"><span class="todo-fname">${j.todoMd?'TODO.md':'New TODO.md'} — ${esc(j.company)}</span><span class="todo-saved"></span><button class="todo-save addsubmit" style="font-size:12px;padding:5px 14px">Save</button><button class="todo-close">Close</button></div><textarea class="todo-textarea" placeholder="# Todos, follow-ups, notes…"></textarea></div>`;
    const unstarBtnHtml=`<button class="star on" data-jobid="${esc(j.jobid)}" title="Starred — click to unstar">★</button>`;
    const st=j.status||'new';
    const wpc=wpClass(j.workplace);
    const wpselHtml=j.workplace?`<span class="dot">·</span><select class="wpsel ${wpc}" data-jobid="${esc(j.jobid)}" title="Set workplace">${WP_OPTS.map(w=>`<option${j.workplace===w?' selected':''}>${esc(w)}</option>`).join('')}</select>`:'';
    const statselHtml=`<select class="statsel st-${st}" data-jobid="${esc(j.jobid)}" title="Set status">${ST_OPTS.map(s=>`<option value="${s}"${st===s?' selected':''}>${esc(ST_LABELS[s])}</option>`).join('')}</select>`;
    return`<div class="card" data-jobid="${esc(j.jobid)}"><div class="chead">
      <div class="score ${sc}">${j.score}</div>
      <div class="ctitle">
        <h3>${titleHtml}</h3>
        <div class="meta">
          ${targetBadge(j)}
          <span>${esc(j.company)}</span>
          ${wpselHtml}
          ${j.salary?`<span class="dot">·</span><span>${esc(j.salary)}</span>`:''}
          ${j.incentives&&j.incentives.length?`<span class="dot">·</span>${incentivePills(j.incentives)}`:''}
          ${j.posted?`<span class="dot">·</span><span title="Posted ${esc(j.posted)}" style="color:${postedColor(j.posted)};font-weight:600">${relDate(j.posted)}</span>`:''}
          ${j.addedAt?`<span class="dot">·</span><span title="Scraped ${esc(j.addedAt)}" style="color:${postedColor(j.addedAt)};font-weight:600">Scraped ${relDate(j.addedAt)}</span>`:''}
        </div>
      </div>
      <div class="card-actions">
        ${statselHtml}
        ${gmailBtnHtml}${resBtnHtml}${clBtnHtml}${todoBtnHtml}
        <button class="editbtn" title="Edit fields">✎</button>
        <button class="cpbtn" title="Save job info to file">⤓</button>
        ${sourceBadge(j.source)}
        ${unstarBtnHtml}
        <button class="chevron" title="Collapse">▾</button>
      </div>
      </div>
      <div class="row-body">
        ${j.why?`<div class="why">${esc(j.why)}</div>`:''}
        <textarea class="inline-notes" data-jobid="${esc(j.jobid)}" placeholder="Notes…">${esc(j.notes||'')}</textarea>
        ${todoPanelHtml}
      </div>
      <div class="editform" style="display:none">
        <div class="addrow">
          <input class="ef_title" value="${esc(j.title)}" placeholder="Title" style="flex:2">
          <input class="ef_company" value="${esc(j.company)}" placeholder="Company">
        </div>
        <div class="addrow">
          <input class="ef_link" value="${esc(j.link||'')}" placeholder="Link URL" style="flex:2">
          <input class="ef_posted" type="date" value="${esc(j.posted||'')}" title="Posted date">
        </div>
        <textarea class="ef_description" placeholder="Full job description (for AI scoring on next scheduled run)…">${esc(j.description||'')}</textarea>
        <div class="addrow">
          <button class="ef_save addsubmit">Save changes</button>
          <button class="ef_cancel">Cancel</button>
        </div>
      </div>
    </div>`;
  }).join('');
  document.getElementById('todo-list').querySelectorAll('.inline-notes').forEach(grow);
}

// Active status filters: empty = show all, otherwise only matching statuses shown
let activeStatusFilters=new Set();

function renderStatusFilterBtns(){
  const statusLabels={interviewing:'Interviewing',applied:'Applied',drafting:'Drafting',rejected:'Rejected','not-applying':'Not Applying'};
  const allCounts={};
  ROLES.forEach(r=>{const s=r.status||'applied';allCounts[s]=(allCounts[s]||0)+1;});
  const container=document.getElementById('sf-btns');
  container.innerHTML=['interviewing','applied','drafting','rejected','not-applying']
    .filter(s=>allCounts[s])
    .map(s=>{
      const active=activeStatusFilters.size===0||activeStatusFilters.has(s);
      return`<button class="sf-btn st-${s}${active?'':' inactive'}" data-status="${s}">${statusLabels[s]} ${allCounts[s]}</button>`;
    }).join('');
  container.querySelectorAll('.sf-btn').forEach(btn=>{
    btn.addEventListener('click',()=>{
      const s=btn.dataset.status;
      if(activeStatusFilters.has(s)){
        activeStatusFilters.delete(s);
        // If nothing selected, show all (clear filter)
        // If still some selected, keep filter
      } else {
        activeStatusFilters.add(s);
      }
      renderApps();
      saveTrackerState();
    });
  });
}

function renderApps(){
  const q=document.getElementById('appq').value.trim().toLowerCase();
  const ss=document.getElementById('ss').value;

  let rows=ROLES.filter(r=>{
    if(activeStatusFilters.size>0&&!activeStatusFilters.has(r.status||'applied'))return false;
    if(q){const t=(r.company+' '+r.role).toLowerCase();if(!t.includes(q))return false;}
    return true;
  });

  rows.sort((a,b)=>{
    if(ss==='status'){
      const sinkOrder={'not-applying':2,'rejected':1};
      const sA=sinkOrder[a.status]??0, sB=sinkOrder[b.status]??0;
      if(sA!==sB)return sA-sB;
      const d=(ST_ORDER[a.status]??99)-(ST_ORDER[b.status]??99);
      if(d!==0)return d;
      return(b.appliedAt||'')>(a.appliedAt||'')?1:-1;
    }
    if(ss==='date-asc')return(a.appliedAt||'')<(b.appliedAt||'')?-1:1;
    if(ss==='posted-desc')return(b.posted||'')>(a.posted||'')?1:-1;
    if(ss==='posted-asc')return(a.posted||'')<(b.posted||'')?-1:1;
    if(ss==='company')return(a.company||'').localeCompare(b.company||'');
    // date-desc (default)
    return(b.appliedAt||'')>(a.appliedAt||'')?1:-1;
  });

  // Update filter buttons to reflect current counts
  renderStatusFilterBtns();
  document.getElementById('app-count').innerHTML=`(${rows.length})`;
  const el=document.getElementById('app-list');
  if(!rows.length){el.innerHTML='<p class="empty">No applications match.</p>';return;}

  el.innerHTML='<div class="app-list">'+rows.map(r=>{
    const roleHtml=r.link
      ?`<a href="${esc(r.link)}" target="_blank" rel="noopener noreferrer">${esc(r.role||'—')} ↗</a>`
      :esc(r.role||'—');
    const gmailBtn=`<a class="abtn gmail" href="${esc(r.gmailSearch)}" target="_blank" rel="noopener noreferrer" title="Search Gmail for ${esc(r.company)}">Gmail</a>`;
    const resBtn=r.resume?`<button class="abtn res" data-path="${esc(r.resume)}" onclick="openFile(this.dataset.path)" title="Open resume PDF">Resume</button>`:'';
    const clBtn=r.coverLetter?`<button class="abtn cl" data-path="${esc(r.coverLetter)}" onclick="openFile(this.dataset.path)" title="Open cover letter PDF">CL</button>`:'';
    const stA=r.status||'applied';
    const statusHtml=r.jobid
      ?`<select class="statsel st-${esc(stA)}" data-jobid="${esc(r.jobid)}">${ST_OPTS.map(s=>`<option value="${s}"${stA===s?' selected':''}>${esc(ST_LABELS[s])}</option>`).join('')}</select>`
      :`<span class="st-badge st-${esc(stA)}">${esc(ST_LABEL[stA]||stA)}</span>`;
    const editBtn=r.jobid?`<button class="app-editbtn" title="Edit">✎</button>`:'<span></span>';
    const expBtn=r.jobid?`<button class="cpbtn" title="Save job info to file">⤓</button>`:'';
    const wpcA=wpClass(r.workplace);
    const wpselA=r.workplace?`<select class="wpsel ${wpcA}" data-jobid="${esc(r.jobid||'')}" title="Set workplace">${WP_OPTS.map(w=>`<option${r.workplace===w?' selected':''}>${esc(w)}</option>`).join('')}</select>`:'';
    const metaHtml=`<div class="app-meta">${r.salary?`<span class="app-salary">${esc(r.salary)}</span>`:''}<span class="app-age app-appliedat" title="Applied ${r.appliedAt||'—'}" style="${r.appliedAt?`color:${appliedColor(r.appliedAt)};font-weight:600`:''}">${r.appliedAt?relDate(r.appliedAt):'—'}</span>${r.addedAt?`<span class="app-age" title="Scraped ${esc(r.addedAt)}" style="color:${postedColor(r.addedAt)};font-weight:600">· Scraped ${relDate(r.addedAt)}</span>`:''}${r.incentives&&r.incentives.length?` ${incentivePills(r.incentives)}`:''}</div>`;
    const todoBtn=`<button class="abtn todo-btn${r.todoMd?' has-todo':''}" title="${r.todoMd?'Edit TODO.md':'New TODO.md'}">📋</button>`;
    const todoPanel=`<div class="todo-panel" data-path="${esc(r.todoMd||'')}" data-folder="${esc(r.companyFolder||'')}" data-company="${esc(r.company||'')}" data-role="${esc(r.role||'')}" style="display:none"><div class="todo-toolbar"><span class="todo-fname">${r.todoMd?'TODO.md':'New TODO.md'} — ${esc(r.company)}</span><span class="todo-saved"></span><button class="todo-save addsubmit" style="font-size:12px;padding:5px 14px">Save</button><button class="todo-close">Close</button></div><textarea class="todo-textarea" placeholder="# Todos, follow-ups, notes…"></textarea></div>`;
    const fuPill=followupPill(r.followUpDate||'');
    const fuRow=fuPill?`<div class="app-row-sub">${fuPill}</div>`:'';
    const notesArea=r.jobid?`<textarea class="inline-notes" data-jobid="${esc(r.jobid)}" placeholder="Notes…">${esc(r.notes||'')}</textarea>`:'';
    const subRowHtml=`${fuRow}${notesArea}`;
    const editForm=r.jobid?`<div class="app-editform" style="display:none">
      <div class="addrow">
        <input class="ae_role" value="${esc(r.role||'')}" placeholder="Role title" style="flex:2">
        <input class="ae_company" value="${esc(r.company||'')}" placeholder="Company">
      </div>
      <div class="addrow">
        <input class="ae_location" value="${esc(r.location||'')}" placeholder="Location (e.g. Remote, Irvine CA)" style="flex:2">
        <select class="ae_workplace" style="flex:1;min-width:110px">
          <option value="">Workplace</option>
          <option value="Remote"${r.workplace==='Remote'?' selected':''}>Remote</option>
          <option value="Hybrid"${r.workplace==='Hybrid'?' selected':''}>Hybrid</option>
          <option value="On-site"${r.workplace==='On-site'?' selected':''}>On-site</option>
        </select>
      </div>
      <div class="addrow">
        <input class="ae_link" value="${esc(r.link||'')}" placeholder="Job posting URL" style="flex:2">
        <input class="ae_salary" value="${esc(r.salary||'')}" placeholder="Salary" style="flex:1">
        <input class="ae_score" type="number" min="0" max="10" value="${r.score!=null?r.score:''}" placeholder="Score 0–10" style="width:90px;flex:none">
      </div>
      <div class="ae-label-row"><span>Posted</span><span>Applied</span><span>Follow-up date</span></div>
      <div class="addrow">
        <input class="ae_posted" type="date" value="${esc(r.posted||'')}" title="Date posted">
        <input class="ae_date" type="date" value="${esc(r.appliedAt||'')}" title="Date applied">
        <input class="ae_followup" type="date" value="${esc(r.followUpDate||'')}" title="Follow-up reminder date">
      </div>
      <textarea class="ae_notes" placeholder="Notes…" style="width:100%;box-sizing:border-box;resize:vertical;min-height:52px;font:inherit;font-size:13px;padding:7px 9px;border:1px solid var(--line);border-radius:8px;background:#fff;outline:none">${esc(r.notes||'')}</textarea>
      <div class="ae_link_section" style="border-top:1px solid var(--line);padding-top:8px;margin-top:4px">
        <div style="font-size:11px;color:var(--muted);margin-bottom:5px;font-weight:500;letter-spacing:.04em">MERGE INTO MASTER JOB RECORD</div>
        <div style="position:relative">
          <input class="ae_link_search" placeholder="Search for the master job record to merge into…" autocomplete="off"
            style="width:100%;box-sizing:border-box;font:inherit;font-size:13px;padding:7px 9px;border:1px solid var(--line);border-radius:8px;background:#fff;outline:none">
          <div class="ae_link_results" style="display:none"></div>
        </div>
        <div class="ae_linked_display" style="display:none;margin-top:6px;font-size:12px;color:var(--ink);align-items:center;gap:6px">
          <span class="ae_linked_label" style="background:var(--accent-soft);color:var(--accent);padding:2px 8px;border-radius:5px;font-weight:500"></span>
          <button class="ae_unlink" style="font:inherit;font-size:11px;padding:2px 7px;border:1px solid var(--line);border-radius:5px;background:none;cursor:pointer;color:var(--muted)">✕ Unlink</button>
        </div>
        <input class="ae_linked_jobid" type="hidden" value="">
      </div>
      <div class="addrow" style="margin-top:8px">
        <button class="ae_save addsubmit" style="font-size:12px;padding:5px 14px">Save</button>
        <button class="ae_cancel">Cancel</button>
      </div>
    </div>`:'';
    const whyHtml=r.why?`<div style="padding:4px 14px 8px;font-size:13px;color:#4b5563;line-height:1.5;border-bottom:1px solid var(--line)">${esc(r.why)}</div>`:'';
    return`<div class="app-row-wrap collapsed" data-jobid="${esc(r.jobid||'')}" data-status="${esc(stA)}">
      <div class="app-row">
        <div class="app-row-top">
          <span class="app-score ${r.score!=null?scoreClass(r.score):'s0'}" title="AI score"><b>${r.score!=null?r.score:'—'}</b></span>
          <span class="app-company" title="${esc(r.company)}">${esc(r.company)}</span>
          <span class="app-role">${roleHtml}</span>
          ${metaHtml}
        </div>
        <div class="app-row-bot">
          ${statusHtml}
          <span class="app-wp-col">${wpselA}</span>
          ${sourceBadge(r.source)}${targetBadge(r)}
          <span class="app-actions">${gmailBtn}${resBtn}${clBtn}${todoBtn}${editBtn}${expBtn}<button class="chevron" title="Toggle details">▾</button></span>
        </div>
      </div>
      <div class="row-body">
        ${whyHtml}
        ${subRowHtml}
      </div>
      ${todoPanel}
      ${editForm}
    </div>`;
  }).join('')+'</div>';
  document.getElementById('app-list').querySelectorAll('.inline-notes').forEach(grow);
}

function renderDrafting(){
  const q=document.getElementById('appq').value.trim().toLowerCase();
  const rows=ROLES.filter(r=>r.status==='drafting'&&(!q||(r.company+' '+r.role).toLowerCase().includes(q)));
  const section=document.getElementById('drafting-section');
  section.style.display=rows.length?'':'none';
  document.getElementById('drafting-count').textContent='('+rows.length+')';
  if(!rows.length){document.getElementById('drafting-list').innerHTML='';return;}
  document.getElementById('drafting-list').innerHTML='<div class="app-list">'+rows.map(r=>{
    const roleHtml=r.link
      ?`<a href="${esc(r.link)}" target="_blank" rel="noopener noreferrer">${esc(r.role||'—')} ↗</a>`
      :esc(r.role||'—');
    const gmailBtn=`<a class="abtn gmail" href="${esc(r.gmailSearch)}" target="_blank" rel="noopener noreferrer" title="Search Gmail for ${esc(r.company)}">Gmail</a>`;
    const resBtn=r.resume?`<button class="abtn res" data-path="${esc(r.resume)}" onclick="openFile(this.dataset.path)" title="Open resume PDF">Resume</button>`:'';
    const clBtn=r.coverLetter?`<button class="abtn cl" data-path="${esc(r.coverLetter)}" onclick="openFile(this.dataset.path)" title="Open cover letter PDF">CL</button>`:'';
    const todoBtn=`<button class="abtn todo-btn${r.todoMd?' has-todo':''}" title="${r.todoMd?'Edit TODO.md':'New TODO.md'}">📋</button>`;
    const todoPanel=`<div class="todo-panel" data-path="${esc(r.todoMd||'')}" data-folder="${esc(r.companyFolder||'')}" data-company="${esc(r.company||'')}" data-role="${esc(r.role||'')}" style="display:none"><div class="todo-toolbar"><span class="todo-fname">${r.todoMd?'TODO.md':'New TODO.md'} — ${esc(r.company)}</span><span class="todo-saved"></span><button class="todo-save addsubmit" style="font-size:12px;padding:5px 14px">Save</button><button class="todo-close">Close</button></div><textarea class="todo-textarea" placeholder="# Todos, follow-ups, notes…"></textarea></div>`;
    const wpcD=wpClass(r.workplace);
    const wpselD=r.workplace?`<select class="wpsel ${wpcD}" data-jobid="${esc(r.jobid||'')}" title="Set workplace">${WP_OPTS.map(w=>`<option${r.workplace===w?' selected':''}>${esc(w)}</option>`).join('')}</select>`:'';
    const metaHtmlD=`<div class="app-meta">${r.salary?`<span class="app-salary">${esc(r.salary)}</span>`:''}<span class="app-age" title="Posted ${r.posted||'—'}">${r.posted?relDate(r.posted):'—'}</span></div>`;
    const statusHtml=r.jobid
      ?`<select class="statsel st-drafting" data-jobid="${esc(r.jobid)}">${ST_OPTS.map(s=>`<option value="${s}"${s==='drafting'?' selected':''}>${esc(ST_LABELS[s])}</option>`).join('')}</select>`
      :`<span class="st-badge st-drafting">Drafting</span>`;
    const editBtn=r.jobid?`<button class="app-editbtn" title="Edit">✎</button>`:'<span></span>';
    const expBtn=r.jobid?`<button class="cpbtn" title="Save job info to file">⤓</button>`:'';
    const notesAreaD=r.jobid?`<textarea class="inline-notes" data-jobid="${esc(r.jobid)}" placeholder="Notes…">${esc(r.notes||'')}` + '</textarea>':'';
    const subRowHtml=notesAreaD||'';
    const whyHtmlD=r.why?`<div style="padding:4px 14px 8px;font-size:13px;color:#4b5563;line-height:1.5;border-bottom:1px solid var(--line)">${esc(r.why)}</div>`:'';
    return`<div class="app-row-wrap" data-jobid="${esc(r.jobid||'')}" data-status="drafting">
      <div class="app-row">
        <div class="app-row-top">
          <span class="app-score ${r.score!=null?scoreClass(r.score):'s0'}" title="AI score"><b>${r.score!=null?r.score:'—'}</b></span>
          <span class="app-company" title="${esc(r.company)}">${esc(r.company)}</span>
          <span class="app-role">${roleHtml}</span>
          ${metaHtmlD}
        </div>
        <div class="app-row-bot">
          ${statusHtml}
          <span class="app-wp-col">${wpselD}</span>
          ${sourceBadge(r.source)}${targetBadge(r)}
          <span class="app-actions">${gmailBtn}${resBtn}${clBtn}${todoBtn}${editBtn}${expBtn}<button class="chevron" title="Toggle details">▾</button></span>
        </div>
      </div>
      <div class="row-body">
        ${whyHtmlD}
        ${subRowHtml}
      </div>
      ${todoPanel}
    </div>`;
  }).join('')+'</div>';
  document.getElementById('drafting-list').querySelectorAll('.inline-notes').forEach(grow);
}

// Section collapse/expand toggle
document.querySelectorAll('.sec-toggle[data-section]').forEach(btn=>{
  btn.addEventListener('click',()=>{
    const sec=btn.dataset.section;
    const list=document.getElementById(sec==='todo'?'todo-list':sec==='drafting'?'drafting-list':'app-list');
    const items=list.querySelectorAll('.app-row-wrap,.card');
    const allCollapsed=[...items].every(el=>el.classList.contains('collapsed'));
    items.forEach(el=>{
      if(allCollapsed)el.classList.remove('collapsed');
      else el.classList.add('collapsed');
    });
    btn.textContent=allCollapsed?'Collapse all':'Expand all';
  });
});
// Compact / Cozy density toggle
(function(){
  const btn=document.getElementById('apps-density-toggle');
  const list=document.getElementById('app-list');
  function applyDensity(compact){
    list.classList.toggle('compact',compact);
    btn.textContent=compact?'Cozy':'Compact';
  }
  applyDensity(localStorage.getItem('apps_compact')==='1');
  btn.addEventListener('click',()=>{
    const nowCompact=!list.classList.contains('compact');
    applyDensity(nowCompact);
    localStorage.setItem('apps_compact',nowCompact?'1':'');
  });
})();

// Chevron click — delegate from each list container
function attachChevronHandler(listId){
  document.getElementById(listId).addEventListener('click',e=>{
    const chev=e.target.closest('.chevron');
    if(!chev)return;
    const wrap=chev.closest('.app-row-wrap,.card');
    if(!wrap)return;
    wrap.classList.toggle('collapsed');
    // Update section toggle button label
    const list=wrap.closest('[id$="-list"],[id="app-list"]');
    const sec=list?.id==='todo-list'?'todo':list?.id==='drafting-list'?'drafting':'apps';
    const btn=document.querySelector(`.sec-toggle[data-section="${sec}"]`);
    if(btn){
      const items=list.querySelectorAll('.app-row-wrap,.card');
      const allC=[...items].every(el=>el.classList.contains('collapsed'));
      const anyC=[...items].some(el=>el.classList.contains('collapsed'));
      btn.textContent=allC?'Expand all':anyC?'Collapse all':'Collapse all';
    }
  });
}
attachChevronHandler('todo-list');
attachChevronHandler('drafting-list');
attachChevronHandler('app-list');

// Save-to-file (export details) button for the Applications and Drafting rows
function attachExportHandler(listId){
  document.getElementById(listId).addEventListener('click',async e=>{
    const btn=e.target.closest('.cpbtn');
    if(!btn)return;
    const jobid=btn.closest('.app-row-wrap')?.dataset.jobid;
    if(!jobid)return;
    btn.style.pointerEvents='none';
    try{
      const r=await fetch('/api/savejobmd',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({jobid})});
      const d=await r.json();
      if(d.ok){btn.textContent='✓';setTimeout(()=>{btn.textContent='⤓';},2000);}
      else{alert('Save failed — '+(d.error||'unknown error'));}
    }catch(ex){alert('Save failed — '+ex);}
    finally{btn.style.pointerEvents='';}
  });
}
attachExportHandler('app-list');
attachExportHandler('drafting-list');

async function load(){
  const[jobsData,appData]=await Promise.all([
    fetch('/api/jobs').then(r=>r.json()).catch(()=>({jobs:[]})),
    fetch('/api/applied').then(r=>r.json()).catch(()=>({roles:[]}))
  ]);
  TODO=(jobsData.jobs||[])
    .filter(j=>j.starred&&j.status==='new'&&!j.dismissed)
    .sort((a,b)=>(b.score-a.score)||((b.posted||'')>(a.posted||'')?1:-1));
  ROLES=appData.roles||[];
  renderDrafting();
  renderTodo();
  renderApps();
}

function saveTrackerState(){
  localStorage.setItem('trackerState',JSON.stringify({
    q:document.getElementById('appq').value,
    sf:[...activeStatusFilters],
    ss:document.getElementById('ss').value,
  }));
}
function restoreTrackerState(){
  try{
    const s=JSON.parse(localStorage.getItem('trackerState')||'{}');
    if(s.q)document.getElementById('appq').value=s.q;
    if(Array.isArray(s.sf))activeStatusFilters=new Set(s.sf);
    if(s.ss)document.getElementById('ss').value=s.ss;
  }catch(e){}
}
restoreTrackerState();
document.getElementById('appq').addEventListener('input',()=>{renderDrafting();renderTodo();renderApps();saveTrackerState();});
document.getElementById('ss').addEventListener('change',()=>{renderApps();saveTrackerState();});

// To-Do edit button + inline form
document.getElementById('todo-list').addEventListener('click',async e=>{
  // Star toggle — unstar removes from To-Do list
  if(e.target.closest('.star')){
    const btn=e.target.closest('.star');
    const jobid=btn.dataset.jobid;
    const isOn=btn.classList.contains('on');
    if(await postStar(jobid,!isOn)){
      pushUndo(async()=>{await postStar(jobid,isOn);load();});
      if(isOn){TODO=TODO.filter(x=>x.jobid!==jobid);renderTodo();}
      else{btn.classList.add('on');btn.title='Starred — click to unstar';btn.textContent='★';}
    }
    return;
  }
  // Edit button toggle
  if(e.target.closest('.editbtn')){
    const card=e.target.closest('.card');
    const form=card.querySelector('.editform');
    form.style.display=form.style.display==='block'?'none':'block';
    return;
  }
  // Cancel
  if(e.target.closest('.ef_cancel')){
    e.target.closest('.editform').style.display='none';
    return;
  }
  // Save
  if(e.target.closest('.ef_save')){
    const card=e.target.closest('.card');
    const jobid=card.dataset.jobid;
    const form=card.querySelector('.editform');
    const body={
      jobid,
      title:form.querySelector('.ef_title').value.trim(),
      company:form.querySelector('.ef_company').value.trim(),
      link:form.querySelector('.ef_link').value.trim(),
      posted:form.querySelector('.ef_posted').value,
      description:form.querySelector('.ef_description').value,
    };
    try{
      const res=await fetch('/api/editjob',{method:'POST',
        headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
      const d=await res.json();
      if(d.ok){
        const j=TODO.find(x=>x.jobid===jobid);
        if(j){j.title=body.title;j.company=body.company;j.link=body.link;
               j.posted=body.posted;j.description=body.description;}
        form.style.display='none';
        renderTodo();
      }else{alert('Could not save changes.');}
    }catch(_){alert('Could not save changes.');}
    return;
  }
  // TODO button — open/close inline editor
  if(e.target.closest('.todo-btn')){
    const wrap=e.target.closest('.app-row-wrap,.card');
    const panel=wrap.querySelector('.todo-panel');
    if(panel.style.display!=='none'){panel.style.display='none';return;}
    panel.style.display='block';
    const ta=panel.querySelector('.todo-textarea');
    if(ta._loaded)return;
    const path=panel.dataset.path;
    if(!path){ta._loaded=true;ta.focus();return;}
    try{
      const d=await fetch('/api/todo?path='+encodeURIComponent(path)).then(r=>r.json());
      ta.value=d.content||'';
    }catch(ex){ta.value='';}
    ta._loaded=true;ta.focus();return;
  }
  // TODO save
  if(e.target.closest('.todo-save')){
    const panel=e.target.closest('.todo-panel');
    const ta=panel.querySelector('.todo-textarea');
    const saved=panel.querySelector('.todo-saved');
    const path=panel.dataset.path;
    const folder=panel.dataset.folder;
    const company=panel.dataset.company||'';
    const role=panel.dataset.role||'';
    try{
      const res=await fetch('/api/todo',{method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({path:path||'',folder:path?'':folder,company:path||folder?'':company,role:path||folder?'':role,content:ta.value})});
      const d=await res.json();
      if(d.ok){
        if(!path){
          panel.dataset.path=d.path;
          if(d.folder)panel.dataset.folder=d.folder;
          const wrap=panel.closest('.app-row-wrap,.card');
          const jobid=wrap?.dataset.jobid;
          const entry=[...ROLES,...TODO].find(r=>r.jobid===jobid);
          if(entry){entry.todoMd=d.path;if(!entry.companyFolder)entry.companyFolder=d.folder;}
          const btn=wrap?.querySelector('.todo-btn');
          if(btn){btn.classList.add('has-todo');btn.title='Edit TODO.md';}
          panel.querySelector('.todo-fname').textContent='TODO.md — '+(entry?.company||'');
        }
        saved.textContent='Saved ✓';
        setTimeout(()=>{saved.textContent='';},2000);
      }else{alert('Could not save: '+(d.error||'unknown error'));}
    }catch(ex){alert('Could not save.');}
    return;
  }
  // TODO close
  if(e.target.closest('.todo-close')){
    e.target.closest('.todo-panel').style.display='none';
    return;
  }
  // Save to file button
  const btn=e.target.closest('.cpbtn');
  if(!btn)return;
  const jobid=btn.closest('.card').dataset.jobid;
  const j=TODO.find(x=>x.jobid===jobid);
  if(!j)return;
  try{
    const r=await fetch('/api/savejobmd',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({jobid:j.jobid})});
    const d=await r.json();
    if(d.ok){btn.textContent='✓';setTimeout(()=>{btn.textContent='⤓';},2000);}
    else{alert('Save failed — '+(d.error||'unknown error'));}
  }catch(e){alert('Save failed — '+e);}
});

// Refresh button — re-reads data.json and applied.json from disk
document.getElementById('refreshbtn').addEventListener('click',async()=>{
  const btn=document.getElementById('refreshbtn');
  btn.textContent='↻ Refreshing…'; btn.disabled=true;
  await load();
  btn.textContent='↻ Refresh'; btn.disabled=false;
});

// Shared workplace handler
async function handleWpChange(sel,entry){
  const nw=sel.value,prev=entry?.workplace||'';
  const jobid=sel.dataset.jobid;
  const ok=await fetch('/api/editjob',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({jobid,workplace:nw})}).then(r=>r.json()).then(d=>d.ok).catch(()=>false);
  if(ok){
    pushUndo(async()=>{
      await fetch('/api/editjob',{method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({jobid,workplace:prev})});
      if(entry)entry.workplace=prev;
      renderTodo();renderDrafting();renderApps();
    });
    sel.className='wpsel '+wpClass(nw);if(entry)entry.workplace=nw;
  }else{sel.value=prev;alert('Could not update workplace.');}
}

// To-Do — status and workplace
document.getElementById('todo-list').addEventListener('change',async e=>{
  const sel=e.target;
  const jobid=sel.dataset.jobid;
  if(sel.classList.contains('wpsel')){
    await handleWpChange(sel,TODO.find(x=>x.jobid===jobid));return;
  }
  if(!sel.classList.contains('statsel'))return;
  const newStatus=sel.value,prevStatus=sel.dataset.prev||'new';
  sel.dataset.prev=newStatus;
  sel.className='statsel st-'+newStatus;
  await fetch('/api/status',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({jobid,status:newStatus})}).catch(()=>{});
  pushUndo(async()=>{
    await fetch('/api/status',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jobid,status:prevStatus})});
    load();
  });
  load();
});

// Inline notes — grow on input
document.addEventListener('input',e=>{
  if(e.target.classList.contains('inline-notes'))grow(e.target);
});
// Inline notes — auto-save on blur
document.addEventListener('focusout',async e=>{
  const ta=e.target;
  if(!ta.classList.contains('inline-notes'))return;
  const jobid=ta.dataset.jobid;
  if(!jobid)return;
  const entry=[...ROLES,...TODO].find(r=>r.jobid===jobid);
  const newNotes=ta.value,prevNotes=entry?.notes||'';
  if(entry&&prevNotes===newNotes)return; // no change
  const ok=await fetch('/api/editjob',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({jobid,notes:newNotes})}).then(r=>r.json()).then(d=>d.ok).catch(()=>false);
  if(ok&&entry){
    pushUndo(async()=>{
      await fetch('/api/editjob',{method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({jobid,notes:prevNotes})});
      entry.notes=prevNotes;
      renderTodo();renderDrafting();renderApps();
    });
    entry.notes=newNotes;
  }
});

// Drafting — status and workplace
document.getElementById('drafting-list').addEventListener('change',async e=>{
  const sel=e.target;
  const wrap=sel.closest('.app-row-wrap');
  const jobid=wrap?.dataset.jobid||sel.dataset.jobid;
  const entry=ROLES.find(r=>r.jobid===jobid);
  if(sel.classList.contains('wpsel')){await handleWpChange(sel,entry);return;}
  if(!sel.classList.contains('statsel'))return;
  const newStatus=sel.value,prevStatus=entry?.status||'drafting';
  sel.className='statsel st-'+newStatus;
  const res=await fetch('/api/status',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({jobid,status:newStatus})}).catch(()=>null);
  const d=res?await res.json().catch(()=>({})):{};
  if(entry){entry.status=newStatus;if(d.appliedAt&&!entry.appliedAt)entry.appliedAt=d.appliedAt;}
  pushUndo(async()=>{
    await fetch('/api/status',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jobid,status:prevStatus})});
    if(entry)entry.status=prevStatus;
    renderDrafting();renderApps();
  });
  if(newStatus==='new'){load();return;}
  renderDrafting();renderApps();
});

// Applications — status and workplace
document.getElementById('app-list').addEventListener('change',async e=>{
  const sel=e.target;
  const wrap=sel.closest('.app-row-wrap');
  const jobid=wrap?.dataset.jobid||sel.dataset.jobid;
  const entry=ROLES.find(r=>r.jobid===jobid);
  if(sel.classList.contains('wpsel')){await handleWpChange(sel,entry);return;}
  if(!sel.classList.contains('statsel'))return;
  const newStatus=sel.value;
  const existingAppliedAt=entry?.appliedAt||'';
  const prevStatus=entry?.status||'applied';
  sel.className='statsel st-'+newStatus;
  if(wrap)wrap.dataset.status=newStatus;
  const res=await fetch('/api/status',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({jobid,status:newStatus,appliedAt:existingAppliedAt||undefined})}).catch(()=>null);
  const d=res?await res.json().catch(()=>({})):{};
  if(entry){entry.status=newStatus;if(d.appliedAt&&!existingAppliedAt)entry.appliedAt=d.appliedAt;}
  pushUndo(async()=>{
    await fetch('/api/status',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jobid,status:prevStatus})});
    if(entry)entry.status=prevStatus;
    renderDrafting();renderApps();
  });
  if(newStatus==='new'){load();return;}
  if(wrap){const span=wrap.querySelector('.app-appliedat');if(span&&entry?.appliedAt){span.textContent=relDate(entry.appliedAt);span.style.color=appliedColor(entry.appliedAt);span.style.fontWeight='600';}}
  renderDrafting();renderApps();
});

// App row edit button + inline form
document.getElementById('app-list').addEventListener('click',async e=>{
  // Edit button toggle
  if(e.target.closest('.app-editbtn')){
    const wrap=e.target.closest('.app-row-wrap');
    const form=wrap.querySelector('.app-editform');
    form.style.display=form.style.display==='block'?'none':'block';
    if(form.style.display==='block') initLinkSearch(form);
    return;
  }
  // Cancel
  if(e.target.closest('.ae_cancel')){
    e.target.closest('.app-editform').style.display='none';
    return;
  }
  // TODO button — open/close inline editor
  if(e.target.closest('.todo-btn')){
    const wrap=e.target.closest('.app-row-wrap,.card');
    const panel=wrap.querySelector('.todo-panel');
    if(panel.style.display!=='none'){panel.style.display='none';return;}
    panel.style.display='block';
    const ta=panel.querySelector('.todo-textarea');
    if(ta._loaded)return;
    const path=panel.dataset.path;
    if(!path){ta._loaded=true;ta.focus();return;}
    try{
      const d=await fetch('/api/todo?path='+encodeURIComponent(path)).then(r=>r.json());
      ta.value=d.content||'';
    }catch(ex){ta.value='';}
    ta._loaded=true;
    ta.focus();
    return;
  }
  // TODO save
  if(e.target.closest('.todo-save')){
    const panel=e.target.closest('.todo-panel');
    const ta=panel.querySelector('.todo-textarea');
    const saved=panel.querySelector('.todo-saved');
    const path=panel.dataset.path;
    const folder=panel.dataset.folder;
    const company=panel.dataset.company||'';
    const role=panel.dataset.role||'';
    try{
      const res=await fetch('/api/todo',{method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({path:path||'',folder:path?'':folder,company:path||folder?'':company,role:path||folder?'':role,content:ta.value})});
      const d=await res.json();
      if(d.ok){
        if(!path){
          panel.dataset.path=d.path;
          if(d.folder)panel.dataset.folder=d.folder;
          const wrap=panel.closest('.app-row-wrap,.card');
          const jobid=wrap?.dataset.jobid;
          const entry=[...ROLES,...TODO].find(r=>(r.jobid||r.jobid)===jobid);
          if(entry){entry.todoMd=d.path;if(!entry.companyFolder)entry.companyFolder=d.folder;}
          const btn=wrap?.querySelector('.todo-btn');
          if(btn){btn.classList.add('has-todo');btn.title='Edit TODO.md';}
          panel.querySelector('.todo-fname').textContent='TODO.md — '+(entry?.company||'');
        }
        saved.textContent='Saved ✓';
        setTimeout(()=>{saved.textContent='';},2000);
      }else{alert('Could not save: '+(d.error||'unknown error'));}
    }catch(ex){alert('Could not save.');}
    return;
  }
  // TODO close
  if(e.target.closest('.todo-close')){
    e.target.closest('.todo-panel').style.display='none';
    return;
  }
  // Save
  if(e.target.closest('.ae_save')){
    const wrap=e.target.closest('.app-row-wrap');
    const jobid=wrap.dataset.jobid;
    const form=wrap.querySelector('.app-editform');
    const linkedJobId=form.querySelector('.ae_linked_jobid')?.value||'';
    const body={
      jobid,
      title:form.querySelector('.ae_role').value.trim(),
      company:form.querySelector('.ae_company').value.trim(),
      location:form.querySelector('.ae_location').value.trim(),
      workplace:form.querySelector('.ae_workplace').value,
      link:form.querySelector('.ae_link').value.trim(),
      salary:form.querySelector('.ae_salary').value.trim(),
      score:form.querySelector('.ae_score').value.trim(),
      posted:form.querySelector('.ae_posted').value,
      appliedAt:form.querySelector('.ae_date').value,
      followUpDate:form.querySelector('.ae_followup').value,
      notes:form.querySelector('.ae_notes').value,
      ...(linkedJobId&&{linkedJobId}),
    };
    try{
      const res=await fetch('/api/editjob',{method:'POST',
        headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
      const d=await res.json();
      if(d.ok){
        const entry=ROLES.find(r=>r.jobid===jobid);
        if(entry){
          entry.role=body.title;entry.company=body.company;
          entry.location=body.location;entry.workplace=body.workplace;
          entry.link=body.link;entry.salary=body.salary;
          entry.score=body.score?parseInt(body.score):entry.score;
          entry.posted=body.posted;entry.appliedAt=body.appliedAt;
          entry.followUpDate=body.followUpDate;
          entry.notes=body.notes;
          if(d.jobid&&d.jobid!==jobid)entry.jobid=d.jobid;
        }
        form.style.display='none';
        renderApps();
      }else{alert('Could not save changes.');}
    }catch(_){alert('Could not save changes.');}
    return;
  }
});

// Single shared dropdown appended to body — escapes all overflow clipping.
const _linkDropdown=(()=>{
  const el=document.createElement('div');
  el.style.cssText='display:none;position:fixed;z-index:9999;background:#fff;border:1px solid #e6e8eb;border-radius:8px;box-shadow:0 4px 12px rgba(0,0,0,.15);max-height:220px;overflow-y:auto;min-width:240px';
  document.body.appendChild(el);
  return el;
})();
let _linkDropdownOwner=null;
function _repositionLinkDropdown(){
  if(!_linkDropdownOwner||_linkDropdown.style.display==='none') return;
  const rect=_linkDropdownOwner.getBoundingClientRect();
  _linkDropdown.style.top=(rect.bottom+4)+'px';
  _linkDropdown.style.left=rect.left+'px';
  _linkDropdown.style.width=rect.width+'px';
}
document.addEventListener('scroll',_repositionLinkDropdown,{capture:true,passive:true});

function initLinkSearch(form){
  const input=form.querySelector('.ae_link_search');
  const display=form.querySelector('.ae_linked_display');
  const label=form.querySelector('.ae_linked_label');
  const hidden=form.querySelector('.ae_linked_jobid');
  const unlinkBtn=form.querySelector('.ae_unlink');
  if(!input||input._linked_init) return;
  input._linked_init=true;
  let debounce;
  function positionDropdown(){
    const rect=input.getBoundingClientRect();
    _linkDropdown.style.top=(rect.bottom+4)+'px';
    _linkDropdown.style.left=rect.left+'px';
    _linkDropdown.style.width=rect.width+'px';
  }
  function hideDropdown(){
    if(_linkDropdownOwner===input){
      _linkDropdown.style.display='none';
      _linkDropdownOwner=null;
    }
  }
  function selectJob(jid,company,title){
    hidden.value=jid;
    label.textContent=company+(title?(' — '+title):'');
    display.style.display='flex';
    input.value='';
    hideDropdown();
  }
  unlinkBtn.addEventListener('click',()=>{
    hidden.value='';
    display.style.display='none';
    input.value='';
  });
  input.addEventListener('input',()=>{
    clearTimeout(debounce);
    const q=input.value.trim();
    if(!q){hideDropdown();return;}
    debounce=setTimeout(async()=>{
      try{
        const d=await fetch('/api/jobs/search?q='+encodeURIComponent(q)).then(r=>r.json());
        if(!d.results.length){_linkDropdown.innerHTML='<div style="padding:8px 12px;font-size:12px;color:#6b7280">No matches</div>';}
        else{_linkDropdown.innerHTML=d.results.map(j=>`<div class="ae_link_result" data-jid="${esc(j.jobid)}" data-company="${esc(j.company)}" data-title="${esc(j.title)}"
          style="padding:7px 12px;cursor:pointer;font-size:13px;border-bottom:1px solid #e6e8eb;display:flex;gap:6px;align-items:baseline">
          <span style="font-weight:500">${esc(j.company)}</span>
          <span style="color:#6b7280;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${esc(j.title)}</span>
        </div>`).join('');}
        _linkDropdownOwner=input;
        positionDropdown();
        _linkDropdown.style.display='block';
      }catch(e){}
    },200);
  });
  _linkDropdown.addEventListener('click',e=>{
    const row=e.target.closest('.ae_link_result');
    if(!row) return;
    selectJob(row.dataset.jid,row.dataset.company,row.dataset.title);
  });
  document.addEventListener('click',e=>{
    if(_linkDropdownOwner===input&&!form.contains(e.target)&&e.target!==_linkDropdown&&!_linkDropdown.contains(e.target)) hideDropdown();
  },{capture:true});
}

// Cmd+S / Ctrl+S inside a TODO textarea → save (covers todo-list, app-list, drafting-list)
document.addEventListener('keydown',async e=>{
  if(!e.target.classList.contains('todo-textarea'))return;
  if(!(e.metaKey||e.ctrlKey)||e.key!=='s')return;
  e.preventDefault();
  e.target.closest('.todo-panel').querySelector('.todo-save').click();
});

load();
</script>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _send(self, code, body, ctype="application/json"):
        data = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", ctype + "; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store, must-revalidate")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        if self.path == "/api/ping":
            self._send(200, json.dumps({"ok": True}))
            return
        if self.path == "/" or self.path.startswith("/index"):
            self._send(200, PAGE, "text/html")
        elif self.path == "/api/jobs":
            try:
                # Prefer data.json (state-first); fall back to xlsx if missing.
                jobs = read_jobs_from_data()
                # Enrich starred jobs with company-folder / PDF info (small subset, so disk I/O is fine)
                for _j in jobs:
                    if _j.get("starred"):
                        _pdfs = _get_company_pdfs(_j.get("company") or "", _j.get("title") or "")
                        _folder = _pdfs.get("folder")
                        _j["companyFolder"] = _folder
                        _j["todoMd"] = os.path.join(_folder, "TODO.md") if _folder and os.path.isfile(os.path.join(_folder, "TODO.md")) else None
                        _j["resume"] = _pdfs.get("resume")
                        _j["coverLetter"] = _pdfs.get("coverLetter")
                try:
                    size_mb = round(os.path.getsize(XLSX) / 1048576, 2)
                except Exception:
                    size_mb = 0
                payload = {"jobs": jobs, "built": read_last_built(), "sizeMB": size_mb, "order": load_order()}
                self._send(200, json.dumps(payload))
            except Exception as e:
                self._send(500, json.dumps({"error": str(e)}))
        elif self.path == "/tracker":
            self._send(200, TRACKER_PAGE, "text/html")
        elif self.path == "/api/applied":
            self._send(200, api_applied())
        elif self.path.startswith("/api/jobs/search"):
            from urllib.parse import urlparse, parse_qs
            qs = parse_qs(urlparse(self.path).query)
            q = (qs.get("q", [""])[0] or "").strip().lower()
            results = []
            if q:
                try:
                    jobs = read_jobs_from_data()
                    for j in jobs:
                        company = (j.get("company") or "").lower()
                        title = (j.get("title") or "").lower()
                        if q in company or q in title:
                            results.append({
                                "jobid": j.get("jobid") or j.get("id") or "",
                                "company": j.get("company") or "",
                                "title": j.get("title") or "",
                            })
                        if len(results) >= 10:
                            break
                except Exception:
                    pass
            self._send(200, json.dumps({"results": results}))
        elif self.path.startswith("/api/todo"):
            from urllib.parse import urlparse, parse_qs
            qs = parse_qs(urlparse(self.path).query)
            path = (qs.get("path", [""])[0] or "").strip()
            if not path or not RESUME_FOLDER or not os.path.abspath(path).startswith(os.path.abspath(RESUME_FOLDER)):
                self._send(403, json.dumps({"ok": False, "error": "path not allowed"}))
            elif not os.path.isfile(path):
                self._send(200, json.dumps({"ok": True, "content": "", "exists": False}))
            else:
                try:
                    with open(path, encoding="utf-8") as fh:
                        content = fh.read()
                    self._send(200, json.dumps({"ok": True, "content": content, "exists": True}))
                except Exception as e:
                    self._send(500, json.dumps({"ok": False, "error": str(e)}))
        else:
            self._send(404, json.dumps({"error": "not found"}))

    def do_POST(self):
        try:
            n = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(n) or b"{}")
        except Exception as e:
            self._send(400, json.dumps({"ok": False, "error": str(e)}))
            return
        try:
            if self.path == "/api/update":
                ok = update_job(body.get("jobid", ""), bool(body.get("dismissed")),
                                body.get("notes", ""), body.get("salary", None))
                self._send(200 if ok else 404, json.dumps({"ok": ok}))
            elif self.path == "/api/bulk":
                count = update_jobs_bulk(body.get("jobids", []), bool(body.get("dismissed")))
                self._send(200, json.dumps({"ok": True, "count": count}))
            elif self.path == "/api/star":
                set_star(body.get("jobid", ""), bool(body.get("starred")))
                self._send(200, json.dumps({"ok": True}))
            elif self.path == "/api/status":
                ok = set_status(body.get("jobid", ""), body.get("status", ""), body.get("appliedAt"))
                entry = scout_state.get("jobs", {}).get(str(body.get("jobid", "")), {})
                self._send(200 if ok else 400, json.dumps({"ok": ok, "appliedAt": entry.get("appliedAt", "")}))
            elif self.path == "/api/order":
                save_order(body.get("order", []))
                self._send(200, json.dumps({"ok": True}))
            elif self.path == "/api/export":
                self._send(200, export_csv(body.get("jobids", [])), ctype="text/csv")
            elif self.path == "/api/addjob":
                ok, msg = add_manual_job(body)
                self._send(200 if ok else 400, json.dumps({"ok": ok, "message": msg}))
            elif self.path == "/api/editjob":
                ok, msg = edit_job(body.get("jobid", ""), body)
                self._send(200 if ok else 400, json.dumps({"ok": ok, "message": msg}))
            elif self.path == "/api/open":
                path = body.get("path", "")
                if not path:
                    self._send(400, json.dumps({"ok": False, "error": "path required"}))
                elif not RESUME_FOLDER or not path.startswith(RESUME_FOLDER):
                    self._send(403, json.dumps({"ok": False, "error": "path outside allowed folder"}))
                elif not os.path.isfile(path):
                    self._send(404, json.dumps({"ok": False, "error": "file not found"}))
                else:
                    subprocess.Popen(["open", path])
                    self._send(200, json.dumps({"ok": True}))
            elif self.path == "/api/exportjob":
                # Per-card single-job CSV export
                jid = str(body.get("jobid") or "").strip()
                if not jid:
                    self._send(400, json.dumps({"ok": False, "error": "jobid required"}))
                else:
                    # Find the job in the current job list
                    all_jobs = read_jobs_from_data()
                    job = next((j for j in all_jobs if j["jobid"] == jid), None)
                    if not job:
                        self._send(404, json.dumps({"ok": False, "error": "job not found"}))
                    else:
                        ok, result = export_single_job_csv(job)
                        self._send(200 if ok else 500, json.dumps({"ok": ok, "path": result if ok else "", "error": "" if ok else result}))
            elif self.path == "/api/savejobmd":
                jid = str(body.get("jobid") or "").strip()
                if not jid:
                    self._send(400, json.dumps({"ok": False, "error": "jobid required"}))
                else:
                    all_jobs = read_jobs_from_data()
                    job = next((j for j in all_jobs if j["jobid"] == jid), None)
                    if not job:
                        self._send(404, json.dumps({"ok": False, "error": "job not found"}))
                    else:
                        try:
                            def _s(v): return str(v) if v is not None else ""
                            company  = _s(job.get("company") or job.get("companyName", ""))
                            title    = _s(job.get("title", ""))
                            def _slug(s):
                                import re
                                return re.sub(r'[\\/:*?"<>|]', '', s).strip()
                            fname = f"{_slug(company)} - {_slug(title)}.md"
                            out_dir = _job_export_dir()
                            os.makedirs(out_dir, exist_ok=True)
                            out_path = os.path.join(out_dir, fname)
                            salary = _s(job.get("salaryRaw") or job.get("salary") or job.get("salaryNormAnnual") or "")
                            md = f"""# {title} — {company}

**Score:** {_s(job.get("score"))} | **Category:** {_s(job.get("category"))} | **Workplace:** {_s(job.get("workplaceNormalized") or job.get("workplace"))}
**Location:** {_s(job.get("location"))} | **Salary:** {salary}
**Geo Fit:** {_s(job.get("geoFit"))} | **CA Eligible:** {_s(job.get("caEligible"))}

## Score Reason
{_s(job.get("scoreReason") or job.get("why"))}

## Notes
{_s(job.get("manualNotes") or job.get("notes"))}

## Details
- **Job ID:** {_s(job.get("jobid") or job.get("id"))}
- **Posted:** {_s(job.get("postedAt"))}
- **Applied:** {_s(job.get("appliedAt"))}
- **Status:** {_s(job.get("reviewStatus") or job.get("status"))}
- **URL:** {_s(job.get("url") or job.get("link"))}

## Description
{_s(job.get("description"))}
"""
                            with open(out_path, "w", encoding="utf-8") as fh:
                                fh.write(md)
                            self._send(200, json.dumps({"ok": True, "path": out_path}))
                        except Exception as ex:
                            self._send(500, json.dumps({"ok": False, "error": str(ex)}))
            elif self.path == "/api/todo":
                path = str(body.get("path") or "").strip()
                folder = str(body.get("folder") or "").strip()
                company = str(body.get("company") or "").strip()
                role = str(body.get("role") or "").strip()
                content = body.get("content") or ""
                if not path:
                    if not folder and company and RESUME_FOLDER:
                        folder_name = f"{company} - {role}" if role else company
                        folder = os.path.join(RESUME_FOLDER, folder_name)
                    if folder:
                        os.makedirs(folder, exist_ok=True)
                        path = os.path.join(folder, "TODO.md")
                if not path or not RESUME_FOLDER or not os.path.abspath(path).startswith(os.path.abspath(RESUME_FOLDER)):
                    self._send(403, json.dumps({"ok": False, "error": "path not allowed"}))
                else:
                    try:
                        os.makedirs(os.path.dirname(path), exist_ok=True)
                        with open(path, "w", encoding="utf-8") as fh:
                            fh.write(content)
                        self._send(200, json.dumps({"ok": True, "path": path, "folder": os.path.dirname(path)}))
                    except Exception as e:
                        self._send(500, json.dumps({"ok": False, "error": str(e)}))
            elif self.path == "/api/restart":
                self._send(200, json.dumps({"ok": True}))
                import threading, time
                def _do_restart():
                    time.sleep(0.2)
                    os.execv(sys.executable, [sys.executable] + sys.argv + ["--no-open"])
                threading.Thread(target=_do_restart, daemon=True).start()
            else:
                self._send(404, json.dumps({"error": "not found"}))
        except Exception as e:
            self._send(500, json.dumps({"ok": False, "error": str(e)}))


def main():
    # The dashboard reads data.json (primary) and falls back to the legacy xlsx.
    # It only needs ONE of them; a brand-new install has neither until the first scrape.
    if not os.path.exists(DATA_JSON) and not os.path.exists(XLSX):
        sys.exit(
            "No jobs found yet.\n"
            "Run a scrape first — `/scrape` in Claude Code, or "
            "`bash scripts/run_pipeline.sh wide` (free, no accounts).\n"
            "The dashboard opens automatically once data.json exists."
        )
    source = DATA_JSON if os.path.exists(DATA_JSON) else XLSX
    url = f"http://127.0.0.1:{PORT}/"
    print("\n  Job Scout Browser")
    print("  " + "-" * 40)
    print(f"  Reading:  {source}")
    print(f"  Open:     {url}")
    print("  Stop:     press Ctrl+C in this window")
    print("  " + "-" * 40 + "\n")
    if "--no-open" not in sys.argv:
        try:
            webbrowser.open(url)
        except Exception:
            pass
    srv = ThreadingHTTPServer(("127.0.0.1", PORT), Handler)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\n  Stopped.\n")
        srv.shutdown()


if __name__ == "__main__":
    main()
