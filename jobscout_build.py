"""Job Scout — build the rolling data store and the spreadsheet.

Usage:
  python3 jobscout_build.py --new scored.json --data data.json --xlsx "/path/Job Scout Jobs.xlsx" [--applied applied.json] [--days 7]

- Reads prior data.json (rolling set) and the prior .xlsx (to capture your
  Dismissed/Notes edits, matched by JobID) so your edits survive regeneration.
- Classifies each job via jobscout_core (evidence-based workplace status, etc.).
- Marks Applied from applied.json (Gmail-derived).
- Merges, prunes to the last N days, sorts by Score then Posted.
- Writes data.json and a formatted, auto-filtered .xlsx.
"""
import argparse, json, os, datetime as dt
import jobscout_core as core

# ---------------------------------------------------------------------------
# Config — loaded once at startup; graceful fallback to empty dict.
# ---------------------------------------------------------------------------
_config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config", "jobscout_config.json")
try:
    with open(_config_path) as _f:
        _config = json.load(_f)
except (FileNotFoundError, json.JSONDecodeError):
    _config = {}

# The build runs in a UTC environment (scheduled-task sandbox), so never rely on
# the machine's local zone for display — format timestamps in the configured zone.
try:
    from zoneinfo import ZoneInfo
    LOCAL_TZ = ZoneInfo("America/Los_Angeles")
    LOCAL_TZ_LABEL = "Pacific time"
except Exception:  # pragma: no cover - fallback if tz database is missing
    LOCAL_TZ = dt.timezone(dt.timedelta(hours=-8))
    LOCAL_TZ_LABEL = "Pacific time (approx)"
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def compute_review_action(job: dict, state_entry: dict, config: dict) -> str:
    """Assign a single review action to a job based on its normalized state.

    Priority order (first matching rule wins):
    1. Already applied
    2. User dismissed
    3. CA restricted / excluded
    4. Below comp floor (with explicit salary)
    5. Apply — strong match, no flags
    6. Verify remote — Remote-unverified, high score
    7. Check salary — salary unknown or top-of-range-only
    8. Review manually — other risk flags present
    9. Dismiss — score too low
    """
    # Normalise the fields we need, handling both canonical and legacy key names.
    score = int(job.get("score") or 0)
    workplace = (job.get("workplaceNormalized") or job.get("workplace") or "").strip()
    geo_fit = job.get("geoFit")
    ca_excluded = bool(job.get("caExcluded"))
    ca_eligible = job.get("caEligible")   # True / False / None
    comp_risk = job.get("compRisk")        # "none" / "top-of-range-only" / "below-floor" / "unknown" / None
    salary_high = job.get("salaryHigh")    # numeric or None
    commute = job.get("commuteEstimateMinutes")   # numeric or None
    source_run = (job.get("sourceRun") or job.get("src") or "remote").lower()

    # Salary floor from config
    comp_cfg = config.get("compensation", {})
    floor = comp_cfg.get("baseSalaryFloorRemote", 120000) if source_run != "oc" \
        else comp_cfg.get("baseSalaryFloorOC", 120000)

    # Helper: salary is known AND below the applicable floor
    def _salary_known_below_floor():
        return isinstance(salary_high, (int, float)) and salary_high < floor

    # Helper: any non-trivial risk flag
    def _has_risk_flag():
        # Commute risk: commute > 40 min
        if isinstance(commute, (int, float)) and commute > 40:
            return True
        # CA uncertain (not definitively eligible, not excluded, non-remote)
        if not ca_excluded and ca_eligible is None and workplace not in ("Remote",):
            return True
        # Contract comp risk
        if comp_risk == "contract-verify":
            return True
        return False

    # 1. Already applied
    if job.get("applied") is True:
        return "Already applied"

    # 2. User dismissed
    if state_entry.get("userDismissed") is True:
        return "Dismissed"

    # 3. CA restricted / excluded
    if ca_excluded or ca_eligible is False:
        return "CA restricted"

    # 4. Below comp floor (salary is known and below floor)
    if _salary_known_below_floor():
        return "Below comp floor"

    # 5. Apply — strong match, no flags
    unverified = "unverified" in workplace.lower()
    if (score >= 8
            and geo_fit is True
            and ca_eligible is True
            and comp_risk in (None, "none")
            and (commute is None or commute <= 40)
            and not unverified):
        return "Apply"

    # 6–8. Remote-unverified branches (tiered by score)
    if unverified:
        if score >= 8:
            return "Verify remote"
        if score >= 5:
            return "Review manually"
        return "Dismiss"

    # 9. Check salary — score OK but salary unknown or just barely clearing the floor
    if score >= 7 and (not isinstance(salary_high, (int, float)) or comp_risk in ("unknown", "top-of-range-only")):
        return "Check salary"

    # 10. Review manually — score OK but some risk flag present
    if score >= 7 and _has_risk_flag():
        return "Review manually"

    # 11. Review manually — score is 7+ with no specific flag caught above
    if score >= 7:
        return "Review manually"

    # 12. Dismiss — low score
    return "Dismiss"


def compute_auto_suppressed(job: dict) -> bool:
    """Return True if a job should be hidden from the Review Queue sheet.

    Records are always written to data.json; suppression only affects the
    workbook Review Queue tab (or the Auto-suppressed column when the workbook
    is single-sheet).
    """
    workplace = (job.get("workplaceNormalized") or job.get("workplace") or "").strip()
    score = int(job.get("score") or 0)
    unverified = "unverified" in workplace.lower()

    if unverified and score <= 4:
        return True
    if score <= 2 and not job.get("starred") and not job.get("applied"):
        return True
    return False


COLS = ["JobID", "Score", "Title", "Company", "Workplace", "Location", "Salary", "Category",
        "Posted", "Days Posted", "Date Scraped", "Status", "Applied date", "Dismissed", "Notes",
        "Matched keywords", "Why", "Link", "Starred", "Geo fit", "Review action",
        "Auto-suppressed", "Description"]
STATUSES = ("new", "applied", "interviewing", "rejected")
WIDTHS = [12, 7, 40, 24, 18, 26, 14, 14, 11, 11, 12, 9, 12, 11, 28, 34, 52, 14, 8, 8, 16, 15, 70]
TRUTHY = {"x", "yes", "true", "1", "✓", "y"}


def load_json(path, default):
    if path and os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            return default
    return default


def read_prior_edits(xlsx_path):
    """Return {jobid: {'dismissed':bool,'notes':str}} from an existing spreadsheet."""
    edits = {}
    if not xlsx_path or not os.path.exists(xlsx_path):
        return edits
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h or "").strip() for h in next(rows)]
        idx = {name: header.index(name) for name in ("JobID", "Dismissed", "Notes") if name in header}
        if "JobID" not in idx:
            return edits
        for r in rows:
            jid = str(r[idx["JobID"]]).strip() if r[idx["JobID"]] is not None else ""
            if not jid:
                continue
            dval = str(r[idx["Dismissed"]] or "").strip().lower() if "Dismissed" in idx else ""
            notes = str(r[idx["Notes"]] or "").strip() if "Notes" in idx else ""
            edits[jid] = {"dismissed": dval in TRUTHY, "notes": notes}
        wb.close()
    except Exception as e:
        print("warn: could not read prior xlsx edits:", e)
    return edits


# Generic/legal tokens that must never be the basis of a company match.
_COMPANY_STOPWORDS = {
    "group", "groups", "inc", "incorporated", "llc", "llp", "pllc", "ltd", "limited",
    "corp", "corporation", "co", "company", "holdings", "global", "international", "intl",
    "technologies", "technology", "solutions", "services", "systems", "labs", "the", "and",
    "plc", "sa", "nv", "ag", "gmbh", "partners", "consulting", "ventures",
}


def _company_tokens(name):
    """Significant, lowercased company tokens (alphanumeric, generics + short tokens dropped)."""
    cleaned = "".join(ch if ch.isalnum() else " " for ch in (name or "").lower())
    return {t for t in cleaned.split() if len(t) > 2 and t not in _COMPANY_STOPWORDS}


def _company_matches(applied_company, job_company):
    """True only when the two names refer to the same company.

    Compares exact significant tokens (no substring matching) so e.g. 'Expedia Group'
    and 'The Voleon Group' (shared only the generic word 'group') do NOT match, and
    'Community Health Plan' does not match 'AmeriHealth' via the substring 'health'.
    A match requires set equality, or one side's significant tokens being a full
    subset of the other's (handles 'Agilent' vs 'Agilent Technologies').
    """
    a, b = _company_tokens(applied_company), _company_tokens(job_company)
    if not a or not b:
        return False
    return a == b or a <= b or b <= a


_ROLE_STOPWORDS = {
    "manager", "senior", "junior", "lead", "director", "specialist", "analyst",
    "program", "project", "principal", "staff", "associate", "assistant", "head",
    "vice", "president", "coordinator", "consultant", "advisor", "remote", "onsite",
    "hybrid", "the", "and", "for", "with", "of",
}


def _role_tokens(title):
    cleaned = "".join(ch if ch.isalnum() else " " for ch in (title or "").lower())
    return {t for t in cleaned.split() if len(t) > 2 and t not in _ROLE_STOPWORDS}


def _role_matches(applied_role, job_title):
    """True when the titles plausibly describe the same posting.

    Degrades to company-only matching when the applied entry has no usable role
    (so we never miss a genuine application that lacked a recorded title).
    """
    a, b = _role_tokens(applied_role), _role_tokens(job_title)
    if not a:
        return True
    if not b:
        return False
    inter = a & b
    return bool(inter) and len(inter) / min(len(a), len(b)) >= 0.5


def applied_lookup(job, applied_roles):
    """Flag a posting as applied only when BOTH the company and the role match.

    Matching the company alone over-suppresses other open roles at a company where
    you applied to a single posting; matching the role alone collides across
    companies. Requiring both keeps the flag meaning 'you applied to this posting'.
    """
    comp = job.get("companyName") or ""
    title = job.get("title") or ""
    for a in applied_roles:
        if _company_matches(a.get("company"), comp) and _role_matches(a.get("role"), title):
            return a.get("appliedAt") or "yes"
    return ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--new", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--applied", default="")
    ap.add_argument("--days", type=int, default=7)
    a = ap.parse_args()

    now = dt.datetime.now(dt.timezone.utc)
    now_iso = now.isoformat()

    new_raw = load_json(a.new, {})
    new_jobs = new_raw.get("jobs") if isinstance(new_raw, dict) else new_raw
    new_jobs = new_jobs or []

    prior = load_json(a.data, {})
    prior_recs = {str(r["id"]): r for r in (prior.get("jobs", []) if isinstance(prior, dict) else prior)}

    edits = read_prior_edits(a.xlsx)  # user's latest Dismissed/Notes from the open spreadsheet
    applied_roles = (load_json(a.applied, {}) or {}).get("appliedRoles", []) if a.applied else []

    # Load consolidated state (starred, userDismissed, reviewStatus, manualNotes, overrides)
    state_path = os.path.join(os.path.dirname(os.path.abspath(a.data)), "job-scout-state.json")
    scout_state = core.load_state(state_path)

    merged = dict(prior_recs)  # start from prior, then upsert new
    for j in new_jobs:
        jid = str(j.get("id"))
        if not jid or jid == "None":
            continue
        src = j.get("_src") or ("oc" if core.is_in_oc(j.get("location")) else "remote")
        d = core.classify(j, src)
        rec = merged.get(jid, {})
        added = rec.get("addedAt") or now_iso
        link = j.get("link") or (f"https://www.linkedin.com/jobs/view/{jid}")
        base_rec = {
            "id": jid, "title": j.get("title", ""), "company": j.get("companyName") or j.get("company") or "",
            "location": j.get("location", ""), "posted": j.get("postedAt") or j.get("posted") or "",
            "employmentType": j.get("employmentType", ""), "seniority": j.get("seniorityLevel", ""),
            "applicants": j.get("applicantsCount", ""),
            # Support both new LLM schema (score/scoreReason/searchLane/matchedKeywords)
            # and legacy schema (_score/_reason/keywords). New fields take priority.
            "score": j.get("score", j.get("_score", rec.get("score", 0))),
            "scoreReason": j.get("scoreReason", j.get("_reason", rec.get("scoreReason", rec.get("reason", "")))),
            "reason": j.get("scoreReason", j.get("_reason", rec.get("reason", ""))),  # legacy alias
            "searchLane": j.get("searchLane", rec.get("searchLane", "")),
            "matchedKeywords": j.get("matchedKeywords", d["keywords"]),
            "addedAt": added, "src": src, "link": link,
            "source": j.get("source") or (
                "manual" if src == "manual"
                else core.source_from_link(link) or "linkedin"
            ),
            "workplace": d["workplace"], "workplaceNormalized": d["workplace"],
            "salary": d["salary"], "caExcluded": d["caExcluded"],
            "keywords": j.get("matchedKeywords", d["keywords"]),  # legacy alias
            "category": d["category"], "ocTarget": d["ocTarget"],
            "geoFit": d["geoFit"],
            # retain source description so the parser can re-reference / re-derive
            # fields (salary, location, remote, keywords) without re-scraping.
            "description": core.get_description(j) or rec.get("description", ""),
            # preserve user edits (prefer the spreadsheet's current values)
            "dismissed": edits.get(jid, {}).get("dismissed", rec.get("dismissed", False)),
            "notes": edits.get(jid, {}).get("notes", rec.get("notes", "")),
        }
        # Overlay starred, userDismissed, reviewStatus, manualNotes, overrides from state
        base_rec = core.merge_state(base_rec, scout_state)
        # Apply _overrides: salaryRaw override -> salary display field
        state_overrides = base_rec.pop("_overrides", {})
        base_rec.pop("_stateUpdatedAt", None)  # internal field, not stored in records
        if state_overrides.get("salaryRaw"):
            base_rec["salary"] = state_overrides["salaryRaw"]
            base_rec["manualSalary"] = True
        # userDismissed from state propagates to dismissed (keep both for compatibility)
        if base_rec.get("userDismissed"):
            base_rec["dismissed"] = True
        # Ensure sourceRun is set before parse_salary (it drives the salary floor)
        if "sourceRun" not in base_rec:
            base_rec["sourceRun"] = base_rec.get("src") or base_rec.get("_src") or "remote"
        # Parse salary into structured fields
        salary_fields = core.parse_salary(base_rec)
        base_rec.update(salary_fields)
        # Parse geo/CA eligibility
        geo_fields = core.parse_geo_eligibility(base_rec)
        base_rec.update(geo_fields)
        # Commute estimate for non-remote jobs
        if base_rec.get("workplace", "").lower() not in ("remote",) or base_rec.get("workplaceNormalized", "").lower() not in ("remote",):
            loc = base_rec.get("location") or base_rec.get("city") or ""
            commute = core.get_commute_estimate(loc)
            base_rec["commuteEstimateMinutes"] = commute.get("estimatedMinutes")
            base_rec["commuteUnknown"] = commute.get("commuteUnknown", True)
        merged[jid] = base_rec

    # also fold edits into prior records that weren't re-scraped this run
    for jid, e in edits.items():
        if jid in merged:
            merged[jid]["dismissed"] = e["dismissed"]
            merged[jid]["notes"] = e["notes"]

    # Apply state to prior records that weren't re-scraped this run
    for jid, rec in merged.items():
        if jid not in {str(j.get("id")) for j in new_jobs}:
            rec_with_state = core.merge_state(rec, scout_state)
            state_overrides = rec_with_state.pop("_overrides", {})
            rec_with_state.pop("_stateUpdatedAt", None)
            if state_overrides.get("salaryRaw"):
                rec_with_state["salary"] = state_overrides["salaryRaw"]
                rec_with_state["manualSalary"] = True
            if rec_with_state.get("userDismissed"):
                rec_with_state["dismissed"] = True
            # Ensure sourceRun is set before parse_salary (it drives the salary floor)
            if "sourceRun" not in rec_with_state:
                rec_with_state["sourceRun"] = rec_with_state.get("src") or rec_with_state.get("_src") or "remote"
            # Backfill source for old records that pre-date the field.
            if not rec_with_state.get("source"):
                if (rec_with_state.get("src") or rec_with_state.get("sourceRun") or "").lower() == "manual":
                    rec_with_state["source"] = "manual"
                else:
                    link = rec_with_state.get("link") or ""
                    rec_with_state["source"] = core.source_from_link(link) or "linkedin"
            # Backfill workplaceNormalized for old records that pre-date the field.
            # parse_geo_eligibility reads workplaceNormalized; without it geo parsing
            # degrades to "workplace type unknown" and caEligible stays None.
            if not rec_with_state.get("workplaceNormalized"):
                rec_with_state["workplaceNormalized"] = rec_with_state.get("workplace", "")
            # Reparse salary and geo so rebuilds propagate new parser logic to all records
            salary_fields = core.parse_salary(rec_with_state)
            rec_with_state.update(salary_fields)
            geo_fields = core.parse_geo_eligibility(rec_with_state)
            rec_with_state.update(geo_fields)
            # Commute estimate for non-remote jobs
            if rec_with_state.get("workplace", "").lower() not in ("remote",) or rec_with_state.get("workplaceNormalized", "").lower() not in ("remote",):
                loc = rec_with_state.get("location") or rec_with_state.get("city") or ""
                commute = core.get_commute_estimate(loc)
                rec_with_state["commuteEstimateMinutes"] = commute.get("estimatedMinutes")
                rec_with_state["commuteUnknown"] = commute.get("commuteUnknown", True)
            merged[jid] = rec_with_state

    # Manual overrides (salary/workplace typed in the dashboard) win over derived
    # values and survive every rebuild. File maps jobid -> {salary?, workplace?}.
    ov_path = os.path.join(os.path.dirname(os.path.abspath(a.data)), "job-scout-overrides.json")
    overrides = load_json(ov_path, {}) or {}
    for jid, ov in overrides.items():
        rec = merged.get(str(jid))
        if not rec:
            continue
        if ov.get("salary") is not None:
            rec["salary"] = ov["salary"]
            rec["manualSalary"] = True
        if ov.get("workplace"):
            rec["workplace"] = ov["workplace"]
            rec["manualWorkplace"] = True
        if ov.get("score") is not None:
            try:
                rec["score"] = max(0, min(10, int(ov["score"])))
                rec["manualScore"] = True
            except Exception:
                pass

    # Starred follow-up flags (set in the dashboard) — kept in a file, surfaced as a column.
    stars_path = os.path.join(os.path.dirname(os.path.abspath(a.data)), "job-scout-stars.json")
    stars = set(str(x) for x in (load_json(stars_path, []) or []))

    _te_path = os.path.join(os.path.dirname(os.path.abspath(a.data)), "config", "target_employers.json")
    _target_employers = (load_json(_te_path, {}) or {}).get("employers", [])

    # Recompute geoFit + approximate keyword provenance + starred for the WHOLE set each
    # run so rule changes (and any workplace overrides) propagate to every rolling record.
    for rec in merged.values():
        # Only fall back to the blunt legacy geo_fit() when parse_geo_eligibility()
        # left geoFit as None (e.g. old records without workplaceNormalized).
        # Never overwrite a value already determined by the structured parser.
        if rec.get("geoFit") is None:
            rec["geoFit"] = core.geo_fit(
                rec.get("workplaceNormalized") or rec.get("workplace", ""),
                rec.get("location", ""),
            )
        rec["searchKeywords"] = core.search_keyword_matches(rec)
        rec["starred"] = str(rec.get("id")) in stars
        rec["targetEmployer"] = False
        rec["targetEmployerName"] = ""
        rec_co = rec.get("company") or ""
        for te in _target_employers:
            if _company_matches(te.get("name", ""), rec_co):
                rec["targetEmployer"] = True
                rec["targetEmployerName"] = te["name"]
                break

    # Application status — manual statuses (set in the dashboard) are the source of
    # truth and survive rebuilds; a genuine Gmail confirmation only sets the DEFAULT
    # ("applied") for jobs you haven't set yourself. Stale dates can't create false
    # positives because nothing is sticky.
    status_path = os.path.join(os.path.dirname(os.path.abspath(a.data)), "job-scout-status.json")
    status_store = load_json(status_path, {}) or {}
    for rec in merged.values():
        ad = applied_lookup({"companyName": rec["company"], "title": rec["title"]}, applied_roles)
        manual = str(status_store.get(str(rec.get("id")), "")).strip().lower()
        if manual in STATUSES:
            rec["status"] = manual
        else:
            rec["status"] = "applied" if ad else "new"
        rec["appliedDate"] = ad if ad and ad != "yes" else (rec.get("appliedDate", "") if rec["status"] != "new" else "")
        rec["applied"] = rec["status"] != "new"

    # Compute reviewAction, autoSuppressed, and salaryNormAnnual for every record
    # now that all fields (applied, geoFit, starred, salary, etc.) have been finalised.
    for rec in merged.values():
        jid = str(rec.get("id", ""))
        state_entry = scout_state.get("jobs", {}).get(jid, {})
        rec["reviewAction"] = compute_review_action(rec, state_entry, _config)
        rec["autoSuppressed"] = compute_auto_suppressed(rec)
        if not rec.get("incentives"):
            rec["incentives"] = core.detect_incentives(rec.get("description") or "")
        # Normalised annual salary — used by the dashboard salary range filter.
        # Hourly roles use annualizedHigh (rate × 2080); annual roles use salaryHigh.
        sal_type = rec.get("salaryType") or "unknown"
        if sal_type == "hourly":
            rec["salaryNormAnnual"] = int(rec["annualizedHigh"]) if rec.get("annualizedHigh") else 0
        elif sal_type == "annual":
            rec["salaryNormAnnual"] = int(rec["salaryHigh"]) if rec.get("salaryHigh") else 0
        else:
            rec["salaryNormAnnual"] = 0

    # prune by addedAt window
    cutoff = now - dt.timedelta(days=a.days)
    def fresh(rec):
        try:
            return dt.datetime.fromisoformat(rec["addedAt"].replace("Z", "+00:00")) >= cutoff
        except Exception:
            return True
    def keep(rec):
        if fresh(rec):
            return True
        if rec.get("manual") or rec.get("applied"):
            return True
        # Never silently drop starred jobs or jobs the user is actively progressing
        if rec.get("starred"):
            return True
        status = rec.get("reviewStatus") or rec.get("status") or "new"
        if status not in ("new", "dismissed"):
            return True
        return False
    recs = [r for r in merged.values() if keep(r)]

    # Collapse duplicate (title, company) postings — reposts and location variants
    # get distinct JobIDs, so dedup by exact title + fuzzy company (token overlap).
    # "BeOne" and "BeOne Medicines" share the token "beone" → same company.
    import re as _re
    from collections import defaultdict as _defaultdict
    def _title_norm(r):
        return _re.sub(r"\s+", " ", (r.get("title") or "").strip().lower())
    def _ddrank(r):
        return (0 if r.get("dismissed") else 1, 1 if r.get("geoFit") else 0,
                r.get("score") or 0, 1 if (r.get("salary") or "").strip() else 0,
                len(r.get("description") or ""), str(r.get("addedAt") or ""))

    # Group by exact normalized title, then within each group union-find on company
    # token overlap so name variants (with/without suffixes) collapse to one record.
    _title_groups = _defaultdict(list)
    for r in recs:
        _title_groups[_title_norm(r)].append(r)

    _kept = []
    for _title, _group in _title_groups.items():
        if len(_group) == 1:
            _kept.append(_group[0])
            continue
        # Union-find: merge records whose company token sets share any token.
        _parent = list(range(len(_group)))
        def _find(i):
            while _parent[i] != i:
                _parent[i] = _parent[_parent[i]]
                i = _parent[i]
            return i
        _co_tok = [_company_tokens(r.get("company") or r.get("companyName") or "") for r in _group]
        for _i in range(len(_group)):
            for _j in range(_i + 1, len(_group)):
                if _co_tok[_i] and _co_tok[_j] and (_co_tok[_i] & _co_tok[_j]):
                    _pi, _pj = _find(_i), _find(_j)
                    if _pi != _pj:
                        _parent[_pi] = _pj
        _clusters = _defaultdict(list)
        for _i, r in enumerate(_group):
            _clusters[_find(_i)].append(r)
        for _cluster in _clusters.values():
            best = max(_cluster, key=_ddrank)
            if len(_cluster) > 1:
                dupes = [r for r in _cluster if r is not best]
                for d in dupes:
                    print(f"[title-dedup] {d.get('title')!r} @ {d.get('company')!r} (id={d.get('id')}) "
                          f"→ kept {best.get('company')!r} (id={best.get('id')})")
            _kept.append(best)
    recs = _kept

    # Description fingerprint dedup — catches cross-posts where different staffing
    # agencies post the same underlying job with different company names / IDs.
    # Fingerprints chars 150-450 of normalized description to skip agency intro text.
    import hashlib as _hashlib
    def _desc_fp(r):
        raw = (r.get("description") or "")
        # strip HTML entities and collapse whitespace
        cleaned = _re.sub(r"&\w+;", " ", raw.lower())
        cleaned = _re.sub(r"\s+", " ", cleaned).strip()
        body = cleaned[150:450]
        if len(body) < 80:
            body = cleaned[:300]  # fallback for very short descriptions
        if len(body) < 40:
            return None  # not enough content to fingerprint
        return _hashlib.md5(body.encode()).hexdigest()

    _fp_best: dict = {}
    for r in recs:
        fp = _desc_fp(r)
        if fp is None:
            continue
        if fp not in _fp_best or _ddrank(r) > _ddrank(_fp_best[fp]):
            _fp_best[fp] = r

    _fp_to_canonical: dict = {fp: str(rec["id"]) for fp, rec in _fp_best.items()}

    deduped = []
    for r in recs:
        fp = _desc_fp(r)
        canonical_id = _fp_to_canonical.get(fp) if fp else None
        if canonical_id and str(r.get("id")) != canonical_id:
            r["dupeOf"] = canonical_id
            print(f"[desc-dedup] {r.get('title')!r} @ {r.get('company')!r} → dupeOf {canonical_id}")
        deduped.append(r)
    recs = deduped
    # Geographic fit first (verified-remote anywhere, or SoCal-local), then score,
    # then most-recently posted — so out-of-area roles sink to the bottom.
    recs.sort(key=lambda r: (1 if r.get("geoFit") else 0, r.get("score") or 0, str(r.get("posted") or "")), reverse=True)

    payload = {"jobs": recs, "updatedAt": now_iso, "count": len(recs)}
    with open(a.data, "w") as f:
        json.dump(payload, f, indent=1)

    # Version history via append-only dated snapshots (this folder blocks file
    # deletion, so git can't be used; snapshots never need deleting).
    snap_dir = os.path.join(os.path.dirname(os.path.abspath(a.data)), "snapshots")
    os.makedirs(snap_dir, exist_ok=True)
    with open(os.path.join(snap_dir, f"data-{now.strftime('%Y-%m-%d')}.json"), "w") as f:
        json.dump(payload, f)

    write_xlsx(a.xlsx, recs, now, applied_roles=applied_roles)
    n_oc = sum(1 for r in recs if r["ocTarget"])
    n_unv = sum(1 for r in recs if r["workplace"] == "Remote — unverified")
    print(f"build: {len(new_jobs)} scraped -> {len(recs)} in rolling set | OC {n_oc} | remote-unverified {n_unv}")


def _salary_display(rec):
    """Format salary as '$143K–$190K', '~$143K', raw salary field, or 'Unknown'."""
    low = rec.get("salaryLow")
    high = rec.get("salaryHigh")
    if isinstance(low, (int, float)) and isinstance(high, (int, float)):
        return f"${int(low / 1000)}K–${int(high / 1000)}K"
    if isinstance(high, (int, float)):
        return f"~${int(high / 1000)}K"
    if isinstance(low, (int, float)):
        return f"~${int(low / 1000)}K"
    raw = str(rec.get("salary") or "").strip()
    return raw if raw else "Unknown"


def _fmt_date_scraped(added_at: str) -> str:
    """Return the date portion of an addedAt ISO timestamp, e.g. '2026-06-13'."""
    if not added_at:
        return ""
    try:
        return str(added_at)[:10]
    except Exception:
        return ""


def _days_posted(posted: str) -> str:
    """Return how many days ago the job was posted, e.g. '3d ago'."""
    if not posted:
        return ""
    try:
        post_date = dt.date.fromisoformat(str(posted)[:10])
        today = dt.datetime.now(dt.timezone.utc).date()
        delta = (today - post_date).days
        if delta < 0:
            return "0d ago"
        if delta == 0:
            return "Today"
        return f"{delta}d ago"
    except Exception:
        return ""


def _set_header_row(ws, cols, header_fill, header_font):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _write_review_queue(wb, recs, now):
    """Sheet 1: Review Queue — active jobs only, sorted and colour-coded."""
    ws = wb.create_sheet("Review Queue")

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    RQ_COLS = [
        "Score", "Review Action", "★", "Status", "Title", "Company", "Location",
        "Workplace", "Salary", "Comp Risk", "Geo Fit", "CA Eligible",
        "Search Lane", "Reason", "URL", "Date Scraped", "Days Posted", "Notes",
    ]
    RQ_WIDTHS = [7, 16, 4, 12, 32, 22, 20, 14, 16, 14, 8, 10, 24, 40, 30, 12, 11, 28]

    _set_header_row(ws, RQ_COLS, header_fill, header_font)

    # Review Action fill colours
    RA_FILLS = {
        "Apply":             PatternFill("solid", start_color="C6EFCE"),
        "Verify remote":     PatternFill("solid", start_color="FFEB9C"),
        "Check salary":      PatternFill("solid", start_color="FFEB9C"),
        "Review manually":   PatternFill("solid", start_color="DDEBF7"),
        "Dismissed":         PatternFill("solid", start_color="FFC7CE"),
        "Below comp floor":  PatternFill("solid", start_color="FFC7CE"),
        "CA restricted":     PatternFill("solid", start_color="FFC7CE"),
        "Already applied":   PatternFill("solid", start_color="FFC7CE"),
        "Dismiss":           PatternFill("solid", start_color="D9D9D9"),
    }

    # Filter: exclude auto-suppressed and dismissed
    queue = [r for r in recs if not r.get("autoSuppressed") and not r.get("userDismissed")]
    # Sort: score descending, then reviewAction alphabetically
    _RA_ORDER = {
        "Apply": 0, "Verify remote": 1, "Check salary": 2,
        "Review manually": 3, "Dismiss": 4,
        "Dismissed": 5, "Below comp floor": 5, "CA restricted": 5, "Already applied": 5,
    }
    queue.sort(key=lambda r: (-(r.get("score") or 0), _RA_ORDER.get(r.get("reviewAction", ""), 9)))

    _WRAP_IDXS = {RQ_COLS.index(h) + 1 for h in ("Title", "Reason", "Notes")}

    for r in queue:
        ra = r.get("reviewAction", "")
        geo = r.get("geoFit")
        ca = r.get("caEligible")
        url = r.get("url") or r.get("linkedInUrl") or r.get("link") or ""

        ws.append([
            r.get("score") or r.get("_score") or 0,
            ra,
            "x" if r.get("starred") else "",
            r.get("reviewStatus") or r.get("status") or "new",
            r.get("title", ""),
            r.get("companyName") or r.get("company", ""),
            r.get("location", ""),
            r.get("workplaceNormalized") or r.get("workplace", ""),
            _salary_display(r),
            r.get("compRisk") or "",
            "True" if geo is True else ("False" if geo is False else "?"),
            "True" if ca is True else ("False" if ca is False else "?"),
            r.get("searchLane") or "",
            r.get("scoreReason") or r.get("reason", ""),
            url,
            _fmt_date_scraped(r.get("addedAt")),
            _days_posted(r.get("posted") or r.get("postedAt")),
            r.get("manualNotes") or r.get("notes", ""),
        ])

        row = ws.max_row
        # Row height
        ws.row_dimensions[row].height = 18

        # Review Action fill
        if ra in RA_FILLS:
            ws.cell(row=row, column=2).fill = RA_FILLS[ra]

        # URL as hyperlink
        if url:
            link_cell = ws.cell(row=row, column=15)
            link_cell.hyperlink = url
            link_cell.value = "Open ↗"
            link_cell.font = Font(color="0563C1", underline="single", name="Arial")

        # Set font and alignment for all cells in row
        for c in range(1, len(RQ_COLS) + 1):
            cc = ws.cell(row=row, column=c)
            if not (cc.font and cc.font.color and cc.font.color.value == "0563C1"):
                cc.font = Font(name="Arial")
            cc.alignment = Alignment(vertical="top", wrap_text=(c in _WRAP_IDXS))

    # Header row height
    ws.row_dimensions[1].height = 18

    # Column widths
    for i, w in enumerate(RQ_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze row 1 + first 4 columns (E2 = first unfrozen cell)
    ws.freeze_panes = "E2"

    # AutoFilter
    last_col = get_column_letter(len(RQ_COLS))
    last_row = max(ws.max_row, 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    return ws


def _write_jobs(wb, recs):
    """Sheet 2: Jobs — full list including dismissed and auto-suppressed."""
    ws = wb.create_sheet("Jobs")

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    fills = {"Remote": "DCFCE7", "Remote — unverified": "FEF3C7", "Hybrid": "FFE4CC", "On-site": "FFE4CC"}

    _set_header_row(ws, COLS, header_fill, header_font)

    _WRAP_COLS = {COLS.index(h) + 1 for h in ("Title", "Notes", "Matched keywords", "Why")}

    for r in recs:
        ws.append([
            r.get("id"), r.get("score", 0), r.get("title", ""),
            r.get("companyName") or r.get("company", ""),
            r.get("workplaceNormalized") or r.get("workplace", ""),
            r.get("location", ""), r.get("salary", ""), r.get("category", ""),
            r.get("posted") or r.get("postedAt", ""),
            _days_posted(r.get("posted") or r.get("postedAt")),
            _fmt_date_scraped(r.get("addedAt")),
            r.get("status", "new"),
            r.get("appliedDate", "") if r.get("appliedDate") != "yes" else "",
            "x" if r.get("dismissed") else "", r.get("notes") or r.get("manualNotes", ""),
            ", ".join(r.get("keywords") or r.get("matchedKeywords") or []),
            r.get("reason") or r.get("scoreReason", ""), r.get("link") or r.get("linkedInUrl", ""),
            "x" if r.get("starred") else "",
            "x" if r.get("geoFit") else "",
            r.get("reviewAction", ""),
            "x" if r.get("autoSuppressed") else "",
            r.get("description", ""),
        ])
        row = ws.max_row
        wp = r.get("workplaceNormalized") or r.get("workplace", "")
        if wp in fills:
            ws.cell(row=row, column=5).fill = PatternFill("solid", start_color=fills[wp])
        ra = r.get("reviewAction", "")
        ra_col = COLS.index("Review action") + 1
        if ra == "Apply":
            ws.cell(row=row, column=ra_col).fill = PatternFill("solid", start_color="DCFCE7")
        elif ra in ("Verify remote", "Check salary"):
            ws.cell(row=row, column=ra_col).fill = PatternFill("solid", start_color="FEF3C7")
        elif ra == "Dismiss":
            ws.cell(row=row, column=ra_col).fill = PatternFill("solid", start_color="FEE2E2")
        link_val = r.get("link") or r.get("linkedInUrl", "")
        link_cell = ws.cell(row=row, column=COLS.index("Link") + 1)
        if link_val:
            link_cell.hyperlink = link_val
            link_cell.value = "Open ↗"
            link_cell.font = Font(color="0563C1", underline="single", name="Arial")
        for c in range(1, len(COLS) + 1):
            cc = ws.cell(row=row, column=c)
            if not (cc.font and cc.font.color and cc.font.color.value == "0563C1"):
                cc.font = Font(name="Arial")
            cc.alignment = Alignment(vertical="top", wrap_text=(c in _WRAP_COLS))

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    last = f"{get_column_letter(len(COLS))}{max(ws.max_row, 1)}"
    ws.auto_filter.ref = f"A1:{last}"

    dv = DataValidation(type="list", formula1='"x"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"L2:L{max(2, ws.max_row)}")

    return ws


def _write_descriptions(wb, recs):
    """Sheet 3: Descriptions — only jobs with non-empty description text."""
    ws = wb.create_sheet("Descriptions")

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    DESC_COLS = ["Job ID", "Title", "Company", "Description", "Fetch Status", "Source Field"]
    DESC_WIDTHS = [14, 36, 24, 80, 18, 18]

    _set_header_row(ws, DESC_COLS, header_fill, header_font)

    desc_col_idx = DESC_COLS.index("Description") + 1

    for r in recs:
        desc = r.get("description") or ""
        if not desc:
            continue
        ws.append([
            r.get("id"),
            r.get("title", ""),
            r.get("companyName") or r.get("company", ""),
            desc,
            r.get("descriptionFetchStatus") or "",
            r.get("descriptionSourceField") or "",
        ])
        row = ws.max_row
        for c in range(1, len(DESC_COLS) + 1):
            cc = ws.cell(row=row, column=c)
            cc.font = Font(name="Arial")
            cc.alignment = Alignment(
                vertical="top",
                wrap_text=(c == desc_col_idx),
            )

    for i, w in enumerate(DESC_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    return ws


def _write_applied(wb, applied_roles):
    """Sheet 4: Applied — entries from job-scout-applied.json."""
    ws = wb.create_sheet("Applied")

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    APP_COLS = ["Company", "Role", "Applied At", "Source", "Confidence"]
    APP_WIDTHS = [28, 40, 22, 20, 14]

    _set_header_row(ws, APP_COLS, header_fill, header_font)

    for a in applied_roles:
        ws.append([
            a.get("company", ""),
            a.get("role", ""),
            a.get("appliedAt", ""),
            a.get("source", ""),
            a.get("confidence", ""),
        ])
        row = ws.max_row
        for c in range(1, len(APP_COLS) + 1):
            ws.cell(row=row, column=c).font = Font(name="Arial")
            ws.cell(row=row, column=c).alignment = Alignment(vertical="top")

    for i, w in enumerate(APP_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    return ws


def _write_dashboard(wb, recs, applied_roles, now):
    """Sheet 5: Dashboard — summary key/value table."""
    ws = wb.create_sheet("Dashboard")

    bold_font = Font(name="Arial", bold=True)
    val_font = Font(name="Arial")

    def _kv(key, val):
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=key).font = bold_font
        ws.cell(row=row_num, column=2, value=val).font = val_font

    total = len(recs)
    score7 = sum(1 for r in recs if (r.get("score") or 0) >= 7)
    score8 = sum(1 for r in recs if (r.get("score") or 0) >= 8)
    not_dismissed_score7 = sum(1 for r in recs if not r.get("userDismissed") and (r.get("score") or 0) >= 7)
    starred = sum(1 for r in recs if r.get("starred"))
    applied_count = sum(1 for r in recs if r.get("applied"))
    dismissed_count = sum(1 for r in recs if r.get("userDismissed"))
    auto_sup = sum(1 for r in recs if r.get("autoSuppressed"))
    remote_count = sum(1 for r in recs if "Remote" in (r.get("workplaceNormalized") or r.get("workplace", "")))
    hybrid_count = sum(1 for r in recs if "Hybrid" in (r.get("workplaceNormalized") or r.get("workplace", "")))
    onsite_count = sum(1 for r in recs if "On-site" in (r.get("workplaceNormalized") or r.get("workplace", "")))
    blank_salary = sum(1 for r in recs if r.get("salaryHigh") is None and not str(r.get("salary") or "").strip())
    blank_desc = sum(1 for r in recs if not (r.get("description") or "").strip())
    alert_apply = sum(1 for r in recs if r.get("reviewAction") == "Apply")
    alert_review = sum(1 for r in recs if "Review" in (r.get("reviewAction") or ""))
    alert_verify = sum(1 for r in recs if r.get("reviewAction") == "Verify remote")
    built_at = now.isoformat()

    _kv("Total jobs", total)
    _kv("Score ≥ 7", score7)
    _kv("Score ≥ 8", score8)
    _kv("Not dismissed, score ≥ 7", not_dismissed_score7)
    _kv("Starred", starred)
    _kv("Applied", applied_count)
    _kv("Dismissed", dismissed_count)
    _kv("Auto-suppressed", auto_sup)
    _kv("Remote", remote_count)
    _kv("Hybrid", hybrid_count)
    _kv("On-site", onsite_count)
    _kv("Blank salary", blank_salary)
    _kv("Blank description", blank_desc)
    _kv("Alerts — Apply", alert_apply)
    _kv("Alerts — Review", alert_review)
    _kv("Alerts — Verify remote", alert_verify)
    _kv("Built at", built_at)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    return ws


def _write_config_snapshot(wb, config, now):
    """Sheet 6: Config Snapshot — key/value table of current config."""
    ws = wb.create_sheet("Config Snapshot")

    bold_font = Font(name="Arial", bold=True)
    val_font = Font(name="Arial")

    def _kv(key, val):
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=key).font = bold_font
        ws.cell(row=row_num, column=2, value=val).font = val_font

    comp = config.get("compensation", {})
    geo = config.get("geo", {})
    search = config.get("search", {})
    seen = config.get("seenIds", {})

    remote_kw = search.get("remote", {}).get("keywords", "")
    oc_kw = search.get("orangeCounty", {}).get("keywords", "")

    _kv("Remote salary floor", comp.get("baseSalaryFloorRemote", ""))
    _kv("OC salary floor", comp.get("baseSalaryFloorOC", ""))
    _kv("Target total comp", comp.get("targetTotalComp", ""))
    _kv("Contract adjustment factor", comp.get("contractAdjustmentFactor", ""))
    _kv("Commute max minutes", geo.get("commuteMaxMinutes", ""))
    _kv("Remote keywords", remote_kw[:120] if len(remote_kw) > 120 else remote_kw)
    _kv("OC keywords", oc_kw[:120] if len(oc_kw) > 120 else oc_kw)
    _kv("Seen IDs max", seen.get("maxIds", ""))
    _kv("Built at", now.isoformat())

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    return ws


def write_xlsx(path, recs, now, applied_roles=None):
    wb = Workbook()
    # Remove the default sheet created by Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    _write_review_queue(wb, recs, now)
    _write_jobs(wb, recs)
    _write_descriptions(wb, recs)
    _write_applied(wb, applied_roles or [])
    _write_dashboard(wb, recs, applied_roles or [], now)
    _write_config_snapshot(wb, _config, now)

    # Set Review Queue as the active (leftmost) tab
    wb.active = wb["Review Queue"]

    wb.save(path)


if __name__ == "__main__":
    main()
